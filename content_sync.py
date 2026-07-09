"""
content_sync.py — ~/sima_bot/content_sync.py

Сервис синхронизации контента Ozon ↔ Wildberries.
Сопоставление по offer_id (Ozon) == vendorCode (WB).
"""
from __future__ import annotations

import asyncio
from dataclasses import dataclass, field
from typing import Optional

import aiohttp
from sqlalchemy import select, update

from database import AsyncSessionLocal, WbAccount


# ─────────────────────────────────────────────────
# Структуры данных
# ─────────────────────────────────────────────────

@dataclass
class ProductContent:
    vendor_code: str
    name: str = ""
    description: str = ""
    images: list[str] = field(default_factory=list)
    attributes: list[dict] = field(default_factory=list)
    product_id: Optional[int] = None
    ozon_attributes_raw: list[dict] = field(default_factory=list)
    nm_id: Optional[int] = None


@dataclass
class MatchedProduct:
    vendor_code: str
    ozon: Optional[ProductContent]
    wb:   Optional[ProductContent]


# ─────────────────────────────────────────────────
# Получение всех товаров Ozon (Content API)
# ─────────────────────────────────────────────────

async def fetch_ozon_products(headers: dict) -> dict[str, ProductContent]:
    result: dict[str, ProductContent] = {}

    async with aiohttp.ClientSession() as session:
        product_ids: list[int] = []
        offer_map:   dict[int, str] = {}
        last_id = ""

        while True:
            async with session.post(
                "https://api-seller.ozon.ru/v3/product/list",
                headers=headers,
                json={"filter": {}, "last_id": last_id, "limit": 1000},
            ) as resp:
                data = await resp.json()
            items = data.get("result", {}).get("items", [])
            if not items:
                break
            for item in items:
                product_ids.append(item["product_id"])
                offer_map[item["product_id"]] = item["offer_id"]
            last_id = data.get("result", {}).get("last_id", "")
            if not last_id:
                break

        if not product_ids:
            return result

        for i in range(0, len(product_ids), 100):
            batch = product_ids[i:i + 100]
            async with session.post(
                "https://api-seller.ozon.ru/v3/product/info/list",
                headers=headers,
                json={"product_id": batch},
            ) as resp:
                info_data = await resp.json(content_type=None)
            for p in info_data.get("items", []):
                pid      = p["id"]
                offer_id = offer_map.get(pid, str(pid))
                attrs = [
                    {
                        "name":  a.get("name", ""),
                        "value": " / ".join(v.get("value", "") for v in a.get("values", [])),
                    }
                    for a in p.get("attributes", []) if a.get("name")
                ]
                result[offer_id] = ProductContent(
                    vendor_code         = offer_id,
                    product_id          = pid,
                    name                = p.get("name", ""),
                    description         = p.get("description", ""),
                    images              = p.get("images", []),
                    attributes          = attrs,
                    ozon_attributes_raw = p.get("attributes", []),
                )

    return result


# ─────────────────────────────────────────────────
# Получение всех карточек WB (Content API v2)
# ─────────────────────────────────────────────────

async def fetch_wb_products(wb_api_key: str) -> dict[str, ProductContent]:
    headers = {"Authorization": wb_api_key, "Content-Type": "application/json"}
    result: dict[str, ProductContent] = {}
    cursor: dict = {}
    LIMIT = 100

    async with aiohttp.ClientSession() as session:
        while True:
            async with session.post(
                "https://content-api.wildberries.ru/content/v2/get/cards/list",
                headers=headers,
                json={"settings": {
                    "sort":   {"ascending": False},
                    "cursor": {**cursor, "limit": LIMIT},
                    "filter": {"withPhoto": -1},
                }},
            ) as resp:
                data = await resp.json(content_type=None)

            cards = data.get("cards", [])
            if not cards:
                break

            for card in cards:
                nm_id       = card.get("nmID")
                vendor_code = card.get("vendorCode") or str(nm_id)

                photos = card.get("photos", [])
                images = [p["big"] for p in photos if p.get("big")][:10]

                attrs = [
                    {
                        "name":  ch.get("name", ""),
                        "value": " / ".join(
                            str(v) for v in (
                                ch["value"] if isinstance(ch.get("value"), list)
                                else [ch["value"]] if ch.get("value") is not None
                                else []
                            )
                        ),
                    }
                    for ch in card.get("characteristics", []) if ch.get("name")
                ]
                result[vendor_code] = ProductContent(
                    vendor_code = vendor_code,
                    nm_id       = nm_id,
                    name        = card.get("title", ""),
                    description = card.get("description", ""),
                    images      = images,
                    attributes  = attrs,
                )

            # Правильное условие из документации WB: total < limit = последняя страница
            cur = data.get("cursor", {})
            if cur.get("total", 0) < LIMIT:
                break
            cursor = {"updatedAt": cur["updatedAt"], "nmID": cur["nmID"]}

    return result


# ─────────────────────────────────────────────────
# Сопоставление по артикулу продавца
# ─────────────────────────────────────────────────

async def get_matched_products(ozon_headers: dict, wb_api_key: str) -> list[MatchedProduct]:
    ozon, wb = await asyncio.gather(
        fetch_ozon_products(ozon_headers),
        fetch_wb_products(wb_api_key),
    )
    codes = sorted(set(ozon) & set(wb))
    return [MatchedProduct(vendor_code=c, ozon=ozon[c], wb=wb[c]) for c in codes]


# ─────────────────────────────────────────────────
# Применение контента WB → Ozon
# ─────────────────────────────────────────────────

async def apply_wb_to_ozon(
    ozon_headers: dict,
    wb_api_key:   str,
    tasks:        list[dict],
) -> list[dict]:
    """Применяет контент WB → Ozon только для товаров из tasks."""

    needed_codes = {t["vendor_code"] for t in tasks}

    # 1. WB данные берём из запроса (уже показаны пользователю — стабильно)
    from database import Product as ProductModel
    wb: dict[str, ProductContent] = {}
    for task in tasks:
        if task.get("wb_data"):
            d = task["wb_data"]
            wb[task["vendor_code"]] = ProductContent(
                vendor_code = task["vendor_code"],
                name        = d.get("name", ""),
                description = d.get("description", ""),
                images      = d.get("images", []),
                attributes  = d.get("attributes", []),
                nm_id       = d.get("nm_id"),
            )

    # 2. product_id из локальной БД
    ozon_ids: dict[str, int] = {}
    async with AsyncSessionLocal() as db:
        r = await db.execute(
            select(ProductModel.offer_id, ProductModel.product_id).where(
                ProductModel.offer_id.in_(list(needed_codes))
            )
        )
        ozon_ids = {row.offer_id: row.product_id for row in r.fetchall() if row.product_id}

    print(f"[apply] tasks={len(tasks)} wb_found={len(wb)} ozon_ids_found={len(ozon_ids)}", flush=True)

    results = []
    async with aiohttp.ClientSession() as session:
        for task in tasks:
            code   = task["vendor_code"]
            fields = set(task.get("fields", []))
            wp     = wb.get(code)
            pid    = ozon_ids.get(code)

            print(f"[apply] code={code} pid={pid} fields={fields}", flush=True)

            if not wp:
                results.append({"vendor_code": code, "status": "error", "error": "Нет WB-данных в запросе"})
                continue
            if not pid:
                results.append({"vendor_code": code, "status": "error", "error": "product_id не найден в БД"})
                continue

            errors = []

            # ── Текст (name / description) ────────────────────────────────
            text_fields = fields & {"name", "description"}
            if text_fields:
                try:
                    # Получаем полные данные товара из Ozon (v4 — включая размеры)
                    async with session.post(
                        "https://api-seller.ozon.ru/v4/product/info/attributes",
                        headers=ozon_headers,
                        json={"filter": {"offer_id": [code]}, "limit": 1, "sort_dir": "ASC"},
                    ) as r:
                        info4 = await r.json(content_type=None)
                    current4 = info4.get("result", [{}])[0]

                    # v3 — только цена и ставка НДС
                    async with session.post(
                        "https://api-seller.ozon.ru/v3/product/info/list",
                        headers=ozon_headers,
                        json={"offer_id": [code]},
                    ) as r:
                        info3 = await r.json(content_type=None)
                    current3 = info3.get("items", [{}])[0]

                    if not current4:
                        errors.append("Не удалось получить данные товара с Ozon")
                        raise StopIteration

                    update_item = {
                        "offer_id":                code,
                        "name":                    wp.name if "name" in text_fields else current4.get("name", ""),
                        "description":             wp.description if "description" in text_fields else "",
                        "description_category_id": current4.get("description_category_id"),
                        "type_id":                 current4.get("type_id"),
                        "price":                   current3.get("price", "0"),
                        "vat":                     current3.get("vat", "0"),
                        "attributes":              current4.get("attributes", []),
                        "images":                  current4.get("images", []),
                        "depth":                   current4.get("depth", 0),
                        "width":                   current4.get("width", 0),
                        "height":                  current4.get("height", 0),
                        "dimension_unit":          current4.get("dimension_unit", "mm"),
                        "weight":                  current4.get("weight", 0),
                        "weight_unit":             current4.get("weight_unit", "g"),
                    }

                    print(f"[apply] text import offer_id={code}", flush=True)
                    async with session.post(
                        "https://api-seller.ozon.ru/v3/product/import",
                        headers=ozon_headers,
                        json={"items": [update_item]},
                    ) as r:
                        txt = await r.text()
                        print(f"[apply] text status={r.status} resp={txt[:200]}", flush=True)
                        if r.status != 200:
                            errors.append(f"Текст {r.status}: {txt[:200]}")
                        else:
                            # Обновляем локальную БД
                            db_vals: dict = {}
                            if "name"        in text_fields: db_vals["name"]        = wp.name
                            if "description" in text_fields: db_vals["description"] = wp.description
                            if db_vals:
                                async with AsyncSessionLocal() as db:
                                    await db.execute(
                                        update(ProductModel)
                                        .where(ProductModel.offer_id == code)
                                        .values(**db_vals)
                                    )
                                    await db.commit()

                except StopIteration:
                    pass
                except Exception as e:
                    errors.append(f"Текст: {e}")

            # ── Фото ─────────────────────────────────────────────────────
            if "images" in fields and wp.images:
                PROXY_BASE = "https://simacontrol.ru/api/img-proxy?url="
                images_to_send = [f"{PROXY_BASE}{u}" for u in wp.images[:10] if u]
                print(f"[apply] photo pid={pid} count={len(images_to_send)}", flush=True)
                try:
                    async with session.post(
                        "https://api-seller.ozon.ru/v1/product/pictures/import",
                        headers=ozon_headers,
                        json={"product_id": pid, "images": images_to_send},
                    ) as r:
                        txt = await r.text()
                        print(f"[apply] photo status={r.status} resp={txt[:300]}", flush=True)
                        if r.status != 200:
                            errors.append(f"Фото {r.status}: {txt[:200]}")
                        else:
                            # Обновляем главное фото в локальной БД
                            if wp.images:
                                async with AsyncSessionLocal() as db:
                                    await db.execute(
                                        update(ProductModel)
                                        .where(ProductModel.offer_id == code)
                                        .values(image_url=wp.images[0])
                                    )
                                    await db.commit()
                except Exception as e:
                    errors.append(f"Фото: {e}")

            results.append({
                "vendor_code": code,
                "status": "error" if errors else "ok",
                "error": "; ".join(errors) if errors else None,
            })

    return results


# ─────────────────────────────────────────────────
# Вспомогательные функции
# ─────────────────────────────────────────────────

def _merge_attributes(ozon_raw: list[dict], wb_attrs: list[dict]) -> list[dict]:
    wb_map = {a["name"].lower(): a["value"] for a in wb_attrs if a.get("name")}
    merged = []
    for attr in ozon_raw:
        key = attr.get("name", "").lower()
        if key in wb_map:
            merged.append({**attr, "values": [{"value": wb_map[key]}]})
        else:
            merged.append(attr)
    return merged


async def _validate_wb_images(urls: list[str]) -> list[str]:
    """Возвращаем URL как есть — Ozon сам проверит доступность."""
    return [u for u in urls if u]


_BOOL_TRUE_VALUES  = {"true", "да", "yes", "1", "есть", "имеется", "истина"}
_BOOL_FALSE_VALUES = {"false", "нет", "no", "0", "отсутствует", "ложь"}


def _to_ozon_boolean(text: str) -> Optional[str]:
    """Приводит произвольный текст WB к строгому true/false, которого ждёт Ozon
    для атрибутов типа Boolean. Если распознать не удалось — возвращает None
    (лучше не отправлять атрибут вовсе, чем отправить непонятный Ozon текст)."""
    t = (text or "").strip().lower()
    if t in _BOOL_TRUE_VALUES:
        return "true"
    if t in _BOOL_FALSE_VALUES:
        return "false"
    return None


# Атрибуты обязательной маркировки («Честный знак») — WB не даёт для них аналога,
# это юридически значимое решение, которое продавец должен принять сам.
_MARKING_KEYWORDS = ("маркиров", "честный знак")


def _is_marking_attr(attr_name: str) -> bool:
    low = (attr_name or "").lower()
    return any(k in low for k in _MARKING_KEYWORDS)


def _classify_error_messages(messages: list[str]) -> str:
    """'marking', если ВСЕ сообщения об ошибке касаются обязательной маркировки
    («Честный знак») — тогда это не баг, а решение, которое должен принять продавец;
    иначе — обычная 'error'."""
    if messages and all(_is_marking_attr(m) for m in messages):
        return "marking"
    return "error"


# ─────────────────────────────────────────────────
# Работа с WbAccount в БД
# ─────────────────────────────────────────────────

async def get_active_wb_key() -> Optional[str]:
    async with AsyncSessionLocal() as db:
        r = await db.execute(select(WbAccount).where(WbAccount.is_active == True))
        account = r.scalar_one_or_none()
        return account.api_key if account else None


async def get_all_wb_accounts_db() -> list[WbAccount]:
    async with AsyncSessionLocal() as db:
        r = await db.execute(select(WbAccount).order_by(WbAccount.id))
        return r.scalars().all()


# ═══════════════════════════════════════════════════════════════
# WB → OZON: ПЕРЕНОС КАРТОЧЕК
# ═══════════════════════════════════════════════════════════════

import json as _json
from datetime import datetime, timedelta


async def _ozon_headers_for_account(account_id: int) -> dict:
    from database import OzonAccount
    async with AsyncSessionLocal() as db:
        r = await db.execute(select(OzonAccount).where(OzonAccount.id == account_id))
        acc = r.scalar_one_or_none()
    if not acc:
        raise ValueError(f"Ozon account {account_id} not found")
    return {"Client-Id": str(acc.client_id), "Api-Key": acc.api_key, "Content-Type": "application/json"}


async def _ozon_get_offer_ids(headers: dict) -> set[str]:
    """Получаем все offer_id из кабинета Ozon."""
    ids: set[str] = set()
    last_id = ""
    async with aiohttp.ClientSession() as session:
        while True:
            async with session.post(
                "https://api-seller.ozon.ru/v3/product/list",
                headers=headers,
                json={"filter": {}, "last_id": last_id, "limit": 1000},
            ) as resp:
                data = await resp.json(content_type=None)
            items = data.get("result", {}).get("items", [])
            if not items:
                break
            for item in items:
                ids.add(item["offer_id"])
            last_id = data.get("result", {}).get("last_id", "")
            if not last_id:
                break
    return ids


async def _wb_get_stats(wb_api_key: str, nm_ids: list[int]) -> Optional[dict[int, dict]]:
    """Заказы за неделю и месяц по nm_id через WB nm-report.
    Возвращает None если API недоступен (не тот тип токена).
    """
    if not nm_ids:
        return {}

    today = datetime.now()
    stats_week: dict[int, int]  = {}
    stats_month: dict[int, int] = {}
    headers = {"Authorization": wb_api_key, "Content-Type": "application/json"}
    url = "https://seller-analytics-api.wildberries.ru/api/v2/nm-report/detail"

    async with aiohttp.ClientSession() as session:
        for period_key, days, store in [
            ("week",  7,  stats_week),
            ("month", 30, stats_month),
        ]:
            begin = (today - timedelta(days=days)).strftime("%Y-%m-%d 00:00:00")
            end   = today.strftime("%Y-%m-%d 23:59:59")
            page  = 1
            while True:
                try:
                    async with session.post(
                        url, headers=headers,
                        json={
                            "nmIDs": nm_ids,
                            "timezone": "Europe/Moscow",
                            "period": {"begin": begin, "end": end},
                            "orderBy": {"field": "ordersCount", "mode": "desc"},
                            "page": page,
                        },
                        timeout=aiohttp.ClientTimeout(total=10),
                    ) as resp:
                        if resp.status in (401, 403, 404):
                            print(f"[wb-stats] API недоступен: {resp.status} (нужен токен типа Analytics)", flush=True)
                            return None
                        data = await resp.json(content_type=None)
                    cards = data.get("data", {}).get("cards", [])
                    for card in cards:
                        nm_id = card.get("nmID")
                        for stat in card.get("statistics", []):
                            sp = stat.get("selectedPeriod", stat)
                            store[nm_id] = store.get(nm_id, 0) + sp.get("ordersCount", 0)
                    if not data.get("data", {}).get("isNext"):
                        break
                    page += 1
                except Exception as e:
                    print(f"[wb-stats] {period_key} error: {e}", flush=True)
                    return None

    return {
        nm_id: {"week": stats_week.get(nm_id, 0), "month": stats_month.get(nm_id, 0)}
        for nm_id in nm_ids
    }


async def find_wb_missing_from_ozon(ozon_account_id: int, wb_api_key: str) -> list[dict]:
    """WB-товары, которых нет в указанном Ozon-кабинете, с заказами за нед/мес."""
    from database import WbProductCache

    headers  = await _ozon_headers_for_account(ozon_account_id)
    ozon_ids = await _ozon_get_offer_ids(headers)

    async with AsyncSessionLocal() as db:
        r = await db.execute(select(WbProductCache))
        wb_all = r.scalars().all()

    missing = [p for p in wb_all if p.vendor_code not in ozon_ids]

    nm_ids = [p.nm_id for p in missing if p.nm_id]
    stats: Optional[dict[int, dict]] = None
    try:
        stats = await _wb_get_stats(wb_api_key, nm_ids)
    except Exception as e:
        print(f"[cross] stats failed: {e}", flush=True)

    result = []
    for p in missing:
        images = _json.loads(p.images_json or "[]")
        if stats is None:
            orders_week, orders_month = None, None
        else:
            nm_stat = stats.get(p.nm_id, {"week": 0, "month": 0}) if p.nm_id else {"week": 0, "month": 0}
            orders_week  = nm_stat["week"]
            orders_month = nm_stat["month"]
        result.append({
            "vendor_code":   p.vendor_code,
            "nm_id":         p.nm_id,
            "name":          p.name or "",
            "image":         images[0] if images else "",
            "subject_name":  getattr(p, "subject_name", None) or "",
            "orders_week":   orders_week,
            "orders_month":  orders_month,
        })

    result.sort(key=lambda x: (x["orders_month"] or 0), reverse=True)
    return result


# Кэш дерева категорий Ozon (обновляется раз в сессию)
_ozon_cat_pairs: list[tuple[int, int, str]] = []


async def _load_ozon_cat_pairs(session: aiohttp.ClientSession, headers: dict) -> list[tuple[int, int, str]]:
    """Загружает и кэширует плоский список (desc_cat_id, type_id, name) из дерева Ozon."""
    global _ozon_cat_pairs
    if _ozon_cat_pairs:
        return _ozon_cat_pairs

    try:
        async with session.post(
            "https://api-seller.ozon.ru/v1/description-category/tree",
            headers=headers,
            json={},
        ) as resp:
            tree = await resp.json(content_type=None)
    except Exception as e:
        print(f"[cat-tree] {e}", flush=True)
        return []

    pairs: list[tuple[int, int, str]] = []

    def traverse(nodes: list, parent_desc_cat: int = 0) -> None:
        for n in nodes:
            desc_cat = n.get("description_category_id") or parent_desc_cat
            type_id  = n.get("type_id", 0)
            name     = (n.get("category_name") or n.get("type_name") or "").lower()
            children = n.get("children") or []
            if type_id and not children and not n.get("disabled"):
                pairs.append((desc_cat, type_id, name))
            if children:
                traverse(children, desc_cat)

    traverse(tree.get("result", []))
    _ozon_cat_pairs = pairs
    print(f"[cat-tree] loaded {len(pairs)} leaf categories", flush=True)
    return pairs


async def _ozon_search_category(session: aiohttp.ClientSession, headers: dict, name: str) -> tuple[int, int]:
    """Ищет (description_category_id, type_id) по ключевым словам из названия товара."""
    import re as _re
    pairs = await _load_ozon_cat_pairs(session, headers)
    if not pairs:
        return 0, 0

    # Токенизация: разбиваем по пробелам и дефисам, минимум 3 символа
    raw_words = name.lower().split()
    tokens: set[str] = set()
    for w in raw_words:
        tokens.add(w)
        for part in _re.split(r'[-/]', w):
            if len(part) >= 3:
                tokens.add(part)

    # 1. Префиксный матч: общий префикс ≥ 4 символа → считается совпадением.
    #    effective = raw * matched / total_cat_words * (2 если первый токен совпал) - штраф за длину
    #    Бонус ×2 за совпадение первого (главного) токена гарантирует, что
    #    "крючок" бьёт "настенные часы" для запроса "Крючки настенные".
    first_token = next((w for w in name.lower().split() if len(w) >= 3), None)
    best_score, best = 0.0, (0, 0)
    for desc_cat, type_id, cat_name in pairs:
        cat_words_list = cat_name.split()
        raw = 0.0
        matched_cat_words = 0
        first_token_matched = False
        for cat_word in cat_words_list:
            if len(cat_word) < 3:
                continue
            for token in tokens:
                if len(token) < 3:
                    continue
                # Длина общего (совпадающего) префикса
                common = 0
                for i in range(min(len(token), len(cat_word))):
                    if token[i] == cat_word[i]:
                        common += 1
                    else:
                        break
                if common >= 4:
                    raw += common
                    matched_cat_words += 1
                    if token == first_token:
                        first_token_matched = True
                    break
        if raw > 0 and matched_cat_words > 0:
            effective = raw * matched_cat_words / len(cat_words_list)
            if first_token_matched:
                effective *= 3.0
            effective -= 0.001 * len(cat_name)  # тайбрейкер: предпочитаем короткие имена
            if effective > best_score:
                best_score = effective
                best = (desc_cat, type_id)

    if best_score >= 4.0:
        return best

    # 2. Substring: каждый токен ищем внутри имени категории
    for token in sorted(tokens, key=len, reverse=True):  # длинные слова — приоритет
        if len(token) < 4:
            continue
        for desc_cat, type_id, cat_name in pairs:
            if token in cat_name:
                return desc_cat, type_id

    return 0, 0


async def _find_ozon_dict_value(
    session: aiohttp.ClientSession, headers: dict,
    attr_id: int, desc_cat_id: int, search_val: str,
) -> Optional[int]:
    """Ищет dictionary_value_id в словаре атрибута Ozon по строке.
    Сначала пробует эндпоинт поиска, иначе перебирает первые страницы."""
    # Пробуем search-эндпоинт (если существует). Он ищет нечётко (по подстроке/похожести),
    # поэтому первый результат может оказаться близким, но НЕ тем же значением
    # (например, "Красный" вместо "Красный металлик") — это и приводило к
    # error_attribute_values_out_of_range на стороне Ozon. Принимаем результат
    # только при точном совпадении текста, иначе идём в постраничный fallback.
    search_lower = search_val.lower().strip()
    try:
        async with session.post(
            "https://api-seller.ozon.ru/v3/category/attribute/values/search",
            headers=headers,
            json={"attribute_id": attr_id, "category_id": desc_cat_id, "language": "RU", "value": search_val[:255]},
        ) as resp:
            if resp.status == 200:
                data = await resp.json(content_type=None)
                results = data.get("result", [])
                for r in results:
                    if r.get("value", "").lower().strip() == search_lower:
                        return r["id"]
    except Exception:
        pass

    # Fallback: перебираем словарь постранично (max 5 страниц по 1000)
    last_id = 0
    for _ in range(5):
        try:
            async with session.post(
                "https://api-seller.ozon.ru/v2/category/attribute/values",
                headers=headers,
                json={"attribute_id": attr_id, "category_id": desc_cat_id,
                      "language": "RU", "limit": 1000, "last_value_id": last_id},
            ) as resp:
                data = await resp.json(content_type=None)
            items = data.get("result", [])
            for item in items:
                if item.get("value", "").lower().strip() == search_lower:
                    return item["id"]
            if not data.get("has_next"):
                break
            last_id = items[-1]["id"] if items else 0
        except Exception:
            break
    return None


async def _ozon_category_attrs(
    session: aiohttp.ClientSession, headers: dict, desc_cat_id: int, type_id: int,
    required_only: bool = True,
) -> list[dict]:
    """Атрибуты категории Ozon. required_only=True — только обязательные (как раньше);
    False — полный список (нужен для составления Excel-шаблона)."""
    try:
        async with session.post(
            "https://api-seller.ozon.ru/v1/description-category/attribute",
            headers=headers,
            json={"description_category_id": desc_cat_id, "type_id": type_id, "language": "RU"},
        ) as resp:
            data = await resp.json(content_type=None)
        attrs = data.get("result", [])
        return [a for a in attrs if a.get("is_required")] if required_only else attrs
    except Exception as e:
        print(f"[attrs] {e}", flush=True)
    return []


# ─────────────────────────────────────────────────
# WB → OZON: Excel-шаблон для проверки перед созданием карточек
# ─────────────────────────────────────────────────

@dataclass
class WbOzonAttrDraft:
    id: int
    name: str
    attribute_type: str   # "None" | "Option" | "Tree"
    val_type: str          # "String" | "Integer" | "Float" ...
    is_required: bool
    value_text: str


@dataclass
class WbOzonDraft:
    vendor_code: str
    nm_id: Optional[int]
    name: str
    description: str
    category_name: str
    description_category_id: int
    type_id: int
    price: str
    old_price: str
    currency_code: str
    vat: str
    depth: int
    width: int
    height: int
    weight: int
    barcode: str
    extra_barcodes: list[str]
    images: list[str]
    primary_image: str
    attributes: list[WbOzonAttrDraft]
    issues: list[str]


async def _build_wb_ozon_draft(
    session: aiohttp.ClientSession, headers: dict, p, vc: str,
    cat_attrs_cache: dict[tuple[int, int], list[dict]],
) -> WbOzonDraft:
    """Строит черновик карточки Ozon из записи WbProductCache (или её отсутствия).
    Никогда не пропускает товар — при любой проблеме добавляет заметку в issues
    и продолжает с тем, что удалось определить."""
    import re as _re
    issues: list[str] = []
    PROXY = "https://simacontrol.ru/api/img-proxy?url="

    if p is None:
        issues.append("не найден в кеше WB — обновите синхронизацию WB перед экспортом")

    images_raw = _json.loads(p.images_json or "[]") if p else []
    images = [f"{PROXY}{u}" for u in images_raw[:10] if u]
    if not images:
        issues.append("нет фото")

    cat_query = (getattr(p, "subject_name", None) or p.name or vc) if p else vc
    desc_cat_id, type_id = await _ozon_search_category(session, headers, cat_query)
    category_name = ""
    if not desc_cat_id:
        issues.append("категория не определена — заполните ID категории и типа вручную")
    else:
        for dc, ti, nm in _ozon_cat_pairs:
            if dc == desc_cat_id and ti == type_id:
                category_name = nm
                break

    cache_key = (desc_cat_id, type_id)
    if desc_cat_id and cache_key not in cat_attrs_cache:
        cat_attrs_cache[cache_key] = await _ozon_category_attrs(session, headers, desc_cat_id, type_id, required_only=False)
    cat_attrs = cat_attrs_cache.get(cache_key, [])

    wb_chars = _json.loads(p.attributes_json or "[]") if p else []
    wb_attrs = {a["name"].lower(): str(a.get("value", "")) for a in wb_chars}
    wb_brand = (getattr(p, "brand", None) or "") if p else ""
    wb_desc  = (p.description or "") if p else ""

    attr_drafts: list[WbOzonAttrDraft] = []
    for attr in cat_attrs:
        attr_id     = attr["id"]
        attr_name   = attr.get("name", "")
        attr_name_l = attr_name.lower()
        dict_type   = attr.get("attribute_type", "None")
        val_type    = attr.get("type", "String")
        is_req      = bool(attr.get("is_required"))

        if "бренд" in attr_name_l:
            wb_val = wb_brand or wb_attrs.get(attr_name_l, "") or "Нет бренда"
        elif "тн вэд" in attr_name_l:
            wb_val = wb_attrs.get("код тн вэд") or wb_attrs.get(attr_name_l, "")
        elif "аннотация" in attr_name_l or "annotation" in attr_name_l:
            wb_val = wb_desc[:200].strip()
        else:
            wb_val = wb_attrs.get(attr_name_l, "")

        if dict_type in ("Option", "Tree"):
            value_text = wb_val
        elif val_type == "Boolean":
            value_text = _to_ozon_boolean(wb_val) or ""
        elif val_type in ("Integer", "Float"):
            nums = _re.findall(r'\d+(?:\.\d+)?', wb_val)
            value_text = nums[0] if nums else ""
        elif val_type == "URL" or "ссылка" in attr_name_l or "видео" in attr_name_l:
            # У WB нет источника ссылок (PDF-документы, видео и т.п.) — название товара сюда
            # подставлять нельзя, Ozon отклонит как невалидный URL (это касается и полей,
            # у которых Ozon формально указывает val_type="String", но по смыслу это ссылка,
            # например "Озон.Видеообложка: ссылка"). Берём WB-значение, только если оно само
            # похоже на ссылку, иначе оставляем пустым.
            value_text = wb_val if wb_val.startswith(("http://", "https://")) else ""
        elif wb_val:
            value_text = wb_val
        elif is_req:
            # Только для ОБЯЗАТЕЛЬНЫХ полей подставляем название товара как крайний случай —
            # для необязательных атрибутов лучше оставить пусто, чем угадывать неподходящим
            # текстом (так и раньше делалось для необязательных, теперь явно и системно).
            value_text = (p.name if p else vc) or vc
        else:
            value_text = ""

        if is_req and not value_text:
            issues.append(f"обязательный атрибут «{attr_name}» не заполнен")

        attr_drafts.append(WbOzonAttrDraft(
            id=attr_id, name=attr_name, attribute_type=dict_type,
            val_type=val_type, is_required=is_req, value_text=value_text,
        ))

    # Аннотация (id=4191) — не всегда входит в атрибуты категории как обязательная,
    # но важна для контента, добавляем принудительно если ещё не пришла из cat_attrs.
    if not any(a.id == 4191 for a in attr_drafts):
        attr_drafts.append(WbOzonAttrDraft(
            id=4191, name="Аннотация", attribute_type="None", val_type="String",
            is_required=False, value_text=wb_desc[:4000].strip(),
        ))

    wb_dims = _json.loads(getattr(p, "dimensions_json", None) or "{}") if p else {}

    # WB dimensions_json хранит значения в СМ → конвертируем в мм (×10) для Ozon API
    def _from_dims(key: str) -> Optional[int]:
        v = wb_dims.get(key)
        if v and int(v) > 0:
            return max(10, int(v) * 10)
        return None

    def _from_attrs_cm(keys: list[str]) -> Optional[int]:
        for k in keys:
            v = wb_attrs.get(k, "")
            nums = _re.findall(r'\d+(?:\.\d+)?', v)
            if nums:
                return max(10, int(float(nums[0]) * 10))
        return None

    depth_val  = _from_dims("length") or _from_attrs_cm(["глубина предмета", "длина предмета", "толщина предмета"])
    width_val  = _from_dims("width")  or _from_attrs_cm(["ширина предмета", "ширина"])
    height_val = _from_dims("height") or _from_attrs_cm(["высота предмета", "высота"])
    dims_defaulted = not (depth_val and width_val and height_val)
    depth, width, height = depth_val or 50, width_val or 50, height_val or 50

    wb_weight_brutto = wb_dims.get("weightBrutto") or 0
    weight_defaulted = False
    if float(wb_weight_brutto) > 0:
        weight = max(1, int(float(wb_weight_brutto) * 1000))
    else:
        def _weight_to_g(keys: list[str]) -> Optional[int]:
            for k in keys:
                v = wb_attrs.get(k, "")
                nums = _re.findall(r'\d+(?:\.\d+)?', v)
                if nums:
                    w = float(nums[0])
                    return max(1, int(w * 1000 if w < 100 else w))
            return None
        weight = _weight_to_g(["вес", "вес брутто", "вес нетто", "масса"])
        if weight is None:
            weight, weight_defaulted = 500, True

    if dims_defaulted:
        issues.append("размеры не определены — установлены 50×50×50 мм по умолчанию, проверьте")
    if weight_defaulted:
        issues.append("вес не определён — установлено 500 г по умолчанию, проверьте")

    wb_barcodes = _json.loads(getattr(p, "barcodes_json", None) or "[]") if p else []
    public_barcodes = [b for b in wb_barcodes if b and not b.startswith("29") and len(b) in (8, 12, 13, 14)]
    barcode = public_barcodes[0] if public_barcodes else ""
    extra_barcodes = public_barcodes[1:]
    if not barcode and wb_barcodes:
        issues.append("штрихкоды WB не публичные (внутренние 29xxx) — не перенесены, при необходимости добавьте EAN вручную")

    issues.append("цена не определена — установлено 10000 ₽ по умолчанию, проверьте")

    return WbOzonDraft(
        vendor_code=vc,
        nm_id=getattr(p, "nm_id", None) if p else None,
        name=((p.name if p else vc) or vc),
        description=(p.description or "") if p else "",
        category_name=category_name,
        description_category_id=desc_cat_id,
        type_id=type_id,
        price="10000",
        old_price="",
        currency_code="RUB",
        vat="0",
        depth=depth, width=width, height=height,
        weight=weight,
        barcode=barcode,
        extra_barcodes=extra_barcodes,
        images=images,
        primary_image="",
        attributes=attr_drafts,
        issues=issues,
    )


# Ставки НДС, которые принимает Ozon (доля от цены)
_OZON_VALID_VAT_RATES: set[float] = {0.0, 0.05, 0.07, 0.1, 0.2}

# Колонки видимого листа: (заголовок, ключ поля, вид)
# вид: "base" (строка), "base_int" (число), "base_images" (список ссылок через ;), "reference" (справочно, не уходит в Ozon)
_WTO_BASE_COLS: list[tuple[str, str, str]] = [
    ("Артикул",                  "offer_id",                 "base"),
    ("Проблема",                 "issues",                   "reference"),
    ("Название",                 "name",                      "base"),
    ("ID категории Ozon*",       "description_category_id",  "base_int"),
    ("ID типа Ozon*",            "type_id",                   "base_int"),
    ("Категория (справка)",      "category_name",             "reference"),
    ("Цена*",                    "price",                     "base"),
    ("Цена до скидки",           "old_price",                 "base"),
    ("Валюта",                   "currency_code",             "base"),
    ("НДС*",                     "vat",                        "base"),
    ("Глубина, мм*",             "depth",                      "base_int"),
    ("Ширина, мм*",              "width",                      "base_int"),
    ("Высота, мм*",              "height",                     "base_int"),
    ("Вес, г*",                  "weight",                     "base_int"),
    ("Штрихкод",                 "barcode",                    "base"),
    ("Доп. штрихкоды (справка)", "extra_barcodes",             "reference"),
    ("Фото (ссылки через ;)*",   "images",                     "base_images"),
    ("Главное фото",             "primary_image",              "base"),
    ("Описание",                 "description",                "base"),
    ("WB nm_id (справка)",       "nm_id",                       "reference"),
]


async def build_wb_to_ozon_template(ozon_account_id: int, vendor_codes: list[str]) -> bytes:
    """Строит Excel-шаблон для ручной проверки перед созданием карточек на Ozon.
    Каждый выбранный товар получает строку (даже если что-то не удалось определить —
    это отражается в колонке «Проблема», а не тихим пропуском строки)."""
    from database import WbProductCache
    import io as _io
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    headers = await _ozon_headers_for_account(ozon_account_id)

    async with AsyncSessionLocal() as db:
        r = await db.execute(select(WbProductCache).where(WbProductCache.vendor_code.in_(vendor_codes)))
        cache = {p.vendor_code: p for p in r.scalars().all()}

    cat_attrs_cache: dict[tuple[int, int], list[dict]] = {}
    async with aiohttp.ClientSession() as session:
        sem = asyncio.Semaphore(6)

        async def _one(vc: str) -> WbOzonDraft:
            async with sem:
                return await _build_wb_ozon_draft(session, headers, cache.get(vc), vc, cat_attrs_cache)

        drafts: list[WbOzonDraft] = list(await asyncio.gather(*[_one(vc) for vc in vendor_codes]))

    # Объединение атрибутных колонок по всем черновикам (по id, порядок первого появления)
    attr_order: list[int] = []
    attr_meta: dict[int, dict] = {}
    for d in drafts:
        for a in d.attributes:
            if a.id not in attr_meta:
                attr_order.append(a.id)
                attr_meta[a.id] = {"name": a.name, "attribute_type": a.attribute_type,
                                    "val_type": a.val_type, "is_required": a.is_required}
            elif a.is_required:
                attr_meta[a.id]["is_required"] = True

    attr_headers = [f"{attr_meta[aid]['name']}{'*' if attr_meta[aid]['is_required'] else ''}" for aid in attr_order]
    header_row = [h for h, _, _ in _WTO_BASE_COLS] + attr_headers

    wb_book = openpyxl.Workbook()
    ws = wb_book.active
    ws.title = "Карточки"
    ws.append(header_row)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header_row))}1"

    for d in drafts:
        row = [
            d.vendor_code, "; ".join(d.issues), d.name,
            d.description_category_id or "", d.type_id or "", d.category_name,
            d.price, d.old_price, d.currency_code, d.vat,
            d.depth, d.width, d.height, d.weight,
            d.barcode, "; ".join(d.extra_barcodes),
            "; ".join(d.images), d.primary_image, d.description, d.nm_id or "",
        ]
        attr_by_id = {a.id: a.value_text for a in d.attributes}
        for aid in attr_order:
            row.append(attr_by_id.get(aid, ""))
        ws.append(row)

    base_widths = [16, 34, 30, 12, 10, 22, 9, 12, 8, 7, 10, 10, 10, 9, 16, 18, 42, 20, 30, 12]
    for i, w in enumerate(base_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i in range(len(_WTO_BASE_COLS) + 1, len(header_row) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 24

    # Скрытый служебный лист — связывает колонку с полем Ozon (самодостаточность файла для импорта)
    meta_ws = wb_book.create_sheet("_meta")
    meta_ws.append(["col_index", "col_header", "field_kind", "field_key", "attribute_type", "val_type", "is_required"])
    for i, (h, key, kind) in enumerate(_WTO_BASE_COLS, start=1):
        meta_ws.append([i, h, kind, key, "", "", ""])
    for j, aid in enumerate(attr_order):
        i = len(_WTO_BASE_COLS) + 1 + j
        m = attr_meta[aid]
        meta_ws.append([i, header_row[i - 1], "attribute", aid, m["attribute_type"], m["val_type"], m["is_required"]])
    meta_ws.sheet_state = "hidden"

    buf = _io.BytesIO()
    wb_book.save(buf)
    return buf.getvalue()


async def check_marking_required(ozon_account_id: int, vendor_codes: list[str]) -> dict[str, bool]:
    """Для каждого vendor_code определяет ДО скачивания шаблона, есть ли в его
    категории Ozon обязательный атрибут маркировки («Честный знак») — чтобы
    продавец мог сразу увидеть предупреждение и сам решить, включать ли товар
    в перенос, а не узнавать об этом только после заполнения всего шаблона."""
    from database import WbProductCache

    headers = await _ozon_headers_for_account(ozon_account_id)

    async with AsyncSessionLocal() as db:
        r = await db.execute(select(WbProductCache).where(WbProductCache.vendor_code.in_(vendor_codes)))
        cache = {p.vendor_code: p for p in r.scalars().all()}

    cat_marking_cache: dict[tuple[int, int], bool] = {}
    result: dict[str, bool] = {}

    async with aiohttp.ClientSession() as session:
        sem = asyncio.Semaphore(6)

        async def _one(vc: str) -> None:
            async with sem:
                p = cache.get(vc)
                cat_query = (getattr(p, "subject_name", None) or (p.name if p else None) or vc) if p else vc
                desc_cat_id, type_id = await _ozon_search_category(session, headers, cat_query)
                if not desc_cat_id:
                    result[vc] = False
                    return
                key = (desc_cat_id, type_id)
                if key not in cat_marking_cache:
                    req_attrs = await _ozon_category_attrs(session, headers, desc_cat_id, type_id, required_only=True)
                    cat_marking_cache[key] = any(_is_marking_attr(a.get("name", "")) for a in req_attrs)
                result[vc] = cat_marking_cache[key]

        await asyncio.gather(*[_one(vc) for vc in vendor_codes])

    return result


def parse_wb_to_ozon_workbook(file_bytes: bytes) -> list[dict]:
    """Читает заполненный пользователем шаблон (без сетевых вызовов).
    Возвращает список строк: {vendor_code, base:{...}, attributes:[...], pre_errors:[...]}.
    pre_errors — проблемы, обнаруженные локально (пустые обязательные поля и т.п.) —
    такие строки не уходят в Ozon и сразу репортятся как ошибка."""
    import io as _io
    import openpyxl

    wb_book = openpyxl.load_workbook(_io.BytesIO(file_bytes), read_only=True, data_only=True)
    if "_meta" not in wb_book.sheetnames:
        raise ValueError("Файл повреждён или структура шаблона изменена (нет служебного листа _meta) — используйте оригинальный скачанный файл")
    if "Карточки" not in wb_book.sheetnames:
        raise ValueError("Не найден лист «Карточки» — используйте оригинальный скачанный файл")

    meta_rows = list(wb_book["_meta"].iter_rows(min_row=2, values_only=True))
    if not meta_rows:
        raise ValueError("Пустой служебный лист _meta")

    meta = []
    for row in meta_rows:
        col_index, col_header, field_kind, field_key, attribute_type, val_type, is_required = row
        meta.append({
            "col_index": int(col_index), "col_header": col_header, "field_kind": field_kind,
            "field_key": field_key, "attribute_type": attribute_type or "",
            "val_type": val_type or "", "is_required": bool(is_required),
        })
    meta.sort(key=lambda m: m["col_index"])

    main_ws = wb_book["Карточки"]
    header_row = next(main_ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    for m in meta:
        idx = m["col_index"] - 1
        actual = header_row[idx] if idx < len(header_row) else None
        if actual != m["col_header"]:
            raise ValueError(
                "Не изменяйте структуру столбцов шаблона (не добавляйте/не удаляйте/не переставляйте колонки) — "
                f"заполняйте только ячейки. Ожидался столбец «{m['col_header']}» на позиции {m['col_index']}, найден «{actual}»."
            )

    rows: list[dict] = []
    for data_row in main_ws.iter_rows(min_row=2, values_only=True):
        if data_row is None or all(v in (None, "") for v in data_row):
            continue

        base: dict = {}
        attributes: list[dict] = []
        pre_errors: list[str] = []
        vendor_code = ""

        for m in meta:
            idx = m["col_index"] - 1
            raw = data_row[idx] if idx < len(data_row) else None
            text = "" if raw is None else str(raw).strip()

            if m["field_key"] == "offer_id":
                vendor_code = text

            if m["field_kind"] == "reference":
                continue
            if m["field_kind"] == "attribute":
                if text:
                    attributes.append({
                        "id": int(m["field_key"]), "value_text": text,
                        "attribute_type": m["attribute_type"], "val_type": m["val_type"],
                        "is_required": m["is_required"],
                    })
                elif m["is_required"]:
                    col_clean = m["col_header"].rstrip("*")
                    if _is_marking_attr(col_clean):
                        pre_errors.append(
                            f"⚠ Нужен Честный знак! Заполните «{col_clean}» (true/false) в XLS и загрузите заново"
                        )
                    else:
                        pre_errors.append(f"обязательный атрибут «{col_clean}» не заполнен")
                continue
            if m["field_kind"] == "base_int":
                try:
                    base[m["field_key"]] = int(float(text)) if text else 0
                except ValueError:
                    pre_errors.append(f"«{m['col_header']}»: не число ({text!r})")
                    base[m["field_key"]] = 0
                continue
            if m["field_kind"] == "base_images":
                base[m["field_key"]] = [u.strip() for u in text.split(";") if u.strip()]
                continue
            if m["field_key"] == "vat":
                vat_text = text or "0"
                try:
                    vat_num = float(vat_text.replace(",", "."))
                except ValueError:
                    pre_errors.append(f"«{m['col_header']}»: не число ({text!r})")
                    base["vat"] = "0"
                    continue
                # Ozon ожидает долю (0.2 = 20%), а не проценты — если ввели "20", переводим в 0.2
                if vat_num > 1:
                    vat_num = round(vat_num / 100, 2)
                if not any(abs(vat_num - r) < 0.001 for r in _OZON_VALID_VAT_RATES):
                    pre_errors.append(
                        f"«{m['col_header']}»: ставка НДС {text!r} не поддерживается Ozon — "
                        f"допустимые значения: 0, 5, 7, 10, 20 (%)"
                    )
                    vat_num = 0
                base["vat"] = str(vat_num)
                continue
            base[m["field_key"]] = text

        if not vendor_code:
            continue

        if not base.get("name"):
            pre_errors.append("не заполнено название")
        if not base.get("description_category_id"):
            pre_errors.append("не заполнен ID категории")
        if not base.get("type_id"):
            pre_errors.append("не заполнен ID типа")
        if not base.get("price"):
            pre_errors.append("не заполнена цена")
        if not base.get("images"):
            pre_errors.append("нет ни одного фото")
        for dim_key, dim_name in [("depth", "глубина"), ("width", "ширина"), ("height", "высота"), ("weight", "вес")]:
            if not base.get(dim_key):
                pre_errors.append(f"не заполнен(а) {dim_name}")

        rows.append({"vendor_code": vendor_code, "base": base, "attributes": attributes, "pre_errors": pre_errors})

    return rows


# Статус фонового импорта — глобальный, по конвенции _sync_status/_ozon_sync_status в web_app.py.
# Читается через module-level атрибут (content_sync._wb_to_ozon_import_status), не через `from ... import`,
# чтобы переприсваивание внутри submit_wb_to_ozon_import было видно снаружи модуля.
_wb_to_ozon_import_status: dict = {"running": False, "total": 0, "results": {}, "error": ""}


async def _poll_ozon_import_task(session: aiohttp.ClientSession, headers: dict, task_id: int) -> dict:
    """POST /v1/product/import/info — статус батча импорта по task_id.
    Схема ответа подтверждена по исходникам github.com/diphantxm/ozon-api-client (ozon/products.go):
    {"result": {"items": [{"offer_id","product_id","status": "pending"|"imported"|"failed",
    "errors": [{"code","description","attribute_name",...}]}], "total": int}}.
    Терпима к неожиданной форме ответа — при сбое просто вернёт {} и попытка повторится
    на следующей итерации поллинга."""
    try:
        async with session.post(
            "https://api-seller.ozon.ru/v1/product/import/info",
            headers=headers,
            json={"task_id": task_id},
        ) as resp:
            data = await resp.json(content_type=None)
        print(f"[wto-import] poll task_id={task_id} status={resp.status} resp={str(data)[:500]}", flush=True)
        items = data.get("result", {}).get("items", []) or data.get("items", [])
        out = {}
        for it in items:
            offer_id = it.get("offer_id")
            status   = it.get("status")
            errors   = it.get("errors") or []
            error_text = "; ".join(e.get("description") or e.get("message") or e.get("code", "") for e in errors)
            out[offer_id] = {"status": status, "error_text": error_text}
        return out
    except Exception as e:
        print(f"[wto-import] poll error task_id={task_id}: {e}", flush=True)
        return {}


async def submit_wb_to_ozon_import(ozon_account_id: int, rows: list[dict]) -> None:
    """Резолвит словарные атрибуты, шлёт карточки батчами (≤100) в Ozon, поллит статус
    и обновляет _wb_to_ozon_import_status для живого прогресса в UI."""
    global _wb_to_ozon_import_status
    results: dict[str, dict] = {}
    for row in rows:
        if row["pre_errors"]:
            results[row["vendor_code"]] = {
                "status": _classify_error_messages(row["pre_errors"]),
                "message": "; ".join(row["pre_errors"]),
            }
        else:
            results[row["vendor_code"]] = {"status": "pending", "message": ""}
    _wb_to_ozon_import_status = {"running": True, "total": len(rows), "results": dict(results), "error": ""}

    try:
        headers = await _ozon_headers_for_account(ozon_account_id)
    except Exception as e:
        _wb_to_ozon_import_status = {"running": False, "total": len(rows), "results": results, "error": str(e)}
        return

    valid_rows = [r for r in rows if not r["pre_errors"]]
    if not valid_rows:
        _wb_to_ozon_import_status = {"running": False, "total": len(rows), "results": results, "error": ""}
        return

    dict_cache: dict[tuple, Optional[int]] = {}
    pending_tasks: list[dict] = []  # [{"task_id": int, "vendor_codes": [...]}]

    async with aiohttp.ClientSession() as session:

        async def _resolve_dict(attr_id: int, desc_cat_id: int, text: str) -> Optional[int]:
            key = (attr_id, desc_cat_id, text.lower())
            if key not in dict_cache:
                dict_cache[key] = await _find_ozon_dict_value(session, headers, attr_id, desc_cat_id, text)
            return dict_cache[key]

        items_to_send: list[tuple[str, dict]] = []
        for row in valid_rows:
            base = row["base"]
            desc_cat_id = base.get("description_category_id", 0)
            ozon_attrs = []
            notes = []
            for a in row["attributes"]:
                if a["attribute_type"] in ("Option", "Tree"):
                    dict_val_id = await _resolve_dict(a["id"], desc_cat_id, a["value_text"])
                    if dict_val_id:
                        ozon_attrs.append({"id": a["id"], "complex_id": 0,
                                            "values": [{"dictionary_value_id": dict_val_id, "value": a["value_text"]}]})
                    else:
                        notes.append(f"атрибут «{a['id']}»: значение «{a['value_text']}» не найдено в справочнике Ozon, не отправлено")
                elif a["val_type"] == "Boolean":
                    bool_val = _to_ozon_boolean(a["value_text"])
                    if bool_val:
                        ozon_attrs.append({"id": a["id"], "complex_id": 0, "values": [{"value": bool_val}]})
                    else:
                        notes.append(f"атрибут «{a['id']}»: значение «{a['value_text']}» не распознано как true/false, не отправлено")
                elif a["val_type"] in ("Integer", "Float"):
                    ozon_attrs.append({"id": a["id"], "complex_id": 0, "values": [{"value": a["value_text"]}]})
                else:
                    ozon_attrs.append({"id": a["id"], "complex_id": 0, "values": [{"value": a["value_text"][:500]}]})

            item = {
                "offer_id":                row["vendor_code"],
                "name":                    (base.get("name") or row["vendor_code"])[:500],
                "description":             (base.get("description") or "")[:10000],
                "description_category_id": desc_cat_id,
                "type_id":                 base.get("type_id", 0),
                "price":                   str(base.get("price") or "0"),
                "currency_code":           base.get("currency_code") or "RUB",
                "vat":                     str(base.get("vat") or "0"),
                "images":                  base.get("images") or [],
                "depth":                   base.get("depth", 0),
                "width":                   base.get("width", 0),
                "height":                  base.get("height", 0),
                "dimension_unit":          "mm",
                "weight":                  base.get("weight", 0),
                "weight_unit":             "g",
                "attributes":              ozon_attrs,
            }
            if base.get("old_price"):
                item["old_price"] = str(base["old_price"])
            if base.get("primary_image"):
                item["primary_image"] = base["primary_image"]
            if base.get("barcode"):
                item["barcode"] = base["barcode"]

            if notes:
                results[row["vendor_code"]]["message"] = "; ".join(notes)
            items_to_send.append((row["vendor_code"], item))
            _wb_to_ozon_import_status = {"running": True, "total": len(rows), "results": dict(results), "error": ""}

        for i in range(0, len(items_to_send), 100):
            chunk = items_to_send[i:i + 100]
            try:
                async with session.post(
                    "https://api-seller.ozon.ru/v3/product/import",
                    headers=headers, json={"items": [it for _, it in chunk]},
                ) as resp:
                    resp_data = await resp.json(content_type=None)
                print(f"[wto-import] submit chunk_start={i} status={resp.status} resp={str(resp_data)[:400]}", flush=True)

                if resp.status == 429 or "item_limit_exceeded" in str(resp_data).lower():
                    retry_after = resp.headers.get("Item-Retry-After", "")
                    msg = "лимит Ozon исчерпан, попробуйте позже" + (f" (через {retry_after} мин)" if retry_after else "")
                    for vc, _ in items_to_send[i:]:
                        results[vc] = {"status": "error", "message": msg}
                    _wb_to_ozon_import_status = {"running": True, "total": len(rows), "results": dict(results), "error": ""}
                    break

                if resp.status != 200:
                    err = resp_data.get("message") or str(resp_data)[:300]
                    for vc, _ in chunk:
                        results[vc] = {"status": "error", "message": f"Ozon {resp.status}: {err}"}
                    _wb_to_ozon_import_status = {"running": True, "total": len(rows), "results": dict(results), "error": ""}
                    continue

                task_id = resp_data.get("result", {}).get("task_id")
                if task_id:
                    pending_tasks.append({"task_id": task_id, "vendor_codes": [vc for vc, _ in chunk]})
                else:
                    for vc, _ in chunk:
                        results[vc] = {"status": "error", "message": "Ozon не вернул task_id"}
            except Exception as e:
                for vc, _ in chunk:
                    results[vc] = {"status": "error", "message": str(e)}
                _wb_to_ozon_import_status = {"running": True, "total": len(rows), "results": dict(results), "error": ""}

        deadline = asyncio.get_event_loop().time() + 120
        while pending_tasks and asyncio.get_event_loop().time() < deadline:
            still_pending = []
            for task in pending_tasks:
                statuses = await _poll_ozon_import_task(session, headers, task["task_id"])
                remaining = []
                for vc in task["vendor_codes"]:
                    st = statuses.get(vc)
                    if not st or st["status"] == "pending":
                        remaining.append(vc)
                        continue
                    if st["status"] == "imported":
                        # Ozon иногда помечает товар "imported", но прикладывает предупреждение
                        # (например по конкретному атрибуту) — не отменяет успех, но стоит показать.
                        notes = [n for n in (results.get(vc, {}).get("message", ""), st["error_text"]) if n]
                        results[vc] = {"status": "ok", "message": "; ".join(notes)}
                    else:
                        err_text = st["error_text"] or "Ozon отклонил карточку"
                        err_parts = [p.strip() for p in err_text.split(";") if p.strip()]
                        results[vc] = {"status": _classify_error_messages(err_parts), "message": err_text}
                if remaining:
                    still_pending.append({"task_id": task["task_id"], "vendor_codes": remaining})
            pending_tasks = still_pending
            _wb_to_ozon_import_status = {"running": True, "total": len(rows), "results": dict(results), "error": ""}
            if pending_tasks:
                await asyncio.sleep(4)

        for task in pending_tasks:
            for vc in task["vendor_codes"]:
                results[vc] = {"status": "unknown", "message": "Ozon не подтвердил статус за 2 мин — проверьте в кабинете Ozon"}

    _wb_to_ozon_import_status = {"running": False, "total": len(rows), "results": results, "error": ""}