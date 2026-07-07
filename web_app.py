"""
web_app.py — Веб-интерфейс для управления заказами

📚 УРОК: Структура FastAPI-приложения
--------------------------------------
FastAPI строится на двух типах маршрутов (routes):
  1. Страницы (HTML) — возвращают HTMLResponse через Jinja2-шаблоны
  2. API-эндпоинты (JSON) — возвращают dict/list, используются JS-кодом на странице

Авторизация работает так:
  - POST /login → проверяет логин/пароль → создаёт JWT → кладёт в httpOnly cookie
  - httpOnly = JS не может прочитать cookie (защита от XSS-атак)
  - Каждый запрос → FastAPI читает cookie → декодирует JWT → знает кто вы
"""

import os
import json
import aiohttp
import asyncio
from datetime import datetime, timedelta, date, timezone
from typing import Optional

import json as _json
import asyncio as _asyncio

from fastapi import FastAPI, Request, Form, Depends, HTTPException, status, BackgroundTasks, File, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from jose import JWTError, jwt
from passlib.context import CryptContext
from sqlalchemy import select, func, or_, update, delete
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.dialects.postgresql import insert as pg_insert

from database import (AsyncSessionLocal, User, Order, VirtualOrder, Product, PriceHistory, CostHistory, SalesHistory,
                      FboWatchlist, StockItem, OzonAccount, WbAccount, WbProductCache, FboStorageReport, FboSalesWatch,
                      FboDailyDigest, init_db)

from content_sync import get_matched_products, apply_wb_to_ozon, get_active_wb_key, get_all_wb_accounts_db

# --- КОНФИГУРАЦИЯ ---
# SECRET_KEY используется для подписи JWT. Если утечёт — злоумышленник сможет
# создать поддельный токен с любой ролью. Поэтому храним в .env, не в коде.
SECRET_KEY = os.getenv("JWT_SECRET_KEY", "ЗАМЕНИТЕ_ЭТО_НА_СЛУЧАЙНУЮ_СТРОКУ_В_ENV")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_HOURS = 12

# CryptContext — менеджер хэширования паролей.
# bcrypt автоматически добавляет "соль" (случайные данные), поэтому
# одинаковые пароли дают разные хэши. Это защита от радужных таблиц.
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

app = FastAPI(title="SimaBot Dashboard")
templates = Jinja2Templates(directory="templates")
app.mount("/static", StaticFiles(directory="static"), name="static")

# --- OZON API КОНФИГУРАЦИЯ ---
OZON_CLIENT_ID = os.getenv("OZON_CLIENT_ID", "")
OZON_API_KEY = os.getenv("OZON_API_KEY", "")
try:
    OZON_WAREHOUSE_ID = int(os.getenv("OZON_WAREHOUSE_ID", "0"))
except:
    OZON_WAREHOUSE_ID = 0
try:
    OZON_WAREHOUSE_ID_UZSPACE = int(os.getenv("OZON_WAREHOUSE_ID_UZSPACE", "0"))
except:
    OZON_WAREHOUSE_ID_UZSPACE = 0

_active_account = {
    "id": None,
    "name": "Основной",
    "client_id": OZON_CLIENT_ID,
    "api_key": OZON_API_KEY,
    "warehouse_id_uzspace": OZON_WAREHOUSE_ID_UZSPACE,
}

OZON_HEADERS = {
    "Client-Id": OZON_CLIENT_ID,
    "Api-Key": OZON_API_KEY,
    "Content-Type": "application/json"
}

def get_active_headers():
    return {
        "Client-Id": _active_account["client_id"],
        "Api-Key": _active_account["api_key"],
        "Content-Type": "application/json"
    }

async def load_active_account():
    global _active_account, OZON_HEADERS
    print("load_active_account called", flush=True)
    try:
        async with AsyncSessionLocal() as db:
            r = await db.execute(
                select(OzonAccount).where(OzonAccount.is_active == True).limit(1)
            )
            account = r.scalars().first()
            print(f"load_active_account found: {account}", flush=True)
            if account:
                _active_account = {
                    "id": account.id,
                    "name": account.name,
                    "client_id": account.client_id,
                    "api_key": account.api_key,
                    "warehouse_id_uzspace": account.warehouse_id_uzspace or 0,
                }
                OZON_HEADERS = {
                    "Client-Id": account.client_id,
                    "Api-Key": account.api_key,
                    "Content-Type": "application/json"
                }
                print(f"Active account: {account.name}", flush=True)
    except Exception as e:
        print(f"load_active_account error: {e}", flush=True)

def get_ozon_headers():
    return {
        "Client-Id": _active_account["client_id"],
        "Api-Key": _active_account["api_key"],
        "Content-Type": "application/json"
    }

OZON_HEADERS = {
    "Client-Id": OZON_CLIENT_ID,
    "Api-Key": OZON_API_KEY,
    "Content-Type": "application/json"
}

# Храним время последней синхронизации и статус фоновой задачи
_last_sync: Optional[datetime] = None
_sync_status = {"running": False, "progress": "", "synced": 0, "error": ""}


async def fetch_ozon_postings(statuses: list) -> list:
    """Получает отправления с Ozon по списку статусов."""
    url = "https://api-seller.ozon.ru/v3/posting/fbs/list"
    date_to = datetime.now() + timedelta(days=1)
    date_from = date_to - timedelta(days=60)
    all_postings = []

    async with aiohttp.ClientSession() as session:
        for status in statuses:
            payload = {
                "dir": "ASC",
                "filter": {
                    "status": status,
                    "warehouse_id": [OZON_WAREHOUSE_ID],
                    "since": date_from.strftime("%Y-%m-%dT%H:%M:%S.000Z"),
                    "to": date_to.strftime("%Y-%m-%dT%H:%M:%S.000Z")
                },
                "limit": 1000,
                "with": {"analytics_data": False, "financial_data": False}
            }
            try:
                async with session.post(url, json=payload, headers=get_active_headers()) as resp:
                    if resp.status == 200:
                        data = await resp.json()
                        all_postings.extend(data.get("result", {}).get("postings", []))
            except Exception as e:
                pass

    return all_postings


async def fetch_images_for_skus(skus: list) -> dict:
    """Батч-запрос фото по SKU через /v3/product/info/list.
    Ответ: {"items": [{..., "sources": [{"sku": 123}], "primary_image": [...]}]}
    SKU находится в sources[].sku, а не на верхнем уровне.
    """
    result = {}
    url = "https://api-seller.ozon.ru/v3/product/info/list"
    async with aiohttp.ClientSession() as session:
        for i in range(0, len(skus), 100):
            batch = skus[i:i+100]
            try:
                async with session.post(url, json={"sku": batch}, headers=get_active_headers()) as resp:
                    if resp.status == 200:
                        data = await resp.json()
                        # Ответ: {"items": [...]} без обёртки result
                        for item in data.get("items", []):
                            images = item.get("primary_image", [])
                            if not images:
                                images = item.get("images", [])
                            image_url = images[0] if isinstance(images, list) and images else ""
                            if not image_url:
                                continue
                            # SKU лежит в sources — берём все и маппим
                            for source in item.get("sources", []):
                                sku = source.get("sku")
                                if sku:
                                    result[int(sku)] = image_url
            except Exception as e:
                import logging
                logging.warning(f"fetch_images_for_skus error: {e}")
    return result


async def sync_from_ozon() -> dict:
    """
    Основная функция синхронизации: получает свежие данные с Ozon,
    обновляет/добавляет заказы в БД, возвращает статистику.

    📚 УРОК: Upsert (INSERT ... ON CONFLICT DO UPDATE)
    Позволяет одной командой вставить запись или обновить если она уже есть.
    Это атомарная операция — не нужно делать SELECT перед INSERT.
    """
    global _last_sync

    postings = await fetch_ozon_postings(["awaiting_packaging", "awaiting_deliver"])

    if not postings:
        return {"synced": 0, "error": "Нет данных от Ozon или ошибка API"}

    # Собираем все SKU для батч-запроса фото
    all_skus = []
    for p in postings:
        for prod in p.get("products", []):
            if prod.get("sku"):
                all_skus.append(int(prod["sku"]))

    sku_images = {}
    if all_skus:
        sku_images = await fetch_images_for_skus(list(set(all_skus)))

    # Множество актуальных номеров отправлений от Ozon
    active_posting_numbers = {p.get("posting_number") for p in postings}

    def parse_dt(s):
        if not s:
            return None
        try:
            return datetime.fromisoformat(s.rstrip("Z").split(".")[0])
        except:
            return None

    async with AsyncSessionLocal() as db:
        # 1. Upsert активных заказов
        for p in postings:
            posting_number = p.get("posting_number")
            ozon_status = p.get("status", "unknown")
            shipment_date = p.get("shipment_date")
            in_process_at = p.get("in_process_at")

            products = []
            for prod in p.get("products", []):
                sku = int(prod["sku"]) if prod.get("sku") else None
                products.append({
                    "offer_id": prod.get("offer_id"),
                    "sku": sku,
                    "name": prod.get("name"),
                    "quantity": prod.get("quantity"),
                    "price": prod.get("price", "0"),
                    "image_url": sku_images.get(sku, "") if sku else ""
                })

            stmt = pg_insert(Order).values(
                posting_number=posting_number,
                ozon_status=ozon_status,
                products_json=json.dumps(products, ensure_ascii=False),
                ozon_accepted_at=parse_dt(in_process_at),
                ozon_created_at=parse_dt(shipment_date),
            )
            stmt = stmt.on_conflict_do_update(
                index_elements=["posting_number"],
                set_={
                    "ozon_status": ozon_status,
                    "products_json": json.dumps(products, ensure_ascii=False),
                    "ozon_accepted_at": parse_dt(in_process_at),
                    "ozon_created_at": parse_dt(shipment_date),
                }
            )
            await db.execute(stmt)

        # 2. Заказы которые были активными в нашей БД но пропали у Ozon —
        #    значит отменены/доставлены. Помечаем их соответствующим статусом.
        #    Запрашиваем актуальный статус каждого через /v3/posting/fbs/get
        stale_q = await db.execute(
            select(Order.posting_number)
            .where(Order.ozon_status.in_(["awaiting_packaging", "awaiting_deliver"]))
        )
        stale_postings = [
            row[0] for row in stale_q.all()
            if row[0] not in active_posting_numbers
        ]

        if stale_postings:
            url = "https://api-seller.ozon.ru/v3/posting/fbs/get"
            async with aiohttp.ClientSession() as check_session:
                for p_num in stale_postings:
                    try:
                        async with check_session.post(
                            url,
                            json={"posting_number": p_num},
                            headers=get_active_headers()
                        ) as resp:
                            if resp.status == 200:
                                data = await resp.json()
                                real_status = data.get("result", {}).get("status", "unknown")
                                await db.execute(
                                    Order.__table__.update()
                                    .where(Order.posting_number == p_num)
                                    .values(ozon_status=real_status)
                                )
                    except Exception as e:
                        import logging
                        logging.warning(f"sync stale check error {p_num}: {e}")

        await db.commit()

    _last_sync = datetime.now()
    removed = len(stale_postings) if stale_postings else 0
    return {
        "synced": len(postings),
        "removed": removed,
        "last_sync": _last_sync.strftime("%d.%m.%Y %H:%M")
    }


# --- ЗАВИСИМОСТИ (Dependency Injection) ---
# 📚 УРОК: Depends() в FastAPI
# Вместо того чтобы копировать код проверки токена в каждый маршрут,
# мы выносим его в отдельную функцию и подключаем через Depends().
# FastAPI сам вызовет её перед вызовом нашего обработчика.

async def get_db() -> AsyncSession:
    """Открывает сессию БД и закрывает её после запроса (паттерн 'контекстный менеджер')"""
    async with AsyncSessionLocal() as session:
        yield session


def create_access_token(data: dict) -> str:
    """Создаёт JWT токен с данными пользователя и сроком жизни"""
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(hours=ACCESS_TOKEN_EXPIRE_HOURS)
    to_encode.update({"exp": expire})
    # jwt.encode() шифрует данные с помощью SECRET_KEY + алгоритма HS256
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)


def get_current_user(request: Request) -> Optional[dict]:
    token = request.cookies.get("access_token")
    if not token:
        return None
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        return {"username": payload.get("sub"), "role": payload.get("role"), "permissions": payload.get("permissions", [])}
    except JWTError:
        return None

async def get_current_user_db(request: Request, db: AsyncSession = Depends(get_db)) -> Optional[dict]:
    """Читает permissions из БД — актуальные данные без перелогина"""
    user = get_current_user(request)
    if not user:
        return None
    result = await db.execute(select(User).where(User.username == user["username"]))
    u = result.scalars().first()
    if not u:
        return None
    perms = u.permissions if isinstance(u.permissions, list) else (json.loads(u.permissions) if u.permissions else [])
    return {"username": u.username, "role": u.role, "permissions": perms}


def require_admin(request: Request) -> dict:
    """Depends-функция: требует роль admin, иначе редирект на логин"""
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=303, headers={"Location": "/login"})
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Только для администратора")
    return user


async def require_any_role(request: Request, db: AsyncSession = Depends(get_db)) -> dict:
    """Depends-функция: читает permissions из БД при каждом запросе"""
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=303, headers={"Location": "/login"})
    result = await db.execute(select(User).where(User.username == user["username"]))
    u = result.scalars().first()
    if not u:
        raise HTTPException(status_code=303, headers={"Location": "/login"})
    perms = u.permissions if isinstance(u.permissions, list) else (json.loads(u.permissions) if u.permissions else [])
    return {"username": u.username, "role": u.role, "permissions": perms}


# --- СТРАНИЦЫ (HTML) ---

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    user = get_current_user(request)
    if user:
        return RedirectResponse("/", status_code=302)
    return templates.TemplateResponse("login.html", {"request": request})


@app.post("/login")
async def login_submit(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    db: AsyncSession = Depends(get_db)
):
    """
    📚 УРОК: Form(...) означает что данные ожидаются из HTML-формы (application/x-www-form-urlencoded),
    а не из JSON-тела запроса. Для JSON используется Pydantic BaseModel.
    """
    result = await db.execute(select(User).where(User.username == username))
    user = result.scalars().first()

    # pwd_context.verify() проверяет пароль против хэша в БД
    # Важно: не сравниваем строки напрямую! Это защита от timing-атак
    if not user or not pwd_context.verify(password, user.password_hash):
        return templates.TemplateResponse("login.html", {
            "request": request,
            "error": "Неверный логин или пароль"
        })

    import json as _json
    raw = user.permissions
    if isinstance(raw, list):
        perms = raw
    elif isinstance(raw, str):
        perms = _json.loads(raw)
    else:
        perms = ["dashboard", "orders", "queue", "users", "repricer", "costs"] if user.role == "admin" else ["queue"]
    token = create_access_token({"sub": user.username, "role": user.role, "permissions": perms})
    # Редиректим на нужную страницу в зависимости от роли
    redirect_url = "/" if user.role == "admin" else "/queue"
    response = RedirectResponse(redirect_url, status_code=302)

    # httpOnly=True — JS не может читать эту cookie (защита от XSS)
    # samesite="lax" — cookie не отправляется при запросах с других сайтов (защита от CSRF)
    response.set_cookie("access_token", token, httponly=True, samesite="lax",
                        max_age=ACCESS_TOKEN_EXPIRE_HOURS * 3600)
    return response


@app.get("/logout")
async def logout():
    response = RedirectResponse("/login", status_code=302)
    response.delete_cookie("access_token")
    return response


@app.get("/", response_class=HTMLResponse)
async def dashboard(request: Request, user: dict = Depends(require_any_role)):
    if "dashboard" not in user.get("permissions", []) and user["role"] != "admin":
        return RedirectResponse("/queue", status_code=302)
    """Главная страница — дашборд для администратора"""
    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "user": user,
        "active_tab": "dashboard"
    })


@app.get("/orders", response_class=HTMLResponse)
async def orders_page(request: Request, user: dict = Depends(require_any_role)):
    if "orders" not in user.get("permissions", []) and user["role"] != "admin":
        return RedirectResponse("/queue", status_code=302)
    return templates.TemplateResponse("orders.html", {
        "request": request,
        "user": user,
        "active_tab": "orders"
    })


@app.get("/queue", response_class=HTMLResponse)
async def queue_page(request: Request, user: dict = Depends(require_any_role)):
    """Очередь на сборку — доступна и фулфилменту, и администратору"""
    return templates.TemplateResponse("queue.html", {
        "request": request,
        "user": user,
        "active_tab": "queue"
    })


# --- API ENDPOINTS (JSON) ---
# 📚 УРОК: Почему API отдельно от HTML-страниц?
# Страница загружается один раз. Потом JS делает fetch() к API-эндпоинтам,
# чтобы обновить данные без перезагрузки. Это называется SPA (Single Page App) паттерн.

@app.get("/api/stats")
async def get_stats(
    user: dict = Depends(require_any_role),
    db: AsyncSession = Depends(get_db)
):
    """
    Статистика для дашборда.
    Логика:
    - Всего активных = awaiting_packaging + awaiting_deliver
    - Отправка сегодня = plan_delivery_date == сегодня И статус активный
    - Просроченные = plan_delivery_date < сегодня И sur_number пустой И статус активный
    """
    today = date.today()
    today_iso = today.isoformat()          # "2026-03-05" для сравнения ISO-дат
    today_dot = today.strftime("%d.%m.%Y") # "05.03.2026" для сравнения дд.мм.гггг
    active_statuses = ["awaiting_packaging", "awaiting_deliver"]

    # Вспомогательная функция: конвертирует дд.мм.гггг → гггг-мм-дд для корректного сравнения
    # На уровне SQL используем CASE для обоих форматов
    from sqlalchemy import case, cast, and_
    from sqlalchemy.sql.expression import literal

    def to_iso_expr(col):
        """
        Конвертирует строковую дату любого формата в DATE через PostgreSQL.
        Форматы в БД:
          - '5.03.26'   → d.mm.yy  (короткий, от бота)
          - '06.03.2026' → dd.mm.yyyy (полный)
          - '2026-03-05' → ISO (от сайта)
        Используем to_date() с нужным форматом через CASE по длине строки.
        """
        from sqlalchemy import func as sqlfunc
        return sqlfunc.to_date(
            case(
                # ISO формат yyyy-mm-dd (10 символов с дефисами)
                (col.op('~')(r'^\d{4}-\d{2}-\d{2}$'), col),
                # Полный дд.мм.гггг (10 символов: 06.03.2026)
                (col.op('~')(r'^\d{2}\.\d{2}\.\d{4}$'),
                 sqlfunc.concat(
                     sqlfunc.substring(col, 7, 4), '-',
                     sqlfunc.substring(col, 4, 2), '-',
                     sqlfunc.substring(col, 1, 2)
                 )),
                # Короткий д.мм.гг или дд.мм.гг (6-8 символов, год 2 цифры)
                # '5.03.26' → день может быть 1 цифра, год 2 цифры → добавляем '20' к году
                else_=sqlfunc.concat(
                    '20',
                    sqlfunc.split_part(col, '.', 3), '-',
                    sqlfunc.lpad(sqlfunc.split_part(col, '.', 2), 2, '0'), '-',
                    sqlfunc.lpad(sqlfunc.split_part(col, '.', 1), 2, '0')
                )
            ),
            literal('YYYY-MM-DD')
        )

    today_date_expr = today  # Python date object для сравнения

    def overdue_filter():
        return and_(
            Order.ozon_status.in_(active_statuses),
            Order.plan_delivery_date != None,
            Order.plan_delivery_date != "",
            or_(Order.sur_number == None, Order.sur_number == ""),
            to_iso_expr(Order.plan_delivery_date) < today_date_expr
        )

    def today_filter():
        return and_(
            Order.ozon_status.in_(active_statuses),
            Order.plan_delivery_date != None,
            Order.plan_delivery_date != "",
            to_iso_expr(Order.plan_delivery_date) == today_date_expr
        )

    # Всего активных заказов
    total_q = await db.execute(
        select(func.count(Order.posting_number))
        .where(Order.ozon_status.in_(active_statuses))
    )
    total_count = total_q.scalar()

    # Отправка сегодня
    today_q = await db.execute(
        select(func.count(Order.posting_number)).where(today_filter())
    )
    today_count = today_q.scalar()

    # Просроченные: дата поставки < сегодня И СУР не заполнен
    overdue_q = await db.execute(
        select(func.count(Order.posting_number)).where(overdue_filter())
    )
    overdue_count = overdue_q.scalar()

    # Виртуальных (ручная сборка)
    virtual_q = await db.execute(select(func.count(VirtualOrder.posting_number)))
    virtual_count = virtual_q.scalar()

    # Просроченные заказы для таблицы на дашборде
    overdue_orders_q = await db.execute(
        select(Order)
        .where(overdue_filter())
        .order_by(to_iso_expr(Order.plan_delivery_date).asc())
        .limit(50)
    )
    overdue_orders = overdue_orders_q.scalars().all()

    return {
        "total": total_count,
        "today": today_count,
        "overdue": overdue_count,
        "virtual": virtual_count,
        "last_sync": _last_sync.strftime("%d.%m.%Y %H:%M") if _last_sync else None,
        "overdue_orders": [
            {
                "posting_number": o.posting_number,
                "ozon_status": o.ozon_status,
                "plan_delivery_date": o.plan_delivery_date,
                "sima_order_number": o.sima_order_number,
                "sur_number": o.sur_number,
                "ozon_accepted_at": o.ozon_accepted_at.isoformat() if o.ozon_accepted_at else None,
                "products": json.loads(o.products_json) if o.products_json else [],
            }
            for o in overdue_orders
        ]
    }


@app.get("/api/orders")
async def get_orders(
    user: dict = Depends(require_any_role),
    db: AsyncSession = Depends(get_db),
    page: int = 1,
    per_page: int = 50,
    status: str = "active",
    search: str = "",
    date_from: str = "",
    date_to: str = "",
    sort: str = "accepted_asc",
    sur: str = "",
    invoice: str = "",
):
    """
    📚 УРОК: Пагинация + фильтрация
    Параметры приходят из query string: /api/orders?page=1&status=...&search=...
    SQLAlchemy позволяет цепочкой добавлять .where() — они объединяются через AND.
    """
    offset = (page - 1) * per_page
    # Сортировка
    if sort == "accepted_asc":
        query = select(Order).order_by(Order.ozon_accepted_at.asc().nullslast())
    elif sort == "accepted_desc":
        query = select(Order).order_by(Order.ozon_accepted_at.desc().nullslast())
    else:
        query = select(Order).order_by(Order.added_at.desc())

    ACTIVE_STATUSES = ["awaiting_packaging", "awaiting_deliver"]
    if status == "active":
        query = query.where(Order.ozon_status.in_(ACTIVE_STATUSES))
    elif status:
        query = query.where(Order.ozon_status == status)
    if search:
        # ilike — регистронезависимый LIKE. % — любое количество символов.
        # Ищем по номеру отправления ИЛИ по содержимому products_json (артикулы)
        query = query.where(
            Order.posting_number.ilike(f"%{search}%") |
            Order.products_json.ilike(f"%{search}%")
        )
    if date_from:
        query = query.where(Order.plan_delivery_date >= date_from)
    if date_to:
        query = query.where(Order.plan_delivery_date <= date_to)

    if sur:
        query = query.where(Order.sur_number.ilike(f"%{sur}%"))
    if invoice:
        query = query.where(Order.sima_order_number.ilike(f"%{invoice}%"))

    # Считаем total с теми же фильтрами
    count_query = select(func.count(Order.posting_number))
    if status == "active":
        count_query = count_query.where(Order.ozon_status.in_(ACTIVE_STATUSES))
    elif status:
        count_query = count_query.where(Order.ozon_status == status)
    if search:
        count_query = count_query.where(
            Order.posting_number.ilike(f"%{search}%") |
            Order.products_json.ilike(f"%{search}%")
        )
    if date_from:
        count_query = count_query.where(Order.plan_delivery_date >= date_from)
    if date_to:
        count_query = count_query.where(Order.plan_delivery_date <= date_to)

    if sur:
        count_query = count_query.where(Order.sur_number.ilike(f"%{sur}%"))
    if invoice:
        count_query = count_query.where(Order.sima_order_number.ilike(f"%{invoice}%"))

    total = (await db.execute(count_query)).scalar()

    result = await db.execute(query.offset(offset).limit(per_page))
    orders = result.scalars().all()

    import json
    return {
        "total": total,
        "page": page,
        "per_page": per_page,
        "orders": [
            {
                "posting_number": o.posting_number,
                "not_delivered": bool(o.not_delivered),
                "ozon_status": o.ozon_status,
                "sima_order_number": o.sima_order_number,
                "sima_order_date": o.sima_order_date,
                "plan_delivery_date": o.plan_delivery_date,
                "sur_number": o.sur_number,
                "ff_delivery_date": o.ff_delivery_date.isoformat() if o.ff_delivery_date else None,
                "comment": o.comment,
                "added_at": o.added_at.isoformat() if o.added_at else None,
                "products": json.loads(o.products_json) if o.products_json else [],
                "ozon_accepted_at": o.ozon_accepted_at.isoformat() if o.ozon_accepted_at else None,
            }
            for o in orders
        ]
    }


@app.patch("/api/orders/{posting_number}")
async def update_order(
    posting_number: str,
    request: Request,
    user: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    """
    📚 УРОК: HTTP-методы
    GET — получить данные
    POST — создать новую запись
    PATCH — частично обновить запись (только переданные поля)
    PUT — полностью заменить запись
    DELETE — удалить

    Мы используем PATCH, потому что фулфилмент обновляет только СУР и дату,
    а не весь объект заказа.
    """
    body = await request.json()

    result = await db.execute(select(Order).where(Order.posting_number == posting_number))
    order = result.scalars().first()
    if not order:
        raise HTTPException(status_code=404, detail="Заказ не найден")

    # Обновляем только переданные поля (PATCH-семантика)
    if "sur_number" in body:
        order.sur_number = body["sur_number"]
    if "ff_delivery_date" in body and body["ff_delivery_date"]:
        try:
            order.ff_delivery_date = datetime.fromisoformat(body["ff_delivery_date"])
        except ValueError:
            raise HTTPException(status_code=422, detail="Неверный формат даты")
    if "comment" in body:
        order.comment = body["comment"]
    if "plan_delivery_date" in body:
        order.plan_delivery_date = body["plan_delivery_date"]
    if "sima_order_number" in body:
        order.sima_order_number = body["sima_order_number"]

    await db.commit()
    return {"ok": True, "posting_number": posting_number}

@app.post("/api/orders/{posting_number}/not_delivered")
async def toggle_not_delivered(
    posting_number: str,
    user: dict = Depends(require_any_role),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(select(Order).where(Order.posting_number == posting_number))
    order = result.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="Заказ не найден")
    order.not_delivered = not bool(order.not_delivered)
    await db.commit()
    return {"not_delivered": order.not_delivered}

@app.get("/api/queue")
async def get_queue(
    user: dict = Depends(require_any_role),
    db: AsyncSession = Depends(get_db)
):
    """
    Очередь на ручную сборку — виртуальные заказы с деталями.
    Доступна фулфилменту.
    """
    import json

    # JOIN: берём виртуальные заказы и подтягиваем детали из таблицы orders
    # 📚 УРОК: outerjoin — LEFT JOIN в SQL.
    # Возвращает все строки из VirtualOrder, даже если в Order нет совпадения.
    from sqlalchemy.orm import aliased
    query = (
        select(VirtualOrder, Order)
        .outerjoin(Order, VirtualOrder.posting_number == Order.posting_number)
        .order_by(Order.plan_delivery_date.asc().nullslast())
    )
    result = await db.execute(query)
    rows = result.all()

    queue = []
    for virtual, order in rows:
        products = []
        if order and order.products_json:
            try:
                products = json.loads(order.products_json)
            except Exception:
                pass
        queue.append({
            "posting_number": virtual.posting_number,
            "added_at": virtual.added_at.isoformat() if virtual.added_at else None,
            "sima_order_number": order.sima_order_number if order else None,
            "plan_delivery_date": order.plan_delivery_date if order else None,
            "sur_number": order.sur_number if order else None,
            "comment": order.comment if order else None,
            "products": products,
        })

    return {"queue": queue, "total": len(queue)}


# --- УТИЛИТЫ ДЛЯ ПЕРВОГО ЗАПУСКА ---

@app.post("/api/setup/create-admin")
async def create_admin(
    request: Request,
    db: AsyncSession = Depends(get_db)
):
    """
    Создаёт первого администратора. Работает только если в БД нет ни одного пользователя.
    После создания первого admin — эндпоинт больше не работает (защита).

    📚 УРОК: Такой эндпоинт называется "bootstrap" или "setup route".
    В production его часто удаляют или защищают отдельным секретом.
    """
    existing = await db.execute(select(func.count(User.id)))
    if existing.scalar() > 0:
        raise HTTPException(status_code=403, detail="Пользователи уже созданы")

    body = await request.json()
    username = body.get("username", "admin")
    password = body.get("password")
    if not password:
        raise HTTPException(status_code=422, detail="Нужен пароль")

    hashed = pwd_context.hash(password)
    admin = User(username=username, password_hash=hashed, role="admin")
    db.add(admin)
    await db.commit()
    return {"ok": True, "message": f"Администратор '{username}' создан"}


@app.post("/api/setup/create-fulfillment")
async def create_fulfillment_user(
    request: Request,
    user: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    """Создаёт аккаунт для сотрудника фулфилмента. Только admin."""
    body = await request.json()
    username = body.get("username")
    password = body.get("password")
    if not username or not password:
        raise HTTPException(status_code=422, detail="Нужны username и password")

    existing = await db.execute(select(User).where(User.username == username))
    if existing.scalars().first():
        raise HTTPException(status_code=409, detail="Пользователь уже существует")

    hashed = pwd_context.hash(password)
    ff_user = User(username=username, password_hash=hashed, role="fulfillment")
    db.add(ff_user)
    await db.commit()
    return {"ok": True, "message": f"Пользователь '{username}' создан"}


# =====================================================================
# РЕПРАЙСЕР — OZON API ФУНКЦИИ
# =====================================================================

async def fetch_products_prices_offset(offset: int = 0, limit: int = 1000) -> dict:
    """Получает цены через /v5/product/info/prices с offset-пагинацией."""
    url = "https://api-seller.ozon.ru/v5/product/info/prices"
    async with aiohttp.ClientSession() as session:
        async with session.post(url,
            json={"filter": {"visibility": "ALL"}, "limit": limit, "offset": offset},
            headers=get_active_headers()) as resp:
            if resp.status == 200:
                return await resp.json()
    return {}


async def sync_products_catalog() -> dict:
    global _sync_status, _last_sync
    print("SYNC STARTED", flush=True)
    _sync_status = {"running": True, "progress": "Загружаем цены...", "synced": 0, "error": ""}

    try:
        # ШАГ 1: Все товары через last_id пагинацию /v3/product/list
        all_items = []
        last_id = ""
        limit = 1000

        while True:
            async with aiohttp.ClientSession() as session:
                async with session.post(
                        "https://api-seller.ozon.ru/v3/product/list",
                        headers=get_active_headers(),
                        json={"filter": {"visibility": "ALL"}, "last_id": last_id, "limit": limit}
                ) as resp:
                    list_data = await resp.json()

            result = list_data.get("result", {})
            list_items = result.get("items", [])
            if not list_items:
                break

            # Конвертируем в формат совместимый с дальнейшим кодом
            for li in list_items:
                all_items.append({
                    "offer_id": li.get("offer_id"),
                    "product_id": li.get("product_id"),
                    "price": {},
                    "commissions": {},
                    "price_indexes": {},
                })

            last_id = result.get("last_id", "")
            _sync_status["progress"] = f"Товары: {len(all_items)}..."
            print(f"STEP 1: fetched {len(all_items)}", flush=True)

            if not last_id or len(list_items) < limit:
                break

        if not all_items:
            _sync_status = {"running": False, "progress": "", "synced": 0, "error": "Нет товаров от Ozon"}
            return {"synced": 0, "error": "Нет товаров от Ozon"}

        total = len(all_items)
        product_ids = [item["product_id"] for item in all_items if item.get("product_id")]
        offer_ids = [item["offer_id"] for item in all_items if item.get("offer_id")]



        # ШАГ 2: Фото, название, категория, склад
        print("STEP 2.0: fetching warehouses", flush=True)
        warehouse_name_map = {}
        try:
            async with aiohttp.ClientSession() as session:
                async with session.post(
                        "https://api-seller.ozon.ru/v1/warehouse/list",
                        json={},
                        headers=get_active_headers()
                ) as resp:
                    if resp.status == 200:
                        wh_data = await resp.json()
                        for wh in wh_data.get("result", []):
                            wh_id = wh.get("warehouse_id")
                            wh_name = wh.get("name", "")
                            if wh_id and wh_name:
                                warehouse_name_map[wh_id] = wh_name
            print(f"STEP 2.0 DONE: {warehouse_name_map}", flush=True)
        except Exception as e:
            import logging
            logging.warning(f"warehouse list error: {e}")
        _sync_status["progress"] = f"Загружаем инфо о товарах..."

        # ШАГ 2.1: Загружаем дерево категорий для маппинга id → название
        category_name_map = {}
        try:
            async with aiohttp.ClientSession() as session:
                async with session.post(
                        "https://api-seller.ozon.ru/v1/description-category/tree",
                        json={"language": "RU"},
                        headers=get_active_headers()
                ) as resp:
                    if resp.status == 200:
                        cat_data = await resp.json()

                        def flatten_cats(nodes):
                            for node in nodes:
                                cid = node.get("description_category_id")
                                cname = node.get("category_name", "")
                                if cid and cname:
                                    category_name_map[cid] = cname
                                flatten_cats(node.get("children", []))

                        flatten_cats(cat_data.get("result", []))
            print(f"STEP 2.1 DONE: {len(category_name_map)} categories", flush=True)
        except Exception as e:
            print(f"category tree error: {e}", flush=True)

        info_map = {}
        url_info = "https://api-seller.ozon.ru/v3/product/info/list"
        timeout = aiohttp.ClientTimeout(total=120)
        async with aiohttp.ClientSession(timeout=timeout) as session:
            for i in range(0, len(product_ids), 100):
                batch = product_ids[i:i+100]
                try:
                    async with session.post(url_info, json={"product_id": batch}, headers=get_active_headers()) as resp:
                        if resp.status == 200:
                            d = await resp.json()
                            for it in d.get("items", []):
                                pid = it.get("id")
                                images = it.get("primary_image") or it.get("images", [])
                                img = images[0] if isinstance(images, list) and images else (images or "")
                                sources = it.get("sources", [])
                                sku_fbs = next((s["sku"] for s in sources if s.get("source") == "fbs"), None)
                                sku_sds = next((s["sku"] for s in sources if s.get("source") == "sds"), None)
                                if sources:
                                    print(f"SOURCES SAMPLE: {sources[:2]}", flush=True)
                                warehouse_names = []
                                for s in sources:
                                    wh_id = s.get("warehouse_id")
                                    wh_name = warehouse_name_map.get(wh_id, s.get("source", ""))
                                    if wh_name and wh_name not in warehouse_names:
                                        warehouse_names.append(wh_name)
                                warehouse = ", ".join(warehouse_names)
                                cat_id = it.get("description_category_id")
                                info_map[pid] = {
                                    "name": it.get("name", ""),
                                    "image_url": img,
                                    "category_id": cat_id,
                                    "category_name": category_name_map.get(cat_id, ""),
                                    "warehouse_type": warehouse,
                                    "sku_fbs": sku_fbs,
                                    "sku_sds": sku_sds,
                                }
                except Exception as e:
                    import logging; logging.warning(f"sync info error: {e}")


        print(f"STEP 2 DONE: {len(info_map)} items", flush=True)


        # ШАГ 3: Бренд (attribute_id=85)
        _sync_status["progress"] = f"Загружаем бренды..."
        brand_map = {}
        url_attrs = "https://api-seller.ozon.ru/v4/product/info/attributes"
        timeout = aiohttp.ClientTimeout(total=30)
        async with aiohttp.ClientSession(timeout=timeout) as session:
            for i in range(0, len(product_ids), 100):
                batch = offer_ids[i:i+100]
                try:
                    async with session.post(url_attrs,
                        json={"filter": {"offer_id": batch}, "limit": len(batch)},
                        headers=get_active_headers()) as resp:
                        if resp.status == 200:
                            d = await resp.json()
                            for it in d.get("result", []):
                                oid = it.get("offer_id")
                                for attr in it.get("attributes", []):
                                    if attr.get("id") == 85:
                                        vals = attr.get("values", [])
                                        if vals:
                                            brand_map[oid] = vals[0].get("value", "")
                except Exception as e:
                    import logging; logging.warning(f"sync attrs error: {e}")

        # ШАГ 4: Upsert в БД
        _sync_status["progress"] = f"Записываем в БД ({total} товаров)..."
        async with AsyncSessionLocal() as db:
            for item in all_items:
                offer_id = item.get("offer_id")
                if not offer_id:
                    continue
                product_id = item.get("product_id")
                price_data = item.get("price") or {}
                comm = item.get("commissions") or {}
                idx = item.get("price_indexes") or {}
                info = info_map.get(product_id) or {}
                ext_data = idx.get("external_index_data") or {}
                ozon_data = idx.get("ozon_index_data") or {}
                ext_min = float(ext_data.get("min_price") or 0)
                ozon_min = float(ozon_data.get("min_price") or 0)
                competitor_min = int(min(ext_min if ext_min else 999999, ozon_min if ozon_min else 999999))
                if competitor_min == 999999:
                    competitor_min = 0
                fbs_logistics = int(
                    float(comm.get("fbs_direct_flow_trans_max_amount") or 0) +
                    float(comm.get("fbs_deliv_to_customer_amount") or 0)
                )
                row = dict(
                    offer_id=offer_id,
                    product_id=product_id,
                    name=info.get("name") or "",
                    image_url=info.get("image_url") or "",
                    category_id=info.get("category_id"),
                    category_name=info.get("category_name") or "",
                    warehouse_type=info.get("warehouse_type") or "",
                    sku_fbs=info.get("sku_fbs"),
                    sku_sds=info.get("sku_sds"),
                    brand=brand_map.get(offer_id) or "",
                    price=int(float(price_data.get("price") or 0)),
                    old_price=int(float(price_data.get("old_price") or 0)),
                    min_price=int(float(price_data.get("min_price") or 0)),
                    net_price=int(float(price_data.get("net_price") or 0)),
                    marketing_price=int(float(price_data.get("marketing_seller_price") or 0)),
                    commission_fbs_percent=int(comm.get("sales_percent_fbs") or 0),
                    commission_fbs_logistics=fbs_logistics,
                    price_index_color=idx.get("color_index") or "",
                    price_index_ozon=str(round(float(ozon_data.get("price_index_value") or 0), 2)),
                    price_index_external=str(round(float(ext_data.get("price_index_value") or 0), 2)),
                    competitor_min_price=competitor_min,
                    updated_at=datetime.now(),
                )
                stmt = pg_insert(Product).values(**row)
                upd = {k: v for k, v in row.items() if k != "offer_id"}
                stmt = stmt.on_conflict_do_update(index_elements=["offer_id"], set_=upd)
                await db.execute(stmt)
            await db.commit()

        _last_sync = datetime.now()
        _sync_status = {"running": False, "progress": "Готово", "synced": total, "error": ""}
        return {"synced": total}

    except Exception as e:
        import logging, traceback
        logging.error(f"sync_products error: {traceback.format_exc()}")
        _sync_status = {"running": False, "progress": "", "synced": 0, "error": str(e)}
        return {"synced": 0, "error": str(e)}


@app.post("/api/repricer/sync")
async def api_repricer_sync(user: dict = Depends(require_admin)):
    """Запускает синхронизацию каталога в фоне и сразу отвечает."""
    import asyncio
    if _sync_status.get("running"):
        return {"started": False, "message": "Синхронизация уже запущена", "status": _sync_status}
    asyncio.create_task(sync_products_catalog())
    return {"started": True, "message": "Синхронизация запущена в фоне"}


@app.get("/api/repricer/sync/status")
async def api_repricer_sync_status(user: dict = Depends(require_admin)):
    """Статус текущей синхронизации."""
    return {
        "running": _sync_status.get("running", False),
        "progress": _sync_status.get("progress", ""),
        "synced": _sync_status.get("synced", 0),
        "error": _sync_status.get("error", ""),
        "last_sync": _last_sync.strftime("%d.%m.%Y %H:%M") if _last_sync else None,
    }


@app.get("/api/repricer/filters")
async def api_repricer_filters(
    user: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    """Возвращает уникальные значения для фильтров: бренды, категории, склады."""
    brands_r = await db.execute(
        select(Product.brand).where(Product.brand != None).where(Product.brand != "")
        .distinct().order_by(Product.brand)
    )
    categories_r = await db.execute(
        select(Product.category_name, Product.category_id)
        .where(Product.category_name != None).where(Product.category_name != "")
        .distinct().order_by(Product.category_name)
    )
    warehouses_r = await db.execute(
        select(Product.warehouse_type).where(Product.warehouse_type != None).where(Product.warehouse_type != "")
        .distinct().order_by(Product.warehouse_type)
    )
    return {
        "brands": [r[0] for r in brands_r.all()],
        "categories": [{"id": r[1], "name": r[0]} for r in categories_r.all()],
        "warehouses": [r[0] for r in warehouses_r.all()],
    }


@app.get("/api/repricer/products")
async def api_repricer_products(
    search: str = "",
    index_filter: str = "",
    brand: str = "",
    category_id: str = "",
    warehouse: str = "",
    demand_only: str = "",
    page: int = 1,
    per_page: int = 100,
    user: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    """Список товаров для репрайсера с фильтрацией."""
    from sqlalchemy import or_

    query = select(Product)
    count_q = select(func.count(Product.offer_id))

    filters = []
    if search:
        like = f"%{search}%"
        filters.append(or_(Product.offer_id.ilike(like), Product.name.ilike(like)))
    if index_filter:
        filters.append(Product.price_index_color == index_filter.upper())
    if brand:
        filters.append(Product.brand == brand)
    if category_id:
        filters.append(Product.category_id == int(category_id))
    if warehouse:
        filters.append(Product.warehouse_type.ilike(f"%{warehouse}%"))
    if demand_only == "1":
        filters.append(Product.demand_rule_enabled == True)

    for f in filters:
        query = query.where(f)
        count_q = count_q.where(f)

    query = query.order_by(Product.price_index_color.asc(), Product.offer_id.asc())
    query = query.limit(per_page).offset((page - 1) * per_page)

    total_r = await db.execute(count_q)
    total = total_r.scalar()
    products_r = await db.execute(query)
    products = products_r.scalars().all()

    def calc_margin(p: Product) -> Optional[float]:
        if not p.cost_price or not p.price or p.price == 0:
            return None
        # Чистая выручка = цена - комиссия% - логистика
        net = p.price * (1 - (p.commission_fbs_percent or 0) / 100) - (p.commission_fbs_logistics or 0)
        if net <= 0:
            return None
        margin = (net - p.cost_price) / net * 100
        return round(margin, 1)

    def calc_min_price_for_margin(p: Product, target_pct: int) -> Optional[int]:
        """Минимальная цена для достижения целевой маржи."""
        if not p.cost_price:
            return None
        # net = price * (1 - comm%) - logistics
        # margin = (net - cost) / net = target/100
        # Решаем: price * (1 - comm%) - logistics = cost / (1 - target/100)
        comm = (p.commission_fbs_percent or 0) / 100
        logistics = p.commission_fbs_logistics or 0
        target = target_pct / 100
        if comm >= 1 or target >= 1:
            return None
        needed_net = p.cost_price / (1 - target)
        price = (needed_net + logistics) / (1 - comm)
        return int(price) + 1

    return {
        "total": total,
        "page": page,
        "per_page": per_page,
        "products": [
            {
                "offer_id": p.offer_id,
                "product_id": p.product_id,
                "name": p.name,
                "image_url": p.image_url,
                "price": p.price,
                "old_price": p.old_price,
                "min_price": p.min_price,
                "net_price": p.net_price,
                "marketing_price": p.marketing_price,
                "commission_fbs_percent": p.commission_fbs_percent,
                "commission_fbs_logistics": p.commission_fbs_logistics,
                "price_index_color": p.price_index_color,
                "price_index_ozon": p.price_index_ozon,
                "competitor_min_price": p.competitor_min_price,
                "cost_price": p.cost_price,
                "target_margin_pct": p.target_margin_pct,
                "auto_reprice_enabled": p.auto_reprice_enabled,
                "current_margin": calc_margin(p),
                "suggested_price": calc_min_price_for_margin(p, p.target_margin_pct) if p.target_margin_pct else None,
                "updated_at": p.updated_at.isoformat() if p.updated_at else None,
                "demand_rule_enabled": bool(p.demand_rule_enabled) if p.demand_rule_enabled is not None else False,
                "demand_min_orders": p.demand_min_orders or 3,
                "demand_step_pct": p.demand_step_pct or 5,
            }
            for p in products
        ]
    }

@app.get("/api/repricer/demand")
async def api_repricer_demand(
    offer_ids: str = "",          # comma-separated, пустой = все товары с demand_rule_enabled
    user: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    """
    Считает спрос за последние 7 дней по каждому артикулу.
    Источник 1: таблица orders (products_json)
    Источник 2: price_history (косвенно — если нет заказов в orders)

    Возвращает: {offer_id: {orders_7d, trend, recommended_action, recommended_price}}
    """
    import json
    from datetime import timedelta
    from sqlalchemy import text

    ids = [i.strip() for i in offer_ids.split(",") if i.strip()] if offer_ids else []

    week_ago = datetime.now() - timedelta(days=7)
    two_weeks_ago = datetime.now() - timedelta(days=14)

    # --- Считаем заказы из таблицы orders за 7 и 14 дней ---
    # products_json хранит список объектов с offer_id
    # Используем PostgreSQL JSON-операторы для эффективного поиска
    raw = await db.execute(
        text("""
            SELECT
                prod->>'offer_id' AS offer_id,
                COUNT(*) FILTER (WHERE o.added_at >= :week_ago) AS orders_7d,
                COUNT(*) FILTER (WHERE o.added_at >= :two_weeks_ago AND o.added_at < :week_ago) AS orders_prev_7d
            FROM orders o,
                 jsonb_array_elements(o.products_json::jsonb) AS prod
            WHERE o.added_at >= :two_weeks_ago
              AND o.ozon_status NOT IN ('cancelled')
            GROUP BY prod->>'offer_id'
        """),
        {"week_ago": week_ago, "two_weeks_ago": two_weeks_ago}
    )
    demand_from_orders = {row.offer_id: {"orders_7d": row.orders_7d, "orders_prev_7d": row.orders_prev_7d}
                         for row in raw.all() if row.offer_id}

    # --- Получаем настройки товаров ---
    if ids:
        prod_q = await db.execute(select(Product).where(Product.offer_id.in_(ids)))
    else:
        prod_q = await db.execute(select(Product).where(Product.demand_rule_enabled == True))
    products = prod_q.scalars().all()

    result = {}
    for p in products:
        d = demand_from_orders.get(p.offer_id, {"orders_7d": 0, "orders_prev_7d": 0})
        orders_7d = int(d["orders_7d"])
        orders_prev = int(d["orders_prev_7d"])

        # Тренд: сравниваем текущую и прошлую неделю
        if orders_prev == 0 and orders_7d == 0:
            trend = "flat"
        elif orders_prev == 0:
            trend = "up"
        elif orders_7d > orders_prev * 1.1:
            trend = "up"
        elif orders_7d < orders_prev * 0.9:
            trend = "down"
        else:
            trend = "flat"

        min_orders = p.demand_min_orders or 3
        step_pct = p.demand_step_pct or 5

        # Рекомендация
        if not p.price:
            recommended_action = None
            recommended_price = None
        elif orders_7d >= min_orders:
            # Спрос достаточный — можно поднять цену
            recommended_action = "raise"
            recommended_price = int(p.price * (1 + step_pct / 100))
        else:
            # Спрос низкий — снизить цену
            recommended_action = "lower"
            new_price = int(p.price * (1 - step_pct / 100))
            # Не опускаем ниже min_price и не ниже цены с минимальной маржой 10%
            floor = p.min_price or 0
            if p.cost_price:
                # Минимум: себестоимость / 0.9 (маржа не ниже 10%)
                margin_floor = int(p.cost_price / 0.9)
                floor = max(floor, margin_floor)
            recommended_price = max(new_price, floor) if floor else new_price

        result[p.offer_id] = {
            "orders_7d": orders_7d,
            "orders_prev_7d": orders_prev,
            "trend": trend,
            "min_orders": min_orders,
            "step_pct": step_pct,
            "recommended_action": recommended_action,
            "recommended_price": recommended_price,
            "demand_rule_enabled": bool(p.demand_rule_enabled),
        }

    return {"demand": result}


@app.patch("/api/repricer/bulk-demand-settings")
async def api_bulk_demand_settings(
    body: dict,
    user: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    """
    Массовое назначение настроек demand-репрайсера для выбранных артикулов.
    body: {
        offer_ids: ["art1", "art2", ...],
        demand_rule_enabled: bool,
        demand_min_orders: int,   # порог заказов/неделю
        demand_step_pct: int,     # шаг % изменения цены
    }
    """
    offer_ids = body.get("offer_ids", [])
    if not offer_ids:
        raise HTTPException(400, "offer_ids required")

    updates = {}
    if "demand_rule_enabled" in body:
        updates["demand_rule_enabled"] = bool(body["demand_rule_enabled"])
    if "demand_min_orders" in body and body["demand_min_orders"]:
        updates["demand_min_orders"] = int(body["demand_min_orders"])
    if "demand_step_pct" in body and body["demand_step_pct"]:
        updates["demand_step_pct"] = int(body["demand_step_pct"])

    if not updates:
        raise HTTPException(400, "no fields to update")

    from sqlalchemy import update as sa_update
    await db.execute(
        sa_update(Product)
        .where(Product.offer_id.in_(offer_ids))
        .values(**updates)
    )
    await db.commit()
    return {"ok": True, "updated": len(offer_ids)}

@app.patch("/api/repricer/products/{offer_id}")
async def api_update_product(
    offer_id: str,
    body: dict,
    user: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    """Обновляет себестоимость, целевую маржу, правило автоприценки."""
    result = await db.execute(select(Product).where(Product.offer_id == offer_id))
    product = result.scalars().first()

    if not product:
        # Создаём запись если нет
        product = Product(offer_id=offer_id)
        db.add(product)

    allowed = ["cost_price", "target_margin_pct", "auto_reprice_enabled", "auto_rule", "auto_rule_value", "min_price"]
    for field in allowed:
        if field in body:
            setattr(product, field, body[field])

    await db.commit()
    return {"ok": True}


@app.post("/api/repricer/apply-price")
async def api_apply_price(
    body: dict,
    user: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    """
    Применяет новую цену через Ozon API.
    body: {offer_id, new_price, old_price?, min_price?}
    """
    offer_id = body.get("offer_id")
    new_price = body.get("new_price")

    if not offer_id or not new_price:
        raise HTTPException(400, "offer_id and new_price required")

    # Получаем текущие данные товара
    result = await db.execute(select(Product).where(Product.offer_id == offer_id))
    product = result.scalars().first()
    old_price_val = product.price if product else 0

    # Применяем через Ozon API
    url = "https://api-seller.ozon.ru/v1/product/import/prices"
    payload = {"prices": [{
        "offer_id": offer_id,
        "price": str(new_price),
        "old_price": str(body.get("old_price", 0)),
        "min_price": str(body.get("min_price", 0)),
    }]}

    async with aiohttp.ClientSession() as session:
        async with session.post(url, json=payload, headers=get_active_headers()) as resp:
            data = await resp.json()
            result_item = data.get("result", [{}])[0]

            if result_item.get("updated"):
                # Обновляем кэш в БД
                if product:
                    product.price = new_price
                    await db.commit()

                # Пишем в историю
                history = PriceHistory(
                    offer_id=offer_id,
                    old_price=old_price_val,
                    new_price=new_price,
                    reason=body.get("reason", "manual"),
                    changed_by=user.get("sub", "unknown"),
                )
                db.add(history)
                await db.commit()
                return {"ok": True, "updated": True}
            else:
                errors = result_item.get("errors", [])
                return {"ok": False, "errors": errors}


@app.get("/api/repricer/history/{offer_id}")
async def api_price_history(
    offer_id: str,
    user: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    """История изменений цены по товару."""
    result = await db.execute(
        select(PriceHistory)
        .where(PriceHistory.offer_id == offer_id)
        .order_by(PriceHistory.changed_at.desc())
        .limit(50)
    )
    history = result.scalars().all()
    return {"history": [
        {
            "old_price": h.old_price,
            "new_price": h.new_price,
            "reason": h.reason,
            "changed_by": h.changed_by,
            "changed_at": h.changed_at.isoformat(),
        }
        for h in history
    ]}



# =====================================================================
# СЕБЕСТОИМОСТЬ
# =====================================================================

@app.get("/costs", response_class=HTMLResponse)
async def costs_page(request: Request, user: dict = Depends(require_admin)):
    return templates.TemplateResponse("costs.html", {
        "request": request,
        "user": user,
        "active_tab": "costs"
    })


@app.get("/api/costs/products")
async def api_costs_products(
    search: str = "",
    brand: str = "",
    page: int = 1,
    per_page: int = 100,
    direction: str = "",
    user: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    """Список товаров с себестоимостью."""
    from sqlalchemy import or_

    query = select(Product)
    count_q = select(func.count(Product.offer_id))

    filters = []
    if search:
        like = f"%{search}%"
        filters.append(or_(Product.offer_id.ilike(like), Product.name.ilike(like)))
    if brand:
        filters.append(Product.brand == brand)
    if direction == "uzspace":
        filters.append(Product.brand.ilike("UZSPACE"))
    elif direction == "sima":
        filters.append(Product.brand.notilike("UZSPACE"))

    for f in filters:
        query = query.where(f)
        count_q = count_q.where(f)

    query = query.order_by(Product.offer_id.asc())
    query = query.limit(per_page).offset((page - 1) * per_page)

    total_r = await db.execute(count_q)
    total = total_r.scalar()
    products_r = await db.execute(query)
    products = products_r.scalars().all()

    return {
        "total": total,
        "products": [
            {
                "offer_id": p.offer_id,
                "name": p.name,
                "brand": p.brand,
                "image_url": p.image_url,
                "cost_price": p.cost_price,
                "cost_updated_at": p.updated_at.isoformat() if p.updated_at else None,
                "price": p.price,
                "net_price": p.net_price,
                "commission_fbs_percent": p.commission_fbs_percent,
                "commission_fbs_logistics": p.commission_fbs_logistics,
            }
            for p in products
        ]
    }


@app.post("/api/costs/upload")
async def api_costs_upload(
    body: dict,
    user: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    """
    Массовое обновление себестоимости из Excel.
    body: {items: [{offer_id, cost_price}]}
    """
    items = body.get("items", [])
    if not items:
        raise HTTPException(400, "No items")

    updated = 0
    for item in items:
        offer_id = item.get("offer_id")
        cost_price = item.get("cost_price")
        if not offer_id or not cost_price:
            continue

        # Получаем текущую себестоимость для истории
        result = await db.execute(select(Product).where(Product.offer_id == offer_id))
        product = result.scalars().first()

        if product:
            old_cost = product.cost_price
            if old_cost != cost_price:
                # Пишем в историю
                history = CostHistory(
                    offer_id=offer_id,
                    old_cost=old_cost,
                    new_cost=cost_price,
                    source="excel_upload",
                    changed_by=user.get("sub", "unknown"),
                )
                db.add(history)
            product.cost_price = cost_price
            updated += 1
        else:
            # Создаём минимальную запись если товар ещё не синхронизирован
            product = Product(offer_id=offer_id, cost_price=cost_price)
            db.add(product)
            updated += 1

    await db.commit()
    return {"updated": updated}


@app.get("/api/costs/history/{offer_id}")
async def api_cost_history(
    offer_id: str,
    user: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    """История изменений себестоимости."""
    result = await db.execute(
        select(CostHistory)
        .where(CostHistory.offer_id == offer_id)
        .order_by(CostHistory.changed_at.desc())
        .limit(50)
    )
    history = result.scalars().all()
    return {"history": [
        {
            "old_cost": h.old_cost,
            "new_cost": h.new_cost,
            "source": h.source,
            "changed_by": h.changed_by,
            "changed_at": h.changed_at.isoformat(),
        }
        for h in history
    ]}


_ozon_sync_status = {"running": False, "fetched": 0, "total": 0, "saved": 0, "done": False}


@app.get("/api/costs/sync-ozon/status")
async def api_costs_sync_status(user: dict = Depends(require_any_role)):
    return _ozon_sync_status


@app.post("/api/costs/sync-ozon")
async def api_costs_sync_ozon(user: dict = Depends(require_admin)):
    asyncio.create_task(sync_products_catalog())
    return {"ok": True, "message": "Синхронизация запущена в фоне"}


async def _sync_ozon_task():
    global _ozon_sync_status
    _ozon_sync_status = {"running": True, "fetched": 0, "total": 0, "saved": 0, "done": False}
    try:
        print("SYNC OZON: started", flush=True)
        all_offer_ids = []
        last_id = ""
        limit = 1000

        while True:
            async with aiohttp.ClientSession() as session:
                async with session.post(
                    "https://api-seller.ozon.ru/v3/product/list",
                    headers=get_active_headers(),
                    json={"filter": {"visibility": "ALL"}, "last_id": last_id, "limit": limit}
                ) as resp:
                    data = await resp.json()

            result = data.get("result", {})
            items = result.get("items", [])
            if not items:
                break

            all_offer_ids.extend(items)
            last_id = result.get("last_id", "")
            _ozon_sync_status["fetched"] = len(all_offer_ids)
            print(f"SYNC OZON: fetched {len(all_offer_ids)}", flush=True)

            if not last_id or len(items) < limit:
                break

        _ozon_sync_status["total"] = len(all_offer_ids)
        print(f"SYNC OZON: total {len(all_offer_ids)}, saving...", flush=True)

        batch_size = 500
        saved = 0
        async with AsyncSessionLocal() as db:
            for i in range(0, len(all_offer_ids), batch_size):
                batch = all_offer_ids[i:i + batch_size]
                for item in batch:
                    offer_id = str(item.get("offer_id", "")).strip()
                    product_id = item.get("product_id")
                    if not offer_id:
                        continue
                    stmt = pg_insert(Product).values(
                        offer_id=offer_id,
                        product_id=product_id,
                    ).on_conflict_do_nothing(index_elements=["offer_id"])
                    await db.execute(stmt)
                    saved += 1
                await db.commit()
                _ozon_sync_status["saved"] = saved
                print(f"SYNC OZON: saved {saved}", flush=True)

        _ozon_sync_status = {"running": False, "fetched": len(all_offer_ids), "total": len(all_offer_ids), "saved": saved, "done": True}
        print(f"SYNC OZON: done", flush=True)

    except Exception as e:
        _ozon_sync_status = {"running": False, "fetched": 0, "total": 0, "saved": 0, "done": False, "error": str(e)}
        print(f"SYNC OZON ERROR: {e}", flush=True)
        import traceback
        traceback.print_exc()



@app.post("/api/sync")
async def api_sync(user: dict = Depends(require_admin)):
    """
    Синхронизация с Ozon API по запросу пользователя.
    Вызывается кнопкой 'Обновить' на сайте.
    """
    result = await sync_from_ozon()
    return result


@app.get("/api/sync/status")
async def api_sync_status(user: dict = Depends(require_admin)):
    """Возвращает время последней синхронизации."""
    return {
        "last_sync": _last_sync.strftime("%d.%m.%Y %H:%M") if _last_sync else None
    }


@app.get("/users", response_class=HTMLResponse)
async def users_page(request: Request, user: dict = Depends(require_admin)):
    """Страница управления пользователями — только для admin"""
    return templates.TemplateResponse("users.html", {
        "request": request,
        "user": user,
        "active_tab": "users"
    })


@app.get("/repricer", response_class=HTMLResponse)
async def repricer_page(request: Request, user: dict = Depends(require_admin)):
    return templates.TemplateResponse("repricer.html", {
        "request": request,
        "user": user,
        "active_tab": "repricer"
    })


@app.get("/api/users")
async def get_users(
    user: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    """
    Список всех пользователей системы.
    📚 УРОК: Никогда не возвращаем password_hash клиенту — только нужные поля.
    """
    result = await db.execute(select(User).order_by(User.id))
    users = result.scalars().all()
    return {
        "users": [
            {
                "id": u.id,
                "username": u.username,
                "role": u.role,
                # created_at нет в модели — добавим None, фронт обработает
                "created_at": None,
                "permissions": (u.permissions if isinstance(u.permissions, list) else json.loads(u.permissions)) if u.permissions else (["dashboard","orders","queue","users","repricer","costs"] if u.role == "admin" else ["queue"]),
            }
            for u in users
        ]
    }

@app.patch("/api/users/{user_id}/permissions")
async def update_user_permissions(
    user_id: int,
    request: Request,
    current_user: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    body = await request.json()
    permissions = body.get("permissions", [])
    import json as _json
    result = await db.execute(select(User).where(User.id == user_id))
    u = result.scalars().first()
    if not u:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    from sqlalchemy import cast
    from sqlalchemy.dialects.postgresql import JSONB
    await db.execute(
        update(User).where(User.id == user_id).values(permissions=cast(_json.dumps(permissions), JSONB))
    )
    await db.commit()
    return {"ok": True, "permissions": permissions}

# =====================================================================
# АНАЛИТИКА ПРОДАЖ
# =====================================================================

async def sync_sales_from_ozon():
    """Синхронизирует историю продаж из Ozon API (delivered + cancelled)"""
    url = "https://api-seller.ozon.ru/v3/posting/fbs/list"
    date_to = datetime.now()
    date_from = date_to - timedelta(days=180)

    # Подтягиваем бренды и категории из таблицы products
    async with AsyncSessionLocal() as db:
        prod_res = await db.execute(select(Product.offer_id, Product.brand, Product.category_id, Product.category_name))
        prod_map = {r[0]: {"brand": r[1], "category_id": r[2], "category_name": r[3]} for r in prod_res.all()}

    statuses_map = {"delivered": "sale", "cancelled": "cancel"}

    async with aiohttp.ClientSession() as session:
        for ozon_status, sale_status in statuses_map.items():
            offset = 0
            while True:
                payload = {
                    "dir": "DESC",
                    "filter": {
                        "status": ozon_status,
                        "since": date_from.strftime("%Y-%m-%dT%H:%M:%S.000Z"),
                        "to": date_to.strftime("%Y-%m-%dT%H:%M:%S.000Z")
                    },
                    "limit": 1000,
                    "offset": offset,
                    "with": {"analytics_data": False, "financial_data": False}
                }
                try:
                    async with session.post(url, json=payload, headers=get_active_headers()) as resp:
                        if resp.status != 200:
                            break
                        data = await resp.json()
                        postings = data.get("result", {}).get("postings", [])
                        if not postings:
                            break

                        async with AsyncSessionLocal() as db:
                            for posting in postings:
                                pnum = posting.get("posting_number")
                                event_date = None
                                in_process = posting.get("in_process_at") or posting.get("shipment_date")
                                if in_process:
                                    try:
                                        event_date = datetime.fromisoformat(in_process.replace("Z", "+00:00")).replace(tzinfo=None)
                                    except:
                                        pass

                                for prod in posting.get("products", []):
                                    oid = prod.get("offer_id", "")
                                    qty = int(prod.get("quantity", 1))
                                    price = int(float(prod.get("price", 0)))
                                    pinfo = prod_map.get(oid, {})

                                    stmt = pg_insert(SalesHistory).values(
                                        posting_number=pnum,
                                        offer_id=oid,
                                        name=prod.get("name", ""),
                                        brand=pinfo.get("brand"),
                                        category_id=pinfo.get("category_id"),
                                        category_name=pinfo.get("category_name"),
                                        quantity=qty,
                                        price=price,
                                        revenue=price * qty,
                                        status=sale_status,
                                        event_date=event_date,
                                    ).on_conflict_do_nothing(index_elements=["posting_number", "offer_id"])
                                    await db.execute(stmt)
                            await db.commit()

                        if len(postings) < 1000:
                            break
                        offset += 1000
                except Exception as e:
                    print(f"sync_sales error: {e}", flush=True)
                    break


@app.get("/analytics", response_class=HTMLResponse)
async def analytics_page(request: Request, user: dict = Depends(require_any_role)):
    if "analytics" not in user.get("permissions", []) and user["role"] != "admin":
        return RedirectResponse("/queue", status_code=302)
    return templates.TemplateResponse("analytics.html", {
        "request": request,
        "user": user,
        "active_tab": "analytics"
    })


@app.post("/api/analytics/sync")
async def api_analytics_sync(user: dict = Depends(require_admin)):
    import asyncio
    asyncio.create_task(sync_sales_from_ozon())
    return {"ok": True}


@app.get("/api/analytics/sales")
async def api_analytics_sales(
    user: dict = Depends(require_any_role),
    db: AsyncSession = Depends(get_db),
    date_from: str = "",
    date_to: str = "",
    status: str = "",
    brand: str = "",
    category_id: str = "",
    search: str = "",
):
    filters = []
    if date_from:
        try:
            filters.append(SalesHistory.event_date >= datetime.fromisoformat(date_from))
        except: pass
    if date_to:
        try:
            filters.append(SalesHistory.event_date <= datetime.fromisoformat(date_to + "T23:59:59"))
        except: pass
    if status in ("sale", "cancel"):
        filters.append(SalesHistory.status == status)
    if brand:
        filters.append(SalesHistory.brand.ilike(f"%{brand}%"))
    if category_id:
        try:
            filters.append(SalesHistory.category_id == int(category_id))
        except: pass
    if search:
        filters.append(or_(
            SalesHistory.offer_id.ilike(f"%{search}%"),
            SalesHistory.name.ilike(f"%{search}%")
        ))

    # Топ товаров
    from sqlalchemy import func as sqlfunc
    top_q = select(
        SalesHistory.offer_id,
        SalesHistory.name,
        SalesHistory.brand,
        SalesHistory.category_name,
        sqlfunc.sum(SalesHistory.quantity).label("total_qty"),
        sqlfunc.sum(SalesHistory.revenue).label("total_revenue"),
        sqlfunc.count(SalesHistory.id).label("orders_count"),
    ).where(*filters).group_by(
        SalesHistory.offer_id, SalesHistory.name, SalesHistory.brand, SalesHistory.category_name
    ).order_by(sqlfunc.sum(SalesHistory.quantity).desc()).limit(200)

    top_res = await db.execute(top_q)
    top_items = top_res.all()

    # График по дням
    chart_q = select(
        sqlfunc.date_trunc('day', SalesHistory.event_date).label("day"),
        SalesHistory.status,
        sqlfunc.sum(SalesHistory.quantity).label("qty"),
        sqlfunc.sum(SalesHistory.revenue).label("revenue"),
    ).where(*filters).group_by("day", SalesHistory.status).order_by("day")

    chart_res = await db.execute(chart_q)
    chart_rows = chart_res.all()

    # Фильтры — бренды и категории
    brands_res = await db.execute(
        select(SalesHistory.brand).where(SalesHistory.brand.isnot(None)).distinct()
    )
    cats_res = await db.execute(
        select(SalesHistory.category_id, SalesHistory.category_name)
        .where(SalesHistory.category_id.isnot(None)).distinct()
    )

    return {
        "top": [
            {
                "offer_id": r.offer_id,
                "name": r.name or "",
                "brand": r.brand or "",
                "category_name": r.category_name or "",
                "total_qty": int(r.total_qty or 0),
                "total_revenue": int(r.total_revenue or 0),
                "orders_count": int(r.orders_count or 0),
            }
            for r in top_items
        ],
        "chart": [
            {
                "day": r.day.strftime("%Y-%m-%d") if r.day else None,
                "status": r.status,
                "qty": int(r.qty or 0),
                "revenue": int(r.revenue or 0),
            }
            for r in chart_rows
        ],
        "filters": {
            "brands": sorted([r[0] for r in brands_res.all() if r[0]]),
            "categories": [{"id": r[0], "name": r[1]} for r in cats_res.all()],
        }
    }

@app.get("/api/analytics/orders-chart")
async def api_analytics_orders_chart(
    user: dict = Depends(require_any_role),
    db: AsyncSession = Depends(get_db),
    date_from: str = "",
    date_to: str = "",
):
    from sqlalchemy import func as sqlfunc, cast, Date
    filters = [Order.ozon_accepted_at.isnot(None)]
    if date_from:
        try: filters.append(Order.ozon_accepted_at >= datetime.fromisoformat(date_from))
        except: pass
    if date_to:
        try: filters.append(Order.ozon_accepted_at <= datetime.fromisoformat(date_to + "T23:59:59"))
        except: pass

    q = select(
        sqlfunc.date_trunc('day', Order.ozon_accepted_at).label("day"),
        sqlfunc.count(Order.posting_number).label("orders_count"),
    ).where(*filters).group_by("day").order_by("day")

    res = await db.execute(q)
    return {
        "orders_chart": [
            {"day": r.day.strftime("%Y-%m-%d"), "count": int(r.orders_count)}
            for r in res.all()
        ]
    }

@app.get("/api/analytics/ozon-chart")
async def api_analytics_ozon_chart(
    user: dict = Depends(require_any_role),
    date_from: str = "",
    date_to: str = "",
):
    if not date_from:
        date_from = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    if not date_to:
        date_to = datetime.now().strftime("%Y-%m-%d")

    try:
        async with aiohttp.ClientSession() as session:
            async with session.post(
                "https://api-seller.ozon.ru/v1/analytics/data",
                json={
                    "date_from": date_from,
                    "date_to": date_to,
                    "metrics": ["ordered_units", "revenue"],
                    "dimension": ["day"],
                    "limit": 1000,
                    "offset": 0
                },
                headers=get_active_headers()
            ) as resp:
                if resp.status != 200:
                    return {"chart": []}
                data = await resp.json()

        chart = [
            {
                "day": row["dimensions"][0]["id"],
                "ordered_units": row["metrics"][0],
                "revenue": row["metrics"][1],
            }
            for row in data.get("result", {}).get("data", [])
        ]
        return {"chart": chart}
    except Exception as e:
        return {"chart": [], "error": str(e)}

@app.get("/api/analytics/top-by-orders")
async def api_analytics_top_by_orders(
    user: dict = Depends(require_any_role),
    db: AsyncSession = Depends(get_db),
    date_from: str = "",
    date_to: str = "",
):
    if not date_from:
        date_from = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    if not date_to:
        date_to = datetime.now().strftime("%Y-%m-%d")

    try:
        # Получаем топ по SKU из Ozon Analytics
        all_items = []
        offset = 0
        async with aiohttp.ClientSession() as session:
            while True:
                async with session.post(
                    "https://api-seller.ozon.ru/v1/analytics/data",
                    json={
                        "date_from": date_from,
                        "date_to": date_to,
                        "metrics": ["ordered_units", "revenue", "cancellations"],
                        "dimension": ["sku"],
                        "limit": 1000,
                        "offset": offset
                    },
                    headers=get_active_headers()
                ) as resp:
                    if resp.status != 200:
                        break
                    data = await resp.json()
                    rows = data.get("result", {}).get("data", [])
                    if not rows:
                        break
                    all_items.extend(rows)
                    if len(rows) < 1000:
                        break
                    offset += 1000

        if not all_items:
            return {"top": []}

        # Строим маппинг SKU → продукт из БД
        skus = [int(r["dimensions"][0]["id"]) for r in all_items]
        from sqlalchemy import or_
        prod_res = await db.execute(
            select(Product.offer_id, Product.name, Product.brand,
                   Product.category_name, Product.sku_fbs, Product.sku_sds)
            .where(or_(Product.sku_fbs.in_(skus), Product.sku_sds.in_(skus)))
        )
        sku_map = {}
        for p in prod_res.all():
            info = {"offer_id": p.offer_id, "name": p.name, "brand": p.brand, "category_name": p.category_name}
            if p.sku_fbs: sku_map[p.sku_fbs] = info
            if p.sku_sds: sku_map[p.sku_sds] = info

        top = []
        for r in all_items:
            sku = int(r["dimensions"][0]["id"])
            name_from_ozon = r["dimensions"][0]["name"]
            prod = sku_map.get(sku, {})
            top.append({
                "offer_id": prod.get("offer_id", f"sku:{sku}"),
                "name": prod.get("name") or name_from_ozon,
                "brand": prod.get("brand", ""),
                "category_name": prod.get("category_name", ""),
                "ordered_units": int(r["metrics"][0]),
                "revenue": int(r["metrics"][1]),
                "cancellations": int(r["metrics"][2]),
            })

            # Для неизвестных SKU — подтягиваем offer_id из Ozon
            unknown_skus = [int(r["dimensions"][0]["id"]) for r in all_items
                            if int(r["dimensions"][0]["id"]) not in sku_map]

            if unknown_skus:
                try:
                    async with aiohttp.ClientSession() as session:
                        for i in range(0, len(unknown_skus), 100):
                            batch = unknown_skus[i:i + 100]
                            async with session.post(
                                    "https://api-seller.ozon.ru/v3/product/info/list",
                                    json={"sku": batch},
                                    headers=get_active_headers()
                            ) as resp:
                                if resp.status == 200:
                                    d = await resp.json()
                                    for it in d.get("items", []):
                                        for src in it.get("sources", []):
                                            s = src.get("sku")
                                            if s in unknown_skus:
                                                sku_map[s] = {
                                                    "offer_id": it.get("offer_id", f"sku:{s}"),
                                                    "name": it.get("name", ""),
                                                    "brand": "",
                                                    "category_name": ""
                                                }
                except Exception:
                    pass

            top = []
            for r in all_items:
                sku = int(r["dimensions"][0]["id"])
                name_from_ozon = r["dimensions"][0]["name"]
                prod = sku_map.get(sku, {})
                top.append({
                    "offer_id": prod.get("offer_id", f"sku:{sku}"),
                    "name": prod.get("name") or name_from_ozon,
                    "brand": prod.get("brand", ""),
                    "category_name": prod.get("category_name", ""),
                    "ordered_units": int(r["metrics"][0]),
                    "revenue": int(r["metrics"][1]),
                    "cancellations": int(r["metrics"][2]),
                })

            top.sort(key=lambda x: x["ordered_units"], reverse=True)
            return {"top": top}

    except Exception as e:
        return {"top": [], "error": str(e)}


@app.get("/stock", response_class=HTMLResponse)
async def stock_page(request: Request, user: dict = Depends(require_any_role)):
    if "stock" not in user.get("permissions", []) and user["role"] != "admin":
        return RedirectResponse("/queue", status_code=302)
    return templates.TemplateResponse("stock.html", {
        "request": request, "user": user, "active_tab": "stock"
    })

# Кэш остатков FBO
_stock_cache = {"items": [], "warehouses": [], "total": 0, "loading": False, "updated_at": None, "watchlist": []}

async def sync_stock_cache():
    global _stock_cache
    if _stock_cache["loading"]:
        return
    _stock_cache["loading"] = True
    print("STOCK SYNC STARTED", flush=True)
    try:
        timeout = aiohttp.ClientTimeout(total=120)
        async with aiohttp.ClientSession(timeout=timeout) as session:

            async def fetch_fbo():
                rows = []
                offset = 0
                while True:
                    async with session.post(
                        "https://api-seller.ozon.ru/v2/analytics/stock_on_warehouses",
                        json={"limit": 1000, "offset": offset, "warehouse_type": "ALL"},
                        headers=get_active_headers()
                    ) as resp:
                        if resp.status != 200:
                            print(f"STOCK FBO error status: {resp.status}", flush=True)
                            break
                        data = await resp.json()
                        batch = data.get("result", {}).get("rows", [])
                        if not batch:
                            break
                        rows.extend(batch)
                        print(f"STOCK FBO loaded: {len(rows)}", flush=True)
                        if len(batch) < 1000:
                            break
                        offset += 1000
                return rows

            async def fetch_fbs():
                result = {}
                cursor = ""
                while True:
                    payload = {"limit": 1000, "filter": {}}
                    if cursor:
                        payload["cursor"] = cursor
                    else:
                        payload["offset"] = 0
                    async with session.post(
                        "https://api-seller.ozon.ru/v4/product/info/stocks",
                        json=payload,
                        headers=get_active_headers()
                    ) as resp:
                        if resp.status != 200:
                            print(f"STOCK FBS error status: {resp.status}", flush=True)
                            break
                        data = await resp.json()
                        items = data.get("items", [])
                        if not items:
                            break
                        for item in items:
                            oid = item.get("offer_id", "")
                            result[oid] = sum(
                                s.get("present", 0) for s in item.get("stocks", [])
                                if s.get("type") == "fbs"
                            )
                        print(f"STOCK FBS loaded: {len(result)}", flush=True)
                        cursor = data.get("cursor", "")
                        if not cursor or len(items) < 1000:
                            break
                return result

            print("STOCK: starting parallel fetch", flush=True)
            import asyncio as _asyncio
            all_rows, fbs_map = await _asyncio.gather(fetch_fbo(), fetch_fbs())
            print(f"STOCK: FBO={len(all_rows)} FBS={len(fbs_map)}", flush=True)

        from collections import defaultdict
        async with AsyncSessionLocal() as db:
            codes = list(set(r.get("item_code", "") for r in all_rows))
            prod_res = await db.execute(
                select(Product.offer_id, Product.brand).where(Product.offer_id.in_(codes))
            )
            brand_map = {r.offer_id: r.brand or "" for r in prod_res.all()}

        by_item = defaultdict(lambda: {
            "item_code": "", "item_name": "", "brand": "",
            "total_free": 0, "total_reserved": 0, "total_promised": 0,
            "fbs_present": 0, "warehouses": {}
        })

        for row in all_rows:
            code = row.get("item_code", "")
            brand = brand_map.get(code, "")
            if brand.upper() == "UZSPACE":
                continue
            wh = row.get("warehouse_name", "")
            free = int(row.get("free_to_sell_amount", 0))
            reserved = int(row.get("reserved_amount", 0))
            promised = int(row.get("promised_amount", 0))
            by_item[code]["item_code"] = code
            by_item[code]["item_name"] = row.get("item_name", "")
            by_item[code]["brand"] = brand
            by_item[code]["total_free"] += free
            by_item[code]["total_reserved"] += reserved
            by_item[code]["total_promised"] += promised
            by_item[code]["fbs_present"] = fbs_map.get(code, 0)
            by_item[code]["warehouses"][wh] = {"free": free, "reserved": reserved, "promised": promised}

        items = sorted(by_item.values(), key=lambda x: x["total_free"], reverse=True)
        warehouses = sorted(set(row.get("warehouse_name", "") for row in all_rows))

        print(f"STOCK SYNC DONE: {len(items)} items", flush=True)
        # Обновляем watchlist
        async with AsyncSessionLocal() as db:
            # Добавляем в watchlist только когда FBO=0 и FBS=0 (пора подключать)
            for item in by_item.values():
                if item["total_free"] == 0 and item["fbs_present"] == 0:
                    stmt = pg_insert(FboWatchlist).values(
                        offer_id=item["item_code"],
                        item_name=item["item_name"],
                        brand=item["brand"],
                    ).on_conflict_do_nothing(index_elements=["offer_id"])
                    await db.execute(stmt)

            # Автоудаляем из watchlist если FBS снова появился
            for item in by_item.values():
                if item["fbs_present"] > 0:
                    await db.execute(
                        delete(FboWatchlist)
                        .where(FboWatchlist.offer_id == item["item_code"])
                    )

            await db.commit()

            # Загружаем watchlist для ответа
            wl_res = await db.execute(
                select(FboWatchlist)
                .where(FboWatchlist.dismissed_at == None)
                .order_by(FboWatchlist.noted_at.desc().nullslast())
            )
            watchlist = [
                {"offer_id": w.offer_id, "item_name": w.item_name, "brand": w.brand,
                 "noted_at": w.noted_at.isoformat() if w.noted_at else None}
                for w in wl_res.scalars().all()
            ]

        _stock_cache["watchlist"] = watchlist
        _stock_cache["items"] = items
        _stock_cache["warehouses"] = warehouses
        _stock_cache["total"] = len(items)
        _stock_cache["updated_at"] = datetime.now().strftime("%d.%m.%Y %H:%M")

    except Exception as e:
        import traceback
        print(f"STOCK SYNC ERROR: {traceback.format_exc()}", flush=True)
    finally:
        _stock_cache["loading"] = False

@app.get("/api/stock/fbo")
async def api_stock_fbo(user: dict = Depends(require_any_role)):
    import asyncio as _asyncio
    if not _stock_cache["items"] and not _stock_cache["loading"]:
        _asyncio.create_task(sync_stock_cache())
    return {
        "items": _stock_cache["items"],
        "warehouses": _stock_cache["warehouses"],
        "total": _stock_cache["total"],
        "loading": _stock_cache["loading"],
        "updated_at": _stock_cache["updated_at"],
        "watchlist": _stock_cache.get("watchlist", []),
    }

@app.post("/api/stock/fbo/sync")
async def api_stock_fbo_sync(user: dict = Depends(require_any_role)):
    import asyncio as _asyncio
    _asyncio.create_task(sync_stock_cache())
    return {"ok": True}

# =====================================================================
# FBO — ПЛАТНОЕ ХРАНЕНИЕ (v3 — загрузка xlsx вручную)
# =====================================================================

import io
from openpyxl import load_workbook

# Кэш данных о хранении из загруженного файла
_storage_report_cache = {
    "data": {},        # offer_id -> {days_left, paid_date, free_qty, paid_qty, daily_cost_28d}
    "updated_at": None,
    "filename": None,
}

async def load_storage_report_from_db():
    """Загружает данные хранения из БД в кэш при старте приложения."""
    global _storage_report_cache
    try:
        async with AsyncSessionLocal() as session:
            rows = await session.execute(select(FboStorageReport))
            items = rows.scalars().all()
            if not items:
                return
            data = {}
            max_updated = None
            for row in items:
                data[row.offer_id] = {
                    "days_left": row.days_left,
                    "paid_date": row.paid_date,
                    "free_qty": row.free_qty,
                    "paid_qty": row.paid_qty,
                    "daily_cost_28d": row.daily_cost_28d,
                }
                if row.updated_at and (max_updated is None or row.updated_at > max_updated):
                    max_updated = row.updated_at
            _storage_report_cache["data"] = data
            if max_updated:
                _storage_report_cache["updated_at"] = max_updated.strftime("%d.%m.%Y %H:%M")
            print(f"FBO STORAGE: loaded {len(data)} rows from DB", flush=True)
    except Exception as e:
        print(f"FBO STORAGE load from DB error: {e}", flush=True)

# Кэш отслеживаемых позиций продаж FBO
_fbo_sales_watch_cache = {}  # offer_id -> {item_name, fbo_snapshot}


async def load_fbo_sales_watch_from_db():
    """Загружает список отслеживаемых позиций из БД в кэш."""
    global _fbo_sales_watch_cache
    try:
        async with AsyncSessionLocal() as session:
            rows = await session.execute(select(FboSalesWatch))
            items = rows.scalars().all()
            _fbo_sales_watch_cache = {
                row.offer_id: {
                    "item_name": row.item_name,
                    "fbo_snapshot": row.fbo_snapshot,
                }
                for row in items
            }
            print(f"FBO SALES WATCH: loaded {len(_fbo_sales_watch_cache)} items from DB", flush=True)
    except Exception as e:
        print(f"FBO SALES WATCH load error: {e}", flush=True)


async def _tg_send_or_edit(bot_token: str, chat_id: str, proxy_url: str, text: str, message_id: Optional[int]) -> Optional[int]:
    """Отправляет новое сообщение или редактирует существующее (если message_id задан и правка прошла).
    Возвращает id итогового сообщения, либо None при ошибке."""
    from aiohttp_socks import ProxyConnector
    connector = ProxyConnector.from_url(proxy_url)
    async with aiohttp.ClientSession(connector=connector) as tg_session:
        if message_id:
            try:
                async with tg_session.post(
                    f"https://api.telegram.org/bot{bot_token}/editMessageText",
                    json={"chat_id": chat_id, "message_id": message_id, "text": text, "parse_mode": "HTML"},
                    timeout=aiohttp.ClientTimeout(total=15)
                ) as r:
                    data = await r.json(content_type=None)
                if data.get("ok"):
                    return message_id
            except Exception as e:
                print(f"FBO DIGEST edit error: {e}", flush=True)
        try:
            async with tg_session.post(
                f"https://api.telegram.org/bot{bot_token}/sendMessage",
                json={"chat_id": chat_id, "text": text, "parse_mode": "HTML"},
                timeout=aiohttp.ClientTimeout(total=15)
            ) as r:
                data = await r.json(content_type=None)
            if data.get("ok"):
                return data["result"]["message_id"]
            print(f"FBO DIGEST send failed: {data}", flush=True)
        except Exception as e:
            print(f"FBO DIGEST send error: {e}", flush=True)
    return None


async def _update_fbo_sold_out_digest(session: AsyncSession, bot_token: str, chat_id: str, proxy_url: str,
                                       offer_id: str, name: str, days_str: str):
    """Добавляет полностью распроданную позицию в накопительный список за сегодня
    и обновляет (редактирует) единое Telegram-сообщение с этим списком.
    Список привязан к дате (Europe/Moscow) — на следующие сутки начинается новое сообщение."""
    today = datetime.now().strftime("%Y-%m-%d")
    row = await session.get(FboDailyDigest, today)
    items = json.loads(row.items_json) if row else []
    items.append({"offer_id": offer_id, "name": name, "days_str": days_str})

    date_display = datetime.now().strftime("%d.%m.%Y")
    count = len(items)
    SHOWN_LIMIT = 40
    shown = items[-SHOWN_LIMIT:]
    lines = [f"✅ <b>Распродано за {date_display} — {count} шт.</b>"]
    if count > len(shown):
        lines.append(f"<i>(показаны последние {len(shown)})</i>")
    for idx, it in enumerate(shown, start=count - len(shown) + 1):
        lines.append(f"{idx}. {it['name']}\nАртикул: <code>{it['offer_id']}</code> | до платного было: {it['days_str']}")
    lines.append("👉 Проверьте акции и включите FBS для этих позиций.")
    text = "\n\n".join(lines)

    message_id = await _tg_send_or_edit(bot_token, chat_id, proxy_url, text, row.message_id if row else None)

    values = {"items_json": json.dumps(items, ensure_ascii=False), "message_id": message_id}
    await session.execute(
        pg_insert(FboDailyDigest).values(date=today, **values)
        .on_conflict_do_update(index_elements=["date"], set_=values)
    )
    await session.commit()


async def check_fbo_sales_and_notify():
    """Проверяет остатки отслеживаемых позиций.
    Полностью распроданные (остаток 0) копятся в единый дневной дайджест (Telegram-сообщение
    редактируется при каждой новой распродаже, на следующие сутки — новое сообщение).
    Частичные продажи отправляются как раньше — отдельным сообщением на каждую продажу."""
    if not _fbo_sales_watch_cache:
        return
    if not _stock_cache["items"]:
        return

    stock_map = {i["item_code"]: i["total_free"] for i in _stock_cache["items"]}
    storage_map = _storage_report_cache.get("data", {})

    bot_token = os.getenv("TELEGRAM_BOT_TOKEN", "") or os.getenv("BOT_TOKEN", "")
    chat_id = os.getenv("TELEGRAM_CHAT_ID", "") or os.getenv("ADMIN_ID", "")
    if not bot_token or not chat_id:
        print("FBO SALES WATCH: no telegram credentials", flush=True)
        return
    proxy_url = os.getenv("PROXY_URL", "socks5://quicknode-tg-proxy:1080")

    updates = {}
    async with AsyncSessionLocal() as session:
        for offer_id, watch in list(_fbo_sales_watch_cache.items()):
            current_fbo = stock_map.get(offer_id, 0)
            prev_fbo = watch["fbo_snapshot"]

            if current_fbo < prev_fbo:
                sold = prev_fbo - current_fbo
                days_left = storage_map.get(offer_id, {}).get("days_left")
                days_str = f"{days_left} дн." if days_left is not None else "—"
                name = watch["item_name"] or offer_id

                try:
                    if current_fbo == 0:
                        await _update_fbo_sold_out_digest(session, bot_token, chat_id, proxy_url, offer_id, name, days_str)
                        print(f"FBO SALES WATCH: digest updated for {offer_id} sold={sold}", flush=True)
                    else:
                        msg = (
                            f"📦 Продажа FBO\n"
                            f"<b>{name}</b>\n"
                            f"Артикул: <code>{offer_id}</code>\n"
                            f"Продано: {sold} шт → остаток: {current_fbo}\n"
                            f"До платного: {days_str}"
                        )
                        await _tg_send_or_edit(bot_token, chat_id, proxy_url, msg, None)
                        print(f"FBO SALES WATCH: notified {offer_id} sold={sold} left={current_fbo}", flush=True)
                except Exception as e:
                    print(f"FBO SALES WATCH telegram error: {e}", flush=True)

                # Обновляем snapshot в БД и кэше
                updates[offer_id] = current_fbo
                _fbo_sales_watch_cache[offer_id]["fbo_snapshot"] = current_fbo

        # Сохраняем обновлённые снапшоты
        for oid, new_fbo in updates.items():
            await session.execute(
                update(FboSalesWatch)
                .where(FboSalesWatch.offer_id == oid)
                .values(fbo_snapshot=new_fbo)
            )
        if updates:
            await session.commit()


@app.get("/fbo-storage", response_class=HTMLResponse)
async def fbo_storage_page(request: Request, user: dict = Depends(require_any_role)):
    if "stock" not in user.get("permissions", []) and user["role"] != "admin":
        return RedirectResponse("/queue", status_code=302)
    return templates.TemplateResponse("fbo_storage.html", {
        "request": request, "user": user, "active_tab": "fbo-storage"
    })


@app.get("/api/fbo-storage/items")
async def api_fbo_storage_items(user: dict = Depends(require_any_role)):
    """Остатки FBO + данные хранения из последнего загруженного файла."""
    import asyncio as _asyncio
    if not _stock_cache["items"] and not _stock_cache["loading"]:
        _asyncio.create_task(sync_stock_cache())
        return {"items": [], "loading": True,
                "report_updated_at": _storage_report_cache.get("updated_at"),
                "report_filename": _storage_report_cache.get("filename")}

    if _stock_cache["loading"]:
        return {"items": [], "loading": True,
                "report_updated_at": _storage_report_cache.get("updated_at"),
                "report_filename": _storage_report_cache.get("filename")}

    stock_items = [i for i in _stock_cache["items"] if i["total_free"] > 0]
    storage = _storage_report_cache["data"]

    result = []
    for item in stock_items:
        oid = item["item_code"]
        s = storage.get(oid, {})
        result.append({
            "offer_id": oid,
            "name": item["item_name"],
            "brand": item.get("brand", ""),
            "fbo": item["total_free"],
            "fbo_reserved": item.get("total_reserved", 0),
            "days_left": s.get("days_left"),
            "paid_date": s.get("paid_date"),
            "free_qty": s.get("free_qty", 0),
            "paid_qty": s.get("paid_qty", 0),
            "daily_cost_28d": s.get("daily_cost_28d", 0.0),
            "warehouses": item.get("warehouses", {}),
        })

    result.sort(key=lambda x: (
        x["days_left"] is None,
        x["days_left"] if x["days_left"] is not None else 9999
    ))

    return {
        "items": result,
        "loading": False,
        "updated_at": _stock_cache.get("updated_at"),
        "report_updated_at": _storage_report_cache.get("updated_at"),
        "report_filename": _storage_report_cache.get("filename"),
    }


@app.post("/api/fbo-storage/upload")
async def api_fbo_storage_upload(
    file: UploadFile = File(...),
    user: dict = Depends(require_any_role)
):
    """
    Принимает xlsx-файл отчёта Ozon 'Стоимость размещения по товарам'.
    Парсит и сохраняет в кэш.
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Нужен файл .xlsx")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active

        # Определяем формат файла по первой строке
        first_row = next(ws.iter_rows(max_row=1, values_only=True))
        headers = [str(c).strip() if c else "" for c in first_row]

        # Сводный отчёт: заголовки в строках 1+2 (merged), данные с 3-й строки
        # Ищем колонки по названию
        col_offer_id = None
        col_days_left = None
        col_paid_date = None
        col_free_qty = None
        col_paid_qty = None
        col_cost_28d = None

        # Читаем первые 3 строки чтобы найти заголовки
        all_headers = []
        for row in ws.iter_rows(max_row=3, values_only=True):
            all_headers.append([str(c).strip() if c else "" for c in row])

        # Ищем во всех строках заголовков
        for row_idx, hrow in enumerate(all_headers):
            for col_idx, h in enumerate(hrow):
                hl = h.lower()
                if "артикул" in hl and col_offer_id is None:
                    col_offer_id = col_idx
                if "дней до" in hl and "платн" in hl:
                    col_days_left = col_idx
                if "дата" in hl and "платн" in hl:
                    col_paid_date = col_idx
                if "бесплатно" in hl and "шт" in hl:
                    col_free_qty = col_idx
                if "платно" in hl and "шт" in hl and col_paid_qty is None:
                    col_paid_qty = col_idx
                if "списано" in hl and "28" in hl or ("списано" in hl and col_cost_28d is None):
                    col_cost_28d = col_idx

        print(f"FBO UPLOAD: cols offer={col_offer_id} days={col_days_left} date={col_paid_date} free={col_free_qty} paid={col_paid_qty} cost={col_cost_28d}", flush=True)

        if col_offer_id is None:
            raise HTTPException(status_code=400, detail="Не найдена колонка 'Артикул'. Загрузите файл 'Стоимость размещения по товарам' из раздела FBO Ozon.")

        new_data = {}
        row_count = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            offer_id = str(row[col_offer_id]).strip() if row[col_offer_id] else None
            if not offer_id or offer_id in ("None", "Артикул", ""):
                continue

            days_left = row[col_days_left] if col_days_left is not None else None
            paid_date = row[col_paid_date] if col_paid_date is not None else None
            free_qty = row[col_free_qty] if col_free_qty is not None else None
            paid_qty = row[col_paid_qty] if col_paid_qty is not None else None
            cost_28d = row[col_cost_28d] if col_cost_28d is not None else None

            paid_date_str = None
            if paid_date:
                try:
                    if hasattr(paid_date, 'strftime'):
                        paid_date_str = paid_date.strftime("%Y-%m-%d")
                    else:
                        paid_date_str = str(paid_date)[:10]
                except Exception:
                    pass

            new_data[offer_id] = {
                "days_left": int(days_left) if days_left is not None else None,
                "paid_date": paid_date_str,
                "free_qty": int(free_qty) if free_qty else 0,
                "paid_qty": int(paid_qty) if paid_qty else 0,
                "daily_cost_28d": float(cost_28d) if cost_28d else 0.0,
            }
            row_count += 1

        if row_count == 0:
            raise HTTPException(status_code=400, detail="Файл не содержит данных. Проверьте формат.")

            # Сохраняем в БД — полная перезапись
        async with AsyncSessionLocal() as session:
            from sqlalchemy import delete as sa_delete
            await session.execute(sa_delete(FboStorageReport))
            for oid, vals in new_data.items():
                session.add(FboStorageReport(
                    offer_id=oid,
                    days_left=vals["days_left"],
                    paid_date=vals["paid_date"],
                    free_qty=vals["free_qty"],
                    paid_qty=vals["paid_qty"],
                    daily_cost_28d=vals["daily_cost_28d"],
                ))
            await session.commit()

        _storage_report_cache["data"] = new_data
        _storage_report_cache["updated_at"] = datetime.now().strftime("%d.%m.%Y %H:%M")
        _storage_report_cache["filename"] = file.filename

        print(f"FBO UPLOAD: parsed {row_count} rows from {file.filename}, saved to DB", flush=True)
        return {"ok": True, "rows": row_count, "filename": file.filename}

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"FBO UPLOAD ERROR: {traceback.format_exc()}", flush=True)
        raise HTTPException(status_code=500, detail=f"Ошибка парсинга: {str(e)}")

@app.get("/api/fbo-storage/sales-watch")
async def api_fbo_sales_watch_list(user: dict = Depends(require_any_role)):
    """Список отслеживаемых позиций."""
    return {"items": list(_fbo_sales_watch_cache.keys())}


@app.post("/api/fbo-storage/sales-watch/{offer_id}")
async def api_fbo_sales_watch_add(offer_id: str, user: dict = Depends(require_any_role)):
    """Добавить позицию в отслеживание."""
    # Берём текущий остаток из кэша
    stock_item = next((i for i in _stock_cache["items"] if i["item_code"] == offer_id), None)
    fbo_now = stock_item["total_free"] if stock_item else 0
    item_name = stock_item["item_name"] if stock_item else offer_id

    async with AsyncSessionLocal() as session:
        from sqlalchemy.dialects.postgresql import insert as pg_insert
        await session.execute(
            pg_insert(FboSalesWatch).values(
                offer_id=offer_id,
                item_name=item_name,
                fbo_snapshot=fbo_now,
            ).on_conflict_do_update(
                index_elements=["offer_id"],
                set_={"fbo_snapshot": fbo_now, "item_name": item_name}
            )
        )
        await session.commit()

    _fbo_sales_watch_cache[offer_id] = {"item_name": item_name, "fbo_snapshot": fbo_now}
    return {"ok": True, "offer_id": offer_id, "fbo_snapshot": fbo_now}


@app.delete("/api/fbo-storage/sales-watch/{offer_id}")
async def api_fbo_sales_watch_remove(offer_id: str, user: dict = Depends(require_any_role)):
    """Убрать позицию из отслеживания."""
    async with AsyncSessionLocal() as session:
        await session.execute(
            delete(FboSalesWatch).where(FboSalesWatch.offer_id == offer_id)
        )
        await session.commit()

    _fbo_sales_watch_cache.pop(offer_id, None)
    return {"ok": True}


@app.post("/api/fbo-storage/test-notify")
async def api_fbo_test_notify(user: dict = Depends(require_any_role)):
    """Тестовое уведомление — имитирует продажу первого отслеживаемого товара."""
    if not _fbo_sales_watch_cache:
        return {"ok": False, "message": "Нет отслеживаемых позиций"}

    offer_id = list(_fbo_sales_watch_cache.keys())[0]
    watch = _fbo_sales_watch_cache[offer_id]
    storage = _storage_report_cache.get("data", {}).get(offer_id, {})
    days_left = storage.get("days_left")
    days_str = f"{days_left} дн." if days_left is not None else "—"

    msg = (
        f"📦 Продажа FBO (ТЕСТ)\n"
        f"<b>{watch['item_name'] or offer_id}</b>\n"
        f"Артикул: <code>{offer_id}</code>\n"
        f"Продано: 1 шт → остаток: {max(0, watch['fbo_snapshot'] - 1)}\n"
        f"До платного: {days_str}"
    )

    bot_token = os.getenv("TELEGRAM_BOT_TOKEN", "") or os.getenv("BOT_TOKEN", "")
    chat_id = os.getenv("TELEGRAM_CHAT_ID", "") or os.getenv("ADMIN_ID", "")

    try:
        from aiohttp_socks import ProxyConnector
        proxy_url = os.getenv("PROXY_URL", "socks5://quicknode-tg-proxy:1080")
        connector = ProxyConnector.from_url(proxy_url)
        async with aiohttp.ClientSession(connector=connector) as tg_session:
            resp = await tg_session.post(
                f"https://api.telegram.org/bot{bot_token}/sendMessage",
                json={"chat_id": chat_id, "text": msg, "parse_mode": "HTML"},
                timeout=aiohttp.ClientTimeout(total=15)
            )
            result = await resp.json()
        return {"ok": True, "offer_id": offer_id, "tg": result.get("ok")}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/api/fbo-storage/actions")
async def api_fbo_storage_actions(user: dict = Depends(require_any_role)):
    """Список доступных акций Ozon."""
    try:
        timeout = aiohttp.ClientTimeout(total=30)
        async with aiohttp.ClientSession(timeout=timeout) as session:
            async with session.get(
                "https://api-seller.ozon.ru/v1/actions",
                headers=get_active_headers()
            ) as resp:
                if resp.status != 200:
                    text = await resp.text()
                    raise HTTPException(status_code=502, detail=f"Ozon API error: {text}")
                data = await resp.json()
                actions = [
                    {
                        "id": a.get("id"),
                        "title": a.get("title"),
                        "action_type": a.get("action_type"),
                        "date_start": a.get("date_start"),
                        "date_end": a.get("date_end"),
                        "potential_products_count": a.get("potential_products_count", 0),
                        "participating_products_count": a.get("participating_products_count", 0),
                    }
                    for a in data.get("result", [])
                ]
                return {"actions": actions}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/fbo-storage/action-products/{action_id}")
async def api_fbo_storage_action_products(action_id: int, user: dict = Depends(require_any_role)):
    """Кандидаты для конкретной акции."""
    try:
        timeout = aiohttp.ClientTimeout(total=30)
        async with aiohttp.ClientSession(timeout=timeout) as session:
            all_products = []
            offset = 0
            while True:
                async with session.post(
                    "https://api-seller.ozon.ru/v1/actions/candidates",
                    json={"action_id": action_id, "limit": 100, "offset": offset},
                    headers=get_active_headers()
                ) as resp:
                    if resp.status != 200:
                        break
                    data = await resp.json()
                    products = data.get("result", {}).get("products", [])
                    all_products.extend(products)
                    if len(products) < 100:
                        break
                    offset += 100
            return {"products": all_products}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/fbo-storage/product-actions/{offer_id}")
async def api_fbo_storage_product_actions(offer_id: str, user: dict = Depends(require_any_role)):
    """Акции в которых товар является кандидатом или уже участвует."""
    import asyncio as _asyncio
    from sqlalchemy import select as sa_select

    product_id_db = None
    async with AsyncSessionLocal() as session:
        row = await session.execute(
            sa_select(Product.product_id).where(Product.offer_id == offer_id)
        )
        result = row.scalar_one_or_none()
        if result:
            product_id_db = int(result)

    def find_in_products(products, offer_id, product_id_db):
        for p in products:
            if (str(p.get("offer_id", "")) == offer_id
                    or (product_id_db and int(p.get("id", 0)) == product_id_db)):
                return p
        return None

    try:
        timeout = aiohttp.ClientTimeout(total=180)
        async with aiohttp.ClientSession(timeout=timeout) as session:
            async with session.get(
                "https://api-seller.ozon.ru/v1/actions",
                headers=get_active_headers()
            ) as resp:
                data = await resp.json()
                all_actions = {a["id"]: a for a in data.get("result", [])}

            result = []
            for action_id, action in all_actions.items():
                await _asyncio.sleep(0.3)
                found_product = None
                is_participating = False

                # Сначала ищем среди участвующих — их меньше, быстрее
                p_offset = 0
                while True:
                    await _asyncio.sleep(0.2)
                    async with session.post(
                        "https://api-seller.ozon.ru/v1/actions/products",
                        json={"action_id": action_id, "limit": 100, "offset": p_offset},
                        headers=get_active_headers()
                    ) as resp:
                        if resp.status != 200:
                            break
                        pdata = await resp.json()
                        p_items = pdata.get("result", {}).get("products", [])
                        p_total = pdata.get("result", {}).get("total", 0)
                        found_product = find_in_products(p_items, offer_id, product_id_db)
                        if found_product:
                            is_participating = True
                            break
                        if p_offset + 100 >= p_total:
                            break
                        p_offset += 100

                # Если не нашли среди участвующих — ищем среди кандидатов
                if not found_product:
                    c_offset = 0
                    while True:
                        await _asyncio.sleep(0.2)
                        async with session.post(
                            "https://api-seller.ozon.ru/v1/actions/candidates",
                            json={"action_id": action_id, "limit": 100, "offset": c_offset},
                            headers=get_active_headers()
                        ) as resp:
                            if resp.status != 200:
                                break
                            cdata = await resp.json()
                            c_items = cdata.get("result", {}).get("products", [])
                            c_total = cdata.get("result", {}).get("total", 0)
                            found_product = find_in_products(c_items, offer_id, product_id_db)
                            if found_product:
                                break
                            if c_offset + 100 >= c_total:
                                break
                            c_offset += 100

                if not found_product:
                    continue

                p = found_product
                is_elastic = action.get("action_type") == "MARKETPLACE_MULTI_LEVEL_DISCOUNT_ON_AMOUNT"
                entry = {
                    "action_id": action_id,
                    "title": action.get("title"),
                    "action_type": action.get("action_type"),
                    "date_end": action.get("date_end"),
                    "is_participating": is_participating,
                    "product_id": p.get("id"),
                    "current_price": p.get("price", 0),
                    "action_price": p.get("action_price", 0),
                    "max_action_price": p.get("max_action_price", 0),
                    "min_stock": p.get("min_stock", 0),
                    "current_boost": p.get("current_boost"),
                }

                if is_elastic:
                    price = p.get("price", 0)
                    price_min = p.get("price_min_elastic", 0)
                    price_max = p.get("price_max_elastic", 0)
                    min_boost = p.get("min_boost", 0)
                    max_boost = p.get("max_boost", 0)
                    tiers = []
                    if price_min and price_max and price:
                        steps = [
                            (price_min, min_boost, "Минимальный"),
                            (int(price_min - (price_min - price_max) * 0.33), int(min_boost + (max_boost - min_boost) * 0.33), "Средний"),
                            (int(price_min - (price_min - price_max) * 0.66), int(min_boost + (max_boost - min_boost) * 0.66), "Высокий"),
                            (price_max, max_boost, "Максимальный"),
                        ]
                        for tier_price, tier_boost, tier_name in steps:
                            discount_pct = round((price - tier_price) / price * 100, 1) if price else 0
                            tiers.append({
                                "name": tier_name,
                                "price": tier_price,
                                "discount_pct": discount_pct,
                                "boost": tier_boost,
                            })
                    entry["elastic_tiers"] = tiers

                result.append(entry)

            return {"actions": result, "offer_id": offer_id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/fbo-storage/actions/add")
async def api_fbo_storage_actions_add(request: Request, user: dict = Depends(require_any_role)):
    """Добавить товары в акцию."""
    body = await request.json()
    action_id = body.get("action_id")
    products = body.get("products", [])
    if not action_id:
        raise HTTPException(status_code=400, detail="action_id обязателен")
    if not products:
        raise HTTPException(status_code=400, detail="products пустой")

    results = {"added": [], "rejected": [], "errors": []}
    try:
        timeout = aiohttp.ClientTimeout(total=60)
        async with aiohttp.ClientSession(timeout=timeout) as session:
            for i in range(0, len(products), 25):
                batch = products[i:i + 25]
                async with session.post(
                    "https://api-seller.ozon.ru/v1/actions/products/activate",
                    json={"action_id": action_id, "products": batch},
                    headers=get_active_headers()
                ) as resp:
                    if resp.status != 200:
                        text = await resp.text()
                        results["errors"].append(f"batch {i}: {text}")
                        continue
                    data = await resp.json()
                    results["added"].extend(data.get("result", {}).get("product_ids", []))
                    results["rejected"].extend(data.get("result", {}).get("rejected", []))

        return {
            "ok": True,
            "added_count": len(results["added"]),
            "rejected_count": len(results["rejected"]),
            "rejected": results["rejected"],
            "errors": results["errors"],
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))





@app.post("/api/stock/fbs/zero")
async def api_fbs_zero(request: Request, user: dict = Depends(require_admin)):
    """Обнуляет FBS остаток для переданных offer_id через Ozon API"""
    body = await request.json()
    offer_ids = body.get("offer_ids", [])
    if not offer_ids:
        raise HTTPException(status_code=400, detail="offer_ids пустой")

    warehouse_id = int(os.getenv("OZON_WAREHOUSE_ID", "0"))
    # Берём только те offer_id у которых реально есть FBS остаток
    items_with_fbs = {
        item["item_code"]: item["fbs_present"]
        for item in _stock_cache.get("items", [])
        if item["fbs_present"] > 0
    }
    stocks_payload = [
        {"offer_id": oid, "stock": 0, "warehouse_id": warehouse_id}
        for oid in offer_ids
        if oid in items_with_fbs
    ]
    if not stocks_payload:
        return {"ok": True, "total": 0, "errors": [], "skipped": len(offer_ids)}

    async with aiohttp.ClientSession() as session:
        async with session.post(
                "https://api-seller.ozon.ru/v2/products/stocks",
                headers={"Client-Id": _active_account["client_id"], "Api-Key": _active_account["api_key"], "Content-Type": "application/json"},
                json={"stocks": stocks_payload}
        ) as resp:
            data = await resp.json()

    errors = [r for r in data.get("result", []) if not r.get("updated")]
    return {
        "ok": True,
        "total": len(stocks_payload),
        "errors": errors
    }

@app.delete("/api/stock/watchlist/{offer_id}")
async def delete_watchlist_item(offer_id: str, user: dict = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    await db.execute(
        update(FboWatchlist)
        .where(FboWatchlist.offer_id == offer_id)
        .values(dismissed_at=datetime.now())
    )
    await db.commit()
    _stock_cache["watchlist"] = [w for w in _stock_cache.get("watchlist", []) if w["offer_id"] != offer_id]
    return {"ok": True}

@app.delete("/api/stock/watchlist")
async def delete_watchlist_bulk(request: Request, user: dict = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    body = await request.json()
    offer_ids = body.get("offer_ids", [])
    await db.execute(
        update(FboWatchlist)
        .where(FboWatchlist.offer_id.in_(offer_ids))
        .values(dismissed_at=datetime.now())
    )
    await db.commit()
    _stock_cache["watchlist"] = [w for w in _stock_cache.get("watchlist", []) if w["offer_id"] not in offer_ids]
    return {"ok": True}

# --- ЗАПУСК ---
# 📚 УРОК: lifespan — современный способ запускать код при старте/остановке приложения.
# Вместо @app.on_event("startup") используем async context manager.
from contextlib import asynccontextmanager

# =====================================================================
# УПРАВЛЕНИЕ ОСТАТКАМИ (не Сима-Ленд)
# =====================================================================

GOOGLE_SHEET_URL = "https://docs.google.com/spreadsheets/d/1EIWUlFEBbaGDpPK72t9mXDW3wD-m6riKilZmdNcM4lI/export?format=csv&gid=1663121958"

@app.get("/stock-manage", response_class=HTMLResponse)
async def stock_manage_page(request: Request, user: dict = Depends(require_any_role)):
    return templates.TemplateResponse("stock_manage.html", {
        "request": request, "user": user, "active_tab": "stock-manage"
    })

@app.get("/api/stock-manage/items")
async def api_stock_manage_items(
    search: str = "",
    page: int = 1,
    per_page: int = 100,
    user: dict = Depends(require_any_role),
    db: AsyncSession = Depends(get_db)
):
    """Список включённых артикулов с FBO/FBS из кэша и остатками из Google Sheets"""

    # Если кэш пустой — запускаем синхронизацию в фоне
    if not _stock_cache.get("items"):
        asyncio.create_task(sync_stock_cache())

    query = select(StockItem).where(StockItem.enabled == True)
    count_q = select(func.count(StockItem.offer_id)).where(StockItem.enabled == True)
    if search:
        like = f"%{search}%"
        query = query.where(
            or_(StockItem.offer_id.ilike(like), StockItem.name.ilike(like))
        )
        count_q = count_q.where(
            or_(StockItem.offer_id.ilike(like), StockItem.name.ilike(like))
        )
    total = (await db.execute(count_q)).scalar()
    items_r = await db.execute(
        query.order_by(StockItem.offer_id).limit(per_page).offset((page - 1) * per_page)
    )
    items = items_r.scalars().all()

    # FBO/FBS из кэша остатков
    stock_map = {i["item_code"]: i for i in _stock_cache.get("items", [])}

    return {
        "total": total,
        "loading": not bool(_stock_cache.get("items")),
        "items": [
            {
                "offer_id": i.offer_id,
                "name": i.name,
                "fbs_override": i.fbs_override,
                "fbo": stock_map.get(i.offer_id, {}).get("total_free", 0),
                "fbs": stock_map.get(i.offer_id, {}).get("fbs_present", 0),
            }
            for i in items
        ]
    }

@app.get("/api/stock-manage/google")
async def api_stock_manage_google(user: dict = Depends(require_any_role)):
    """Читает остатки из Google Sheets"""
    import urllib.request, csv, io as sio
    try:
        data = urllib.request.urlopen(GOOGLE_SHEET_URL, timeout=10).read().decode('utf-8')
        rows = list(csv.reader(sio.StringIO(data)))
        result = {}
        for row in rows[2:]:
            if not row or not row[0].strip():
                continue
            offer_id = row[0].strip()
            try:
                stock = int(float(row[10].replace(',', '.').replace(' ', ''))) if len(row) > 10 and row[10].strip() else 0
            except:
                stock = 0
            result[offer_id] = stock
        return {"ok": True, "data": result}
    except Exception as e:
        return {"ok": False, "error": str(e), "data": {}}

@app.post("/api/stock-manage/import-from-google")
async def api_stock_manage_import_google(
    user: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    """Импортирует артикулы из Google Sheets в stock_items"""
    import urllib.request, csv, io as sio
    try:
        data = urllib.request.urlopen(GOOGLE_SHEET_URL, timeout=10).read().decode('utf-8')
        rows = list(csv.reader(sio.StringIO(data)))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка чтения Google Sheets: {e}")

    added = 0
    skipped = 0
    for row in rows[2:]:
        if not row or not row[0].strip():
            continue
        offer_id = row[0].strip()
        stmt = pg_insert(StockItem).values(
            offer_id=offer_id,
            name="",
            enabled=True,
        ).on_conflict_do_update(
            index_elements=["offer_id"],
            set_={"enabled": True}
        )
        result = await db.execute(stmt)
        if result.rowcount:
            added += 1
        else:
            skipped += 1

    await db.commit()

    # Подтягиваем названия из таблицы products
    await db.execute(
        update(StockItem)
        .where(StockItem.offer_id == Product.offer_id)
        .values(name=Product.name)
    )
    await db.commit()

    return {"ok": True, "added": added, "skipped": skipped}

@app.post("/api/stock-manage/push-fbs")
async def api_stock_manage_push_fbs(
    request: Request,
    user: dict = Depends(require_admin),
):
    """Передаёт новые FBS остатки в Ozon"""
    body = await request.json()
    stocks = body.get("stocks", [])  # [{offer_id, stock}]
    if not stocks:
        raise HTTPException(status_code=400, detail="stocks пустой")

    warehouse_id = _active_account.get("warehouse_id_uzspace") or int(os.getenv("OZON_WAREHOUSE_ID_UZSPACE", os.getenv("OZON_WAREHOUSE_ID", "0")))
    payload = [{"offer_id": s["offer_id"], "stock": s["stock"], "warehouse_id": warehouse_id} for s in stocks]

    async with aiohttp.ClientSession() as session:
        async with session.post(
            "https://api-seller.ozon.ru/v2/products/stocks",
            headers=get_active_headers(),
            json={"stocks": payload}
        ) as resp:
            data = await resp.json()

    errors = [r for r in data.get("result", []) if not r.get("updated")]
    if errors:
        print(f"PUSH FBS ERRORS: {errors[:3]}", flush=True)
    return {"ok": True, "total": len(payload), "errors": errors}

@app.get("/api/stock-manage/export-list")
async def api_stock_manage_export(user: dict = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    """Скачать Excel с включёнными артикулами из stock_items"""
    import openpyxl
    from fastapi.responses import StreamingResponse
    import io as sio

    items_r = await db.execute(
        select(StockItem).where(StockItem.enabled == True).order_by(StockItem.offer_id)
    )
    items = items_r.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Артикулы"
    ws.append(["Артикул", "Название", "Включить (Да/-)"])
    for item in items:
        ws.append([item.offer_id, item.name or "", "Да"])
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 15

    buf = sio.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=stock_items.xlsx"}
    )

@app.post("/api/stock-manage/sync-stocks")
async def api_stock_manage_sync_stocks(
    user: dict = Depends(require_any_role),
    db: AsyncSession = Depends(get_db)
):
    """Загружает FBO+FBS остатки для артикулов из stock_items"""
    # Получаем все включённые артикулы
    r = await db.execute(select(StockItem.offer_id).where(StockItem.enabled == True))
    offer_ids = [row[0] for row in r.fetchall()]
    if not offer_ids:
        return {"ok": True, "stocks": {}}

    result = {}

    # FBS через /v4/product/info/stocks
    async with aiohttp.ClientSession() as session:
        for i in range(0, len(offer_ids), 100):
            batch = offer_ids[i:i+100]
            async with session.post(
                "https://api-seller.ozon.ru/v4/product/info/stocks",
                headers=get_active_headers(),
                json={"filter": {"offer_id": batch, "visibility": "ALL"}, "limit": 100, "offset": 0}
            ) as resp:
                if resp.status == 200:
                    data = await resp.json()
                    for item in data.get("items", []):
                        oid = item.get("offer_id")
                        if not oid:
                            continue
                        fbs = sum(s.get("present", 0) for s in item.get("stocks", []) if s.get("type") == "fbs")
                        fbo = sum(s.get("present", 0) for s in item.get("stocks", []) if s.get("type") == "fbo")
                        result[oid] = {"fbo": fbo, "fbs": fbs}

    return {"ok": True, "stocks": result}



@app.post("/api/stock-manage/import-list")
async def api_stock_manage_import(
    request: Request,
    user: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    """Загрузить Excel с отмеченными артикулами"""
    import openpyxl, io as sio
    form = await request.form()
    file = form.get("file")
    if not file:
        raise HTTPException(status_code=400, detail="Файл не найден")
    content = await file.read()
    wb = openpyxl.load_workbook(sio.BytesIO(content))
    ws = wb.active
    added = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        offer_id = str(row[0]).strip()
        name = str(row[1]).strip() if row[1] else ""
        enabled = str(row[2]).strip().lower() in ("да", "yes", "+", "1", "true") if len(row) > 2 and row[2] else False
        stmt = pg_insert(StockItem).values(
            offer_id=offer_id,
            name=name,
            enabled=enabled
        ).on_conflict_do_update(
            index_elements=["offer_id"],
            set_={"name": name, "enabled": enabled}
        )
        await db.execute(stmt)
        if enabled:
            added += 1
    await db.commit()
    return {"ok": True, "enabled": added}

@app.patch("/api/stock-manage/fbs-override/{offer_id}")
async def api_stock_manage_fbs_override(
    offer_id: str,
    request: Request,
    user: dict = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    body = await request.json()
    value = body.get("value")
    await db.execute(
        update(StockItem).where(StockItem.offer_id == offer_id).values(fbs_override=value)
    )
    await db.commit()
    return {"ok": True}

# =====================================================================
# КАБИНЕТЫ OZON
# =====================================================================

@app.get("/api/accounts")
async def api_get_accounts(user: dict = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    r = await db.execute(select(OzonAccount).order_by(OzonAccount.id))
    accounts = r.scalars().all()
    return {"accounts": [
        {"id": a.id, "name": a.name, "client_id": a.client_id,
         "warehouse_id_uzspace": a.warehouse_id_uzspace, "is_active": a.is_active}
        for a in accounts
    ]}

@app.post("/api/accounts")
async def api_create_account(request: Request, user: dict = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    body = await request.json()
    account = OzonAccount(
        name=body["name"],
        client_id=body["client_id"],
        api_key=body["api_key"],
        warehouse_id_uzspace=int(body.get("warehouse_id_uzspace") or 0),
        is_active=False
    )
    db.add(account)
    await db.commit()
    return {"ok": True, "id": account.id}

@app.post("/api/accounts/{account_id}/activate")
async def api_activate_account(account_id: int, user: dict = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    global _stock_cache, _ozon_sync_status, _active_account, OZON_HEADERS
    await db.execute(update(OzonAccount).values(is_active=False))
    await db.execute(update(OzonAccount).where(OzonAccount.id == account_id).values(is_active=True))
    await db.commit()
    await load_active_account()
    # Сбрасываем кэш остатков и запускаем синхронизацию для нового кабинета
    _stock_cache = {"items": [], "warehouses": [], "total": 0, "loading": False, "updated_at": None, "watchlist": []}
    _ozon_sync_status = {"running": False, "fetched": 0, "total": 0, "saved": 0, "done": False}
    asyncio.create_task(sync_stock_cache())
    return {"ok": True, "active": _active_account["name"]}

@app.delete("/api/accounts/{account_id}")
async def api_delete_account(account_id: int, user: dict = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    await db.execute(delete(OzonAccount).where(OzonAccount.id == account_id))
    await db.commit()
    return {"ok": True}

@app.get("/api/accounts/current")
async def api_current_account(user: dict = Depends(require_any_role)):
    return {"name": _active_account["name"], "id": _active_account["id"]}

@app.post("/api/accounts/warehouses")
async def api_get_warehouses(request: Request, user: dict = Depends(require_admin)):
    """Получает список FBS складов по переданным учётным данным"""
    body = await request.json()
    client_id = body.get("client_id", "")
    api_key = body.get("api_key", "")
    headers = {"Client-Id": client_id, "Api-Key": api_key, "Content-Type": "application/json"}
    try:
        async with aiohttp.ClientSession() as session:
            async with session.post(
                "https://api-seller.ozon.ru/v2/warehouse/list",
                headers=headers,
                json={}
            ) as resp:
                data = await resp.json()
        warehouses = [
            {"id": w["warehouse_id"], "name": w["name"], "type": w.get("warehouse_type", "fbs")}
            for w in data.get("warehouses", [])
        ]
        return {"ok": True, "warehouses": warehouses}
    except Exception as e:
        return {"ok": False, "error": str(e), "warehouses": []}

# Глобальный стейт обновления кеша WB
_wb_refresh = {"running": False, "fetched": 0, "error": None, "finished_at": None}

@app.get("/content-sync", response_class=HTMLResponse)
async def content_sync_page(request: Request, user: dict = Depends(require_any_role)):
    return templates.TemplateResponse("content_sync.html", {
        "request": request, "user": user, "active_tab": "content-sync"
    })


@app.get("/api/content-sync/products")
async def api_content_sync_products(request: Request):
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Не авторизован")
    wb_key = await get_active_wb_key()
    if not wb_key:
        raise HTTPException(status_code=400, detail="Нет активного WB-аккаунта.")

    # Берём Ozon-данные из локальной БД (быстро) вместо API (медленно)
    async with AsyncSessionLocal() as db:
        r = await db.execute(
            select(Product).where(
                Product.offer_id != None,
                Product.name != None
            )
        )
        products_db = r.scalars().all()

    from content_sync import MatchedProduct, ProductContent
    # Читаем WB из кеша в БД — быстро и стабильно
    async with AsyncSessionLocal() as db:
        r = await db.execute(select(WbProductCache))
        wb_cache = {row.vendor_code: row for row in r.scalars().all()}

    import json as _json
    matched = []
    for p in products_db:
        wc = wb_cache.get(p.offer_id)
        if not wc:
            continue
        matched.append(MatchedProduct(
            vendor_code=p.offer_id,
            ozon=ProductContent(
                vendor_code=p.offer_id,
                product_id=p.product_id,
                name=p.name or "",
                description=p.description or "",
                images=[p.image_url] if p.image_url else [],
                attributes=[],
            ),
            wb=ProductContent(
                vendor_code=wc.vendor_code,
                nm_id=wc.nm_id,
                name=wc.name or "",
                description=wc.description or "",
                images=_json.loads(wc.images_json or "[]"),
                attributes=_json.loads(wc.attributes_json or "[]"),
            ),
        ))

    def serialize(pc):
        if pc is None: return None
        return {
            "vendor_code": pc.vendor_code,
            "product_id":  pc.product_id,
            "nm_id":       pc.nm_id,
            "name":        pc.name,
            "description": pc.description,
            "images":      pc.images[:5],
            "attributes":  pc.attributes[:20],
        }

    return [
        {"vendor_code": m.vendor_code, "ozon": serialize(m.ozon), "wb": serialize(m.wb)}
        for m in matched
    ]

@app.post("/api/content-sync/apply")
async def api_content_sync_apply(request: Request):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=200, content={"error": "unauthorized"})
    body  = await request.json()
    tasks = body.get("tasks", [])
    if not tasks:
        raise HTTPException(status_code=400, detail="Нет задач")
    wb_key = await get_active_wb_key()
    if not wb_key:
        raise HTTPException(status_code=400, detail="Нет активного WB-аккаунта")
    results  = await apply_wb_to_ozon(get_active_headers(), wb_key, tasks)
    ok_count = sum(1 for r in results if r["status"] == "ok")
    return {"results": results, "ok": ok_count, "total": len(results)}

# ── WB-аккаунты ──────────────────────────────────────────────────────
@app.get("/api/wb-accounts")
async def api_wb_accounts_list(request: Request, db: AsyncSession = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=200, content={"error": "unauthorized"})
    accounts = await get_all_wb_accounts_db()
    return [
        {
            "id":           a.id,
            "name":         a.name,
            "is_active":    a.is_active,
            "api_key_hint": a.api_key[:8] + "…" if a.api_key else "",
        }
        for a in accounts
    ]

@app.post("/api/wb-accounts")
async def api_wb_accounts_create(request: Request, db: AsyncSession = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=200, content={"error": "unauthorized"})
    body = await request.json()
    acc  = WbAccount(name=body["name"], api_key=body["api_key"], is_active=False)
    db.add(acc)
    await db.commit()
    return {"ok": True, "id": acc.id, "name": acc.name}

@app.post("/api/wb-accounts/{account_id}/activate")
async def api_wb_accounts_activate(request: Request, account_id: int, db: AsyncSession = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=200, content={"error": "unauthorized"})
    await db.execute(update(WbAccount).values(is_active=False))
    await db.execute(update(WbAccount).where(WbAccount.id == account_id).values(is_active=True))
    await db.commit()
    return {"ok": True, "active_id": account_id}

@app.delete("/api/wb-accounts/{account_id}")
async def api_wb_accounts_delete(request: Request, account_id: int, db: AsyncSession = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return JSONResponse(status_code=200, content={"error": "unauthorized"})
    await db.execute(delete(WbAccount).where(WbAccount.id == account_id))
    await db.commit()
    return {"ok": True}

@app.get("/api/img-proxy")
async def image_proxy(request: Request, url: str):
    """Скачивает фото с WB CDN и отдаёт как JPEG для Ozon."""
    from fastapi.responses import Response
    import io
    from PIL import Image

    if "wbbasket.ru" not in url and "wb.ru" not in url:
        raise HTTPException(status_code=403, detail="Только WB CDN")

    # HEAD-запрос — просто подтверждаем что URL валиден
    if request.method == "HEAD":
        return Response(content=b"", media_type="image/jpeg")

    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(
                url,
                headers={"User-Agent": "Mozilla/5.0"},
                timeout=aiohttp.ClientTimeout(total=15),
            ) as r:
                if r.status != 200:
                    raise HTTPException(status_code=404, detail=f"WB вернул {r.status}")
                raw = await r.read()

        # Конвертируем в JPEG (Ozon принимает только JPG/PNG)
        img = Image.open(io.BytesIO(raw)).convert("RGB")
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=95)
        return Response(content=buf.getvalue(), media_type="image/jpeg")

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=502, detail=str(e))

_ozon_desc_sync = {"running": False, "fetched": 0, "error": None, "finished_at": None}

@app.post("/api/content-sync/sync-ozon-descriptions")
async def sync_ozon_descriptions(request: Request):
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Не авторизован")

    async def _do_sync():
        global _ozon_desc_sync
        _ozon_desc_sync = {"running": True, "fetched": 0, "error": None, "finished_at": None}
        try:
            import aiohttp
            headers = get_active_headers()

            # Только matched товары
            async with AsyncSessionLocal() as db:
                r = await db.execute(
                    select(Product.offer_id).join(
                        WbProductCache, Product.offer_id == WbProductCache.vendor_code
                    ).where(Product.offer_id != None)
                )
                offer_ids = [row[0] for row in r.fetchall()]

            print(f"[ozon-desc] Загружаем описания для {len(offer_ids)} товаров", flush=True)

            DESCRIPTION_ATTR_ID = 4191
            BATCH = 1000

            async with aiohttp.ClientSession() as session:
                for i in range(0, len(offer_ids), BATCH):
                    batch = offer_ids[i:i + BATCH]
                    last_id = ""
                    while True:
                        body = {
                            "filter": {"offer_id": batch},
                            "limit": BATCH,
                            "sort_dir": "ASC",
                        }
                        if last_id:
                            body["last_id"] = last_id

                        async with session.post(
                                "https://api-seller.ozon.ru/v4/product/info/attributes",
                                headers=headers,
                                json=body,
                        ) as resp:
                            data = await resp.json(content_type=None)

                        items = data.get("result", [])
                        if not items:
                            break

                        # Сохраняем описания в БД
                        async with AsyncSessionLocal() as db:
                            for item in items:
                                oid = item.get("offer_id", "")
                                desc = next(
                                    (a["values"][0]["value"]
                                     for a in item.get("attributes", [])
                                     if a.get("id") == DESCRIPTION_ATTR_ID
                                     and a.get("values")),
                                    ""
                                )
                                primary_image = item.get("primary_image", "") or (item.get("images") or [""])[0]
                                update_vals = {}
                                if desc:
                                    update_vals["description"] = desc
                                if primary_image:
                                    update_vals["image_url"] = primary_image
                                if update_vals:
                                    await db.execute(
                                        update(Product)
                                        .where(Product.offer_id == oid)
                                        .values(**update_vals)
                                    )
                            await db.commit()

                        _ozon_desc_sync["fetched"] += len(items)
                        print(f"[ozon-desc] {_ozon_desc_sync['fetched']}/{len(offer_ids)}", flush=True)

                        last_id = data.get("last_id", "")
                        if not last_id or len(items) < BATCH:
                            break

            _ozon_desc_sync["running"] = False
            _ozon_desc_sync["finished_at"] = datetime.now().isoformat()
            print(f"[ozon-desc] ✅ Готово: {_ozon_desc_sync['fetched']}", flush=True)

        except Exception as e:
            import traceback
            _ozon_desc_sync["running"] = False
            _ozon_desc_sync["error"] = str(e)
            print(f"[ozon-desc] ❌ Ошибка: {e}\n{traceback.format_exc()}", flush=True)

    asyncio.create_task(_do_sync())
    return {"status": "started"}


@app.get("/api/content-sync/sync-ozon-descriptions/progress")
async def sync_ozon_descriptions_progress(request: Request):
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Не авторизован")
    return _ozon_desc_sync

# ── Кеш WB ───────────────────────────────────────────────────────────

@app.get("/api/content-sync/wb-cache/status")
async def wb_cache_status(request: Request):
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Не авторизован")
    async with AsyncSessionLocal() as db:
        count = (await db.execute(select(func.count(WbProductCache.vendor_code)))).scalar()
        last  = (await db.execute(
            select(WbProductCache.updated_at).order_by(WbProductCache.updated_at.desc()).limit(1)
        )).scalar()
    return {
        "count":      count or 0,
        "updated_at": last.isoformat() if last else None,
    }


@app.post("/api/content-sync/wb-cache/refresh")
async def wb_cache_refresh(request: Request):
    """Запускает фоновое обновление кеша WB."""
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Не авторизован")
    wb_key = await get_active_wb_key()
    if not wb_key:
        raise HTTPException(status_code=400, detail="Нет активного WB-аккаунта")

    async def _do_refresh():
        global _wb_refresh
        _wb_refresh = {"running": True, "fetched": 0, "error": None, "finished_at": None}
        try:
            import aiohttp, json as _json
            from database import WbProductCache

            headers = {"Authorization": wb_key, "Content-Type": "application/json"}
            cursor: dict = {}

            async with aiohttp.ClientSession() as session:
                while True:
                    # Retry при временных ошибках WB (500/502/503/504)
                    data = None
                    for attempt in range(5):
                        async with session.post(
                            "https://content-api.wildberries.ru/content/v2/get/cards/list",
                            headers=headers,
                            json={"settings": {
                                "sort":   {"ascending": False},
                                "cursor": {**cursor, "limit": 100},
                                "filter": {"withPhoto": -1},
                            }},
                        ) as resp:
                            status = resp.status
                            if status in (500, 502, 503, 504):
                                wait = (attempt + 1) * 15
                                print(f"[wb-cache] WB вернул {status}, попытка {attempt+1}/5, ждём {wait} сек...", flush=True)
                                await asyncio.sleep(wait)
                                continue
                            data = await resp.json(content_type=None)
                            print(f"[wb-cache] HTTP {status} | cards={len(data.get('cards',[]))} | cursor_total={data.get('cursor',{}).get('total')}", flush=True)
                            break

                    if data is None:
                        raise Exception("WB API недоступен после 5 попыток")

                    cards = data.get("cards", [])
                    if not cards:
                        err = data.get("errorText") or data.get("error")
                        if err:
                            print(f"[wb-cache] Ошибка API: {err} | {data.get('additionalErrors')}", flush=True)
                        print("[wb-cache] Пустой ответ — завершаем", flush=True)
                        break

                    batch = []
                    for card in cards:
                        vc     = card.get("vendorCode") or str(card.get("nmID"))
                        photos = card.get("photos", [])
                        images = [p["big"] for p in photos if p.get("big")][:10]
                        attrs  = [
                            {
                                "name":  ch.get("name", ""),
                                "value": " / ".join(
                                    str(v) for v in (
                                        ch["value"] if isinstance(ch.get("value"), list)
                                        else [ch["value"]] if ch.get("value") is not None
                                        else []
                                    )
                                ),
                            }
                            for ch in card.get("characteristics", []) if ch.get("name")
                        ]
                        barcodes = []
                        for size in card.get("sizes", []):
                            barcodes.extend(size.get("skus", []))
                        dims = card.get("dimensions") or {}
                        batch.append(WbProductCache(
                            vendor_code     = vc,
                            nm_id           = card.get("nmID"),
                            name            = card.get("title", ""),
                            description     = card.get("description", ""),
                            brand           = card.get("brand", ""),
                            subject_name    = card.get("subjectName", ""),
                            images_json     = _json.dumps(images,    ensure_ascii=False),
                            attributes_json = _json.dumps(attrs,     ensure_ascii=False),
                            barcodes_json   = _json.dumps(barcodes,  ensure_ascii=False),
                            dimensions_json = _json.dumps(dims,      ensure_ascii=False),
                            updated_at      = datetime.now(),
                        ))

                    async with AsyncSessionLocal() as db:
                        for item in batch:
                            await db.merge(item)
                        await db.commit()

                    _wb_refresh["fetched"] += len(batch)
                    print(f"[wb-cache] Загружено: {_wb_refresh['fetched']}", flush=True)

                    cur = data.get("cursor", {})
                    if cur.get("total", 0) < 100:
                        print(f"[wb-cache] Последняя страница (total={cur.get('total')})", flush=True)
                        break
                    cursor = {"updatedAt": cur["updatedAt"], "nmID": cur["nmID"]}

            _wb_refresh["running"]     = False
            _wb_refresh["finished_at"] = datetime.now().isoformat()
            print(f"[wb-cache] ✅ Готово: {_wb_refresh['fetched']} карточек", flush=True)

        except Exception as e:
            import traceback
            _wb_refresh["running"] = False
            _wb_refresh["error"]   = str(e)
            print(f"[wb-cache] ❌ Ошибка: {e}\n{traceback.format_exc()}", flush=True)

    asyncio.create_task(_do_refresh())
    return {"status": "started", "message": "Обновление запущено в фоне"}


@app.get("/api/content-sync/wb-cache/progress")
async def wb_cache_progress(request: Request):
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Не авторизован")
    return _wb_refresh

@app.get("/api/content-sync/products")
async def api_content_sync_products(request: Request):
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Не авторизован")

    # Ozon из локальной БД
    async with AsyncSessionLocal() as db:
        r = await db.execute(
            select(Product).where(Product.offer_id != None, Product.name != None)
        )
        products_db = r.scalars().all()

    # WB из кеша
    async with AsyncSessionLocal() as db:
        r = await db.execute(select(WbProductCache))
        wb_cache = {row.vendor_code: row for row in r.scalars().all()}

    if not wb_cache:
        raise HTTPException(status_code=400, detail="Кеш WB пуст. Нажмите 'Обновить кеш WB'.")

    from content_sync import MatchedProduct, ProductContent
    matched = []
    for p in products_db:
        wc = wb_cache.get(p.offer_id)
        if not wc:
            continue
        matched.append(MatchedProduct(
            vendor_code = p.offer_id,
            ozon = ProductContent(
                vendor_code = p.offer_id,
                product_id  = p.product_id,
                name        = p.name or "",
                description = "",
                images      = [p.image_url] if getattr(p, 'image_url', None) else [],
                attributes  = [],
            ),
            wb = ProductContent(
                vendor_code = wc.vendor_code,
                nm_id       = wc.nm_id,
                name        = wc.name or "",
                description = wc.description or "",
                images      = _json.loads(wc.images_json     or "[]"),
                attributes  = _json.loads(wc.attributes_json or "[]"),
            ),
        ))

    def serialize(pc):
        if pc is None: return None
        return {
            "vendor_code": pc.vendor_code,
            "product_id":  pc.product_id,
            "nm_id":       pc.nm_id,
            "name":        pc.name,
            "description": pc.description,
            "images":      pc.images[:5],
            "attributes":  pc.attributes[:20],
        }

    return [
        {"vendor_code": m.vendor_code, "ozon": serialize(m.ozon), "wb": serialize(m.wb)}
        for m in matched
    ]

# ── WB → Ozon: перенос карточек ──────────────────────────────────────────────

@app.get("/wb-to-ozon", response_class=HTMLResponse)
async def wb_to_ozon_page(request: Request, user: dict = Depends(require_any_role), db: AsyncSession = Depends(get_db)):
    ozon_accounts = (await db.execute(select(OzonAccount).order_by(OzonAccount.id))).scalars().all()
    wb_accounts   = (await db.execute(select(WbAccount).order_by(WbAccount.id))).scalars().all()
    return templates.TemplateResponse("wb_to_ozon.html", {
        "request": request, "user": user, "active_tab": "wb-to-ozon",
        "ozon_accounts": ozon_accounts,
        "wb_accounts":   wb_accounts,
    })


@app.post("/api/wb-to-ozon/compare")
async def api_wb_to_ozon_compare(request: Request, user: dict = Depends(require_any_role), db: AsyncSession = Depends(get_db)):
    from content_sync import find_wb_missing_from_ozon
    body = await request.json()
    ozon_account_id = int(body["ozon_account_id"])
    wb_account_id   = int(body["wb_account_id"])

    wb_acc = (await db.execute(select(WbAccount).where(WbAccount.id == wb_account_id))).scalar_one_or_none()
    if not wb_acc:
        raise HTTPException(status_code=404, detail="WB-аккаунт не найден")

    result = await find_wb_missing_from_ozon(ozon_account_id, wb_acc.api_key)
    return result


@app.post("/api/wb-to-ozon/create-cards")
async def api_wb_to_ozon_create(request: Request, user: dict = Depends(require_any_role)):
    from content_sync import create_ozon_cards_from_wb
    body = await request.json()
    ozon_account_id = int(body["ozon_account_id"])
    vendor_codes    = list(body["vendor_codes"])
    if not vendor_codes:
        raise HTTPException(status_code=400, detail="Не выбраны артикулы")
    result = await create_ozon_cards_from_wb(ozon_account_id, vendor_codes)
    return result


# ── Заказы Сима ──────────────────────────────────────────────────────────────

@app.get("/sima-orders")
async def sima_orders_page(request: Request, user: dict = Depends(require_any_role)):
    return templates.TemplateResponse("sima_orders.html", {
        "request": request, "user": user, "active_tab": "sima-orders"
    })


@app.post("/api/sima/orders")
async def api_sima_get_orders(request: Request):
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401)
    from ozon_api import get_new_orders
    import pandas as pd, io, base64
    orders = await get_new_orders()
    if not orders:
        return {"orders": [], "articles": "", "excel_b64": None}

    all_articles = []
    for o in orders:
        for p in o["products"]:
            qty = int(p.get("quantity", 1))
            art = str(p.get("offer_id") or p.get("sku"))
            for _ in range(qty):
                all_articles.append(art)

    data_for_excel = []
    for order in orders:
        for product in order["products"]:
            data_for_excel.append({
                "Номер заказа": order["number"],
                "Дата отгрузки": order["ship_date"],
                "Название":      product["name"],
                "Артикул":       product.get("offer_id"),
                "Количество":    product["quantity"],
                "Цена":          product["price"],
            })
    df = pd.DataFrame(data_for_excel)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False)
    buf.seek(0)
    excel_b64 = base64.b64encode(buf.read()).decode()

    return {
        "orders":    orders,
        "articles":  " ".join(all_articles),
        "count":     len(orders),
        "excel_b64": excel_b64,
        "filename":  f"Zakaz_Sima_{datetime.now().strftime('%d_%m')}.xlsx",
    }


@app.post("/api/sima/check-cart")
async def api_sima_check_cart(request: Request, file: UploadFile = File(...)):
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401)
    import pandas as pd, io
    from ozon_api import get_total_ozon_demand
    from database import CostHistory

    content = await file.read()
    try:
        df_raw = pd.read_excel(io.BytesIO(content), header=None)
    except Exception:
        df_raw = pd.read_csv(io.BytesIO(content), header=None, sep=None, engine="python")

    # Извлекаем номер заказа и итоговую сумму из строк до заголовка
    sima_file_order_number = None
    total_sum_from_file = None
    header_row_index = -1
    for i, row in df_raw.iterrows():
        row_content = " ".join([str(v).lower() for v in row if pd.notna(v)])
        if "артикул" in row_content and ("количество" in row_content or "кол-во" in row_content):
            header_row_index = i
            break
        if sima_file_order_number is None:
            for v in row:
                if isinstance(v, (int, float)) and not pd.isna(v) and v > 0:
                    sima_file_order_number = str(int(v))
                    break
        if "итого" in row_content:
            vals = [v for v in row if isinstance(v, (int, float)) and not pd.isna(v)]
            if vals:
                total_sum_from_file = float(vals[-1])

    if header_row_index == -1:
        raise HTTPException(status_code=400, detail="Не найдены колонки Артикул и Количество")

    headers = [str(h).strip().lower() for h in df_raw.iloc[header_row_index]]
    df = df_raw.iloc[header_row_index + 1:].copy()
    df.columns = headers

    col_art   = next((c for c in df.columns if "артикул" in c or "код" in c), None)
    col_qty   = next((c for c in df.columns if "количество" in c or "кол-во" in c or "кол." in c), None)
    col_price = next((c for c in df.columns if "цена, ₽" in c or ("цена" in c and "₽" in c)), None)
    if col_price is None:
        col_price = next((c for c in df.columns if "цена" in c and "опт" not in c), None)

    file_items: dict[str, int] = {}
    for _, row in df.iterrows():
        try:
            art = str(row[col_art]).strip()
            if not art or art == "nan" or "итого" in art.lower():
                continue
            qty = int(float(row[col_qty]))
            file_items[art] = file_items.get(art, 0) + qty
        except Exception:
            continue

    ozon_demand = await get_total_ozon_demand()

    diffs = []
    all_arts = set(list(file_items.keys()) + list(ozon_demand.keys()))
    for art in sorted(all_arts):
        in_file = file_items.get(art, 0)
        needed  = ozon_demand.get(art, 0)
        if in_file > needed:
            diffs.append({"art": art, "in_file": in_file, "needed": needed, "type": "excess",  "delta": in_file - needed})
        elif in_file < needed:
            diffs.append({"art": art, "in_file": in_file, "needed": needed, "type": "missing", "delta": needed - in_file})

    # Обновляем себестоимость
    cost_changes = []
    if col_price and col_art:
        file_prices: dict[str, int] = {}
        for _, row in df.iterrows():
            try:
                art = str(row[col_art]).strip()
                if not art or art == "nan" or "итого" in art.lower():
                    continue
                price_val = float(str(row[col_price]).replace(",", ".").replace(" ", ""))
                if price_val > 0:
                    file_prices[art] = round(price_val)
            except Exception:
                continue

        if file_prices:
            async with AsyncSessionLocal() as db:
                for offer_id, new_price in file_prices.items():
                    result = await db.execute(select(Product).where(Product.offer_id == offer_id))
                    product = result.scalars().first()
                    if product:
                        old_price = product.cost_price or 0
                        if old_price != new_price:
                            cost_changes.append({"offer_id": offer_id, "old": old_price, "new": new_price})
                            product.cost_price = new_price
                            product.updated_at = datetime.now()
                            db.add(CostHistory(
                                offer_id=offer_id,
                                old_cost=old_price if old_price else None,
                                new_cost=new_price,
                                source="sima_order",
                                changed_at=datetime.now(),
                            ))
                await db.commit()

    # Итого по количеству из файла
    total_qty_from_file = sum(file_items.values())

    # Ищем "итого" в строках данных (ниже заголовка), если не нашли выше
    if total_sum_from_file is None:
        col_sum = next((c for c in df.columns if "сумма" in c), None)
        if col_sum:
            for _, row in df.iterrows():
                try:
                    art_val = str(row.get(col_art, "")).lower()
                    if "итого" in art_val or (art_val in ("nan", "") and pd.isna(row.get(col_art))):
                        v = row.get(col_sum)
                        if isinstance(v, (int, float)) and not pd.isna(v):
                            total_sum_from_file = float(v)
                except Exception:
                    pass

    return {
        "diffs": diffs,
        "cost_changes": cost_changes,
        "ok": len(diffs) == 0,
        "file_order_number": sima_file_order_number,
        "file_total_sum": total_sum_from_file,
        "file_total_qty": total_qty_from_file,
    }


@app.post("/api/sima/assemble")
async def api_sima_assemble(request: Request):
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401)
    body = await request.json()
    sima_num    = body.get("sima_num", "").strip()
    supply_date = body.get("supply_date", "").strip()
    file_order_number = body.get("file_order_number")
    file_total_sum    = body.get("file_total_sum")
    file_total_qty    = body.get("file_total_qty")
    if not sima_num or not supply_date:
        raise HTTPException(status_code=400, detail="Укажите номер заказа и дату поставки")
    from ozon_api import assemble_orders
    from database import SimaOrderHistory
    result = await assemble_orders(sima_order_num=sima_num, supply_date=supply_date)

    order_num_to_save = file_order_number or sima_num
    async with AsyncSessionLocal() as db:
        db.add(SimaOrderHistory(
            sima_order_number=order_num_to_save,
            order_date=datetime.now(),
            total_sum=float(file_total_sum) if file_total_sum is not None else None,
            total_qty=int(file_total_qty) if file_total_qty is not None else None,
            delivery_date=supply_date,
        ))
        await db.commit()

    return {"message": result}


@app.get("/sima-history", response_class=HTMLResponse)
async def sima_history_page(request: Request, user: dict = Depends(require_any_role)):
    return templates.TemplateResponse("sima_history.html", {
        "request": request, "user": user, "active_tab": "sima-history"
    })


@app.get("/api/sima/history")
async def api_sima_history(
    q: str = "",
    date_from: str = "",
    date_to: str = "",
    user: dict = Depends(require_any_role),
    db: AsyncSession = Depends(get_db),
):
    from database import SimaOrderHistory
    filters = []
    if q:
        filters.append(SimaOrderHistory.sima_order_number.ilike(f"%{q}%"))
    if date_from:
        try:
            filters.append(SimaOrderHistory.order_date >= datetime.strptime(date_from, "%Y-%m-%d"))
        except ValueError:
            pass
    if date_to:
        try:
            filters.append(SimaOrderHistory.order_date < datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1))
        except ValueError:
            pass

    query = select(SimaOrderHistory)
    if filters:
        query = query.where(*filters)
    result = await db.execute(query.order_by(SimaOrderHistory.order_date.desc()))
    records = result.scalars().all()
    return [
        {
            "id": r.id,
            "sima_order_number": r.sima_order_number,
            "order_date": r.order_date.strftime("%d.%m.%Y") if r.order_date else "",
            "total_sum": r.total_sum,
            "total_qty": r.total_qty,
            "delivery_date": r.delivery_date or "",
        }
        for r in records
    ]


def _parse_sima_history_body(body: dict) -> dict:
    values = {
        "sima_order_number": str(body.get("sima_order_number") or "").strip(),
        "total_sum": float(body["total_sum"]) if body.get("total_sum") not in (None, "") else None,
        "total_qty": int(body["total_qty"]) if body.get("total_qty") not in (None, "") else None,
        "delivery_date": str(body.get("delivery_date") or "").strip() or None,
    }
    order_date_raw = str(body.get("order_date") or "").strip()
    if order_date_raw:
        values["order_date"] = datetime.strptime(order_date_raw, "%d.%m.%Y")
    return values


@app.post("/api/sima/history")
async def api_sima_history_create(request: Request, user: dict = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    from database import SimaOrderHistory
    body = await request.json()
    if not str(body.get("sima_order_number") or "").strip():
        raise HTTPException(400, "Номер счёта обязателен")
    values = _parse_sima_history_body(body)
    record = SimaOrderHistory(**{k: v for k, v in values.items() if v is not None or k == "sima_order_number"})
    db.add(record)
    await db.commit()
    return {"ok": True, "id": record.id}


@app.put("/api/sima/history/{record_id}")
async def api_sima_history_update(record_id: int, request: Request, user: dict = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    from database import SimaOrderHistory
    body = await request.json()
    if not str(body.get("sima_order_number") or "").strip():
        raise HTTPException(400, "Номер счёта обязателен")
    values = _parse_sima_history_body(body)
    result = await db.execute(
        update(SimaOrderHistory).where(SimaOrderHistory.id == record_id).values(**values)
    )
    await db.commit()
    if result.rowcount == 0:
        raise HTTPException(404, "Запись не найдена")
    return {"ok": True}


@app.delete("/api/sima/history/{record_id}")
async def api_sima_history_delete(record_id: int, user: dict = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    from database import SimaOrderHistory
    await db.execute(delete(SimaOrderHistory).where(SimaOrderHistory.id == record_id))
    await db.commit()
    return {"ok": True}


@asynccontextmanager
async def lifespan(app: FastAPI):
    await init_db()
    await load_active_account()
    await load_storage_report_from_db()
    await load_fbo_sales_watch_from_db()
    import asyncio
    asyncio.create_task(sync_from_ozon())

    async def fbo_sales_watch_loop():
        while True:
            await asyncio.sleep(3600)  # каждый час
            try:
                # Сначала обновляем кэш остатков
                await sync_stock_cache()
                # Ждём пока загрузится
                await asyncio.sleep(30)
                # Потом проверяем продажи
                await check_fbo_sales_and_notify()
            except Exception as e:
                print(f"FBO SALES WATCH loop error: {e}", flush=True)

    asyncio.create_task(fbo_sales_watch_loop())
    yield

app.router.lifespan_context = lifespan


# =====================================================================
# ЮНИ-ЭКОНОМИКА
# =====================================================================

@app.get("/unit-economics", response_class=HTMLResponse)
async def unit_economics_page(request: Request, user: dict = Depends(require_any_role)):
    return templates.TemplateResponse("unit_economics.html", {
        "request": request, "user": user, "active_tab": "unit-economics"
    })

@app.get("/api/unit-economics/products")
async def api_unit_economics_products(
    page: int = 1,
    per_page: int = 100,
    search: str = "",
    brand: str = "",
    direction: str = "",
    user: dict = Depends(require_any_role),
    db: AsyncSession = Depends(get_db)
):
    query = select(Product).where(Product.offer_id != None)
    count_q = select(func.count(Product.offer_id)).where(Product.offer_id != None)
    if search:
        like = f"%{search}%"
        query = query.where(or_(Product.offer_id.ilike(like), Product.name.ilike(like)))
        count_q = count_q.where(or_(Product.offer_id.ilike(like), Product.name.ilike(like)))
    if brand:
        query = query.where(Product.brand == brand)
        count_q = count_q.where(Product.brand == brand)
    if direction == "uzspace":
        query = query.where(Product.offer_id.notlike("%/%") == False)
        count_q = count_q.where(Product.offer_id.notlike("%/%") == False)
    elif direction == "sima":
        query = query.where(Product.brand == "Сима-ленд")
        count_q = count_q.where(Product.brand == "Сима-ленд")

    total = (await db.execute(count_q)).scalar()
    products_r = await db.execute(
        query.order_by(Product.offer_id).limit(per_page).offset((page-1)*per_page)
    )
    products = products_r.scalars().all()

    return {
        "total": total,
        "products": [
            {
                "offer_id": p.offer_id,
                "product_id": p.product_id,
                "name": p.name or "",
                "brand": p.brand or "",
                "image_url": p.image_url or "",
                "price": p.price or 0,
                "old_price": p.old_price or 0,
                "min_price": p.min_price or 0,
                "cost_price": p.cost_price or 0,
                "ff_cost": p.ff_cost or 0,
                "commission_fbs_percent": p.commission_fbs_percent or 0,
                "commission_fbs_logistics": p.commission_fbs_logistics or 0,
                "volume_liters": p.volume_liters or 0,
                "fbs_return_amount": p.fbs_return_amount or 0,
                "category_name": p.category_name or "",
            }
            for p in products
        ]
    }

@app.post("/api/unit-economics/update-prices")
async def api_unit_economics_update_prices(
    request: Request,
    user: dict = Depends(require_admin),
):
    body = await request.json()
    prices = body.get("prices", [])
    if not prices:
        raise HTTPException(status_code=400, detail="prices пустой")

    # Разбиваем на батчи по 1000
    results = []
    async with aiohttp.ClientSession() as session:
        for i in range(0, len(product_ids), 100):
            batch = product_ids[i:i + 100]
            _ue_sync_status["progress"] = f"Габариты: {i}/{len(product_ids)}..."
            try:
                async with session.post(
                        "https://api-seller.ozon.ru/v3/product/info/list",
                        headers=get_active_headers(),
                        json={"product_id": batch}
                ) as resp:
                    data = await resp.json()
                for item in data.get("items", []):
                    oid = item.get("offer_id")
                    if not oid:
                        continue
                    volume_weight = float(item.get("volume_weight") or 0)
                    # volume_weight в кг → в литры (1 кг объёмного веса = 1 л по тарифу Ozon)
                    volume_liters = volume_weight

                    # Логистика из commissions
                    fbs_delivery = 0
                    fbs_return = 0
                    for comm in (item.get("commissions") or []):
                        if comm.get("sale_schema") == "FBS":
                            fbs_delivery = float(comm.get("delivery_amount") or 0)
                            fbs_return = float(comm.get("return_amount") or 0)
                            break

                    volume_map[oid] = {
                        "volume_liters": round(volume_weight * 5, 3),
                        "weight_kg": volume_weight,
                        "fbs_delivery": fbs_delivery,
                        "fbs_return": fbs_return,
                    }
            except Exception as e:
                print(f"volume batch error: {e}", flush=True)

    errors = [r for r in results if not r.get("updated")]
    return {"ok": True, "total": len(prices), "errors": len(errors), "error_items": errors[:5]}

_ue_sync_status = {"running": False, "progress": "", "synced": 0, "error": ""}


@app.get("/api/unit-economics/filters")
async def api_ue_filters(
    user: dict = Depends(require_any_role),
    db: AsyncSession = Depends(get_db)
):
    from sqlalchemy import select, distinct
    brands_r = await db.execute(
        select(distinct(Product.brand)).where(Product.brand != None, Product.brand != '')
    )
    cats_r = await db.execute(
        select(distinct(Product.category_name)).where(Product.category_name != None, Product.category_name != '')
    )
    brands = sorted([r[0] for r in brands_r if r[0]])
    categories = sorted([r[0] for r in cats_r if r[0]])
    return {"brands": brands, "categories": categories}

@app.get("/api/unit-economics/sync/status")
async def api_ue_sync_status(user: dict = Depends(require_any_role)):
    return _ue_sync_status

@app.post("/api/unit-economics/sync")
async def api_ue_sync(user: dict = Depends(require_admin)):
    asyncio.create_task(_ue_sync_task())
    return {"ok": True}

async def _ue_sync_task():
    global _ue_sync_status
    _ue_sync_status = {"running": True, "progress": "Загружаем артикулы...", "synced": 0, "error": ""}
    try:
        # ШАГ 1: Все артикулы через last_id
        all_items = []
        last_id = ""
        while True:
            async with aiohttp.ClientSession() as session:
                async with session.post(
                    "https://api-seller.ozon.ru/v3/product/list",
                    headers=get_active_headers(),
                    json={"filter": {"visibility": "ALL"}, "last_id": last_id, "limit": 1000}
                ) as resp:
                    data = await resp.json()
            result = data.get("result", {})
            items = result.get("items", [])
            if not items:
                break
            all_items.extend(items)
            last_id = result.get("last_id", "")
            _ue_sync_status["progress"] = f"Артикулы: {len(all_items)}..."
            if not last_id or len(items) < 1000:
                break

        product_ids = [item.get("product_id") for item in all_items if item.get("product_id")]

        # ШАГ 2: Цены и комиссии
        price_map = {}
        async with aiohttp.ClientSession() as session:
            for i in range(0, len(product_ids), 1000):
                batch = product_ids[i:i+1000]
                _ue_sync_status["progress"] = f"Цены: {i}/{len(product_ids)}..."
                async with session.post(
                    "https://api-seller.ozon.ru/v5/product/info/prices",
                    headers=get_active_headers(),
                    json={"filter": {"product_id": batch, "visibility": "ALL"}, "limit": 1000, "last_id": ""}
                ) as resp:
                    data = await resp.json()
                for item in data.get("items", []):
                    oid = item.get("offer_id")
                    if not oid:
                        continue
                    price_data = item.get("price") or {}
                    comm = item.get("commissions") or {}
                    price_map[oid] = {
                        "price": int(float(price_data.get("price") or 0)),
                        "old_price": int(float(price_data.get("old_price") or 0)),
                        "min_price": int(float(price_data.get("min_price") or 0)),
                        "commission_fbs_percent": int(float(comm.get("sales_percent_fbs") or 0)),
                        "commission_fbs_logistics": int(
                            float(comm.get("fbs_direct_flow_trans_max_amount") or 0) +
                            float(comm.get("fbs_deliv_to_customer_amount") or 0)),
                    }

        # ШАГ 3: Объём и габариты через /v3/product/info/list
        volume_map = {}
        async with aiohttp.ClientSession() as session:
            for i in range(0, len(product_ids), 100):
                batch = product_ids[i:i + 100]
                _ue_sync_status["progress"] = f"Габариты: {i}/{len(product_ids)}..."
                try:
                    async with session.post(
                            "https://api-seller.ozon.ru/v3/product/info/list",
                            headers=get_active_headers(),
                            json={"product_id": batch}
                    ) as resp:
                        data = await resp.json()
                    for item in data.get("items", []):
                        oid = item.get("offer_id")
                        if not oid:
                            continue
                        volume_weight = float(item.get("volume_weight") or 0)
                        fbs_delivery = 0
                        fbs_return = 0
                        for comm in (item.get("commissions") or []):
                            if comm.get("sale_schema") == "FBS":
                                fbs_delivery = float(comm.get("delivery_amount") or 0)
                                fbs_return = float(comm.get("return_amount") or 0)
                                break
                        volume_map[oid] = {
                            "volume_liters": round(volume_weight * 5, 3),
                            "weight_kg": volume_weight,
                            "fbs_delivery": fbs_delivery,
                            "fbs_return": fbs_return,
                        }
                except Exception as e:
                    print(f"volume batch error: {e}", flush=True)

        # ШАГ 4: Сохраняем в БД
        _ue_sync_status["progress"] = f"Сохраняем {len(all_items)} записей..."
        saved = 0
        async with AsyncSessionLocal() as db:
            for i in range(0, len(all_items), 500):
                batch = all_items[i:i+500]
                for item in batch:
                    oid = str(item.get("offer_id", "")).strip()
                    if not oid:
                        continue
                    pd = price_map.get(oid, {})
                    vd = volume_map.get(oid, {})
                    stmt = pg_insert(Product).values(
                        offer_id=oid,
                        product_id=item.get("product_id"),
                        price=pd.get("price") or None,
                        old_price=pd.get("old_price") or None,
                        min_price=pd.get("min_price") or None,
                        commission_fbs_percent=pd.get("commission_fbs_percent") or None,
                        commission_fbs_logistics=int(vd.get("fbs_delivery") or pd.get("commission_fbs_logistics") or 0),
                        volume_liters=vd.get("volume_liters") or None,
                        weight_kg=vd.get("weight_kg") or None,
                    ).on_conflict_do_update(
                        index_elements=["offer_id"],
                        set_={
                            "price": pd.get("price") or None,
                            "old_price": pd.get("old_price") or None,
                            "min_price": pd.get("min_price") or None,
                            "commission_fbs_percent": pd.get("commission_fbs_percent") or None,
                            "commission_fbs_logistics": int(vd.get("fbs_delivery") or pd.get("commission_fbs_logistics") or 0),
                            "volume_liters": vd.get("volume_liters") or None,
                            "weight_kg": vd.get("weight_kg") or None,
                        }
                    )
                    await db.execute(stmt)
                    saved += 1
                await db.commit()
                _ue_sync_status["synced"] = saved
        _ue_sync_status = {"running": False, "progress": "", "synced": saved, "error": ""}
        print(f"UE SYNC DONE: {saved}", flush=True)
    except Exception as e:
        import traceback
        _ue_sync_status = {"running": False, "progress": "", "synced": 0, "error": str(e)}
        print(f"UE SYNC ERROR: {traceback.format_exc()}", flush=True)

@app.get("/api/active-accounts")
async def api_active_accounts(request: Request):
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401)
    ozon_name = None
    wb_name   = None
    async with AsyncSessionLocal() as db:
        r = await db.execute(select(OzonAccount).where(OzonAccount.is_active == True))
        ozon = r.scalar_one_or_none()
        if ozon: ozon_name = ozon.name
        r = await db.execute(select(WbAccount).where(WbAccount.is_active == True))
        wb = r.scalar_one_or_none()
        if wb: wb_name = wb.name
    return {"ozon": ozon_name, "wb": wb_name}

if __name__ == "__main__":
    import uvicorn
    # reload=True — при изменении файлов сервер перезапускается автоматически.
    # Удобно для разработки, НЕ использовать в production!
    uvicorn.run("web_app:app", host="0.0.0.0", port=8000, reload=True)